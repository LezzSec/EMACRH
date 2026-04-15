# -*- coding: utf-8 -*-
"""
personnel_export_excel — export Excel du profil d'un opérateur.

Aucun import PyQt5. Entrées : données brutes. Sortie : fichier .xlsx.
"""

from infrastructure.logging.logging_config import get_logger
logger = get_logger(__name__)


def export_personnel_profile_excel(
    operateur_id: int,
    nom: str,
    prenom: str,
    matricule: str,
    statut: str,
    infos_data: list,
    polyvalences: list,
    file_path: str,
) -> None:
    """
    Génère un fichier Excel du profil d'un opérateur.

    Args:
        operateur_id:  Identifiant (conservé pour extension future).
        nom:           Nom de famille.
        prenom:        Prénom.
        matricule:     Matricule.
        statut:        Statut (ACTIF / INACTIF).
        infos_data:    Liste de tuples (section_title, [(label, value), ...]).
        polyvalences:  Liste de dicts {poste, niveau, date_evaluation, prochaine_evaluation,
                       anciennete, statut_eval} — chaînes déjà formatées.
        file_path:     Chemin complet du fichier de sortie (.xlsx).

    Raises:
        Exception: en cas d'erreur openpyxl, propagée à l'appelant.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ----- Styles -----
    THIN      = Side(style="thin", color="DDDDDD")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FILL  = PatternFill("solid", fgColor="4472C4")
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    SEC_FILL  = PatternFill("solid", fgColor="EEF2FF")
    TITLE_FONT = Font(bold=True, size=16)
    SUB_FONT  = Font(bold=True, size=12, color="555555")
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left", vertical="center")

    def style_header(row):
        for c in row:
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = CENTER
            c.border = BORDER

    def style_table(ws, start_row, start_col, end_row, end_col, zebra=False):
        for r in range(start_row, end_row + 1):
            fill = (
                PatternFill("solid", fgColor="F8FAFC")
                if zebra and r > start_row and (r - start_row) % 2 == 1
                else None
            )
            for c in range(start_col, end_col + 1):
                cell = ws.cell(r, c)
                cell.border = BORDER
                if fill:
                    cell.fill = fill

    def autofit(ws, min_w=10, max_w=80):
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            longest = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                v = row[0].value
                if v is not None:
                    longest = max(longest, len(str(v)))
            ws.column_dimensions[letter].width = max(min(longest + 2, max_w), min_w)

    wb = Workbook()

    # =========================================================
    # 1) FEUILLE RÉSUMÉ
    # =========================================================
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.sheet_view.zoomScale = 120

    wide_cols = [chr(c) for c in range(ord('A'), ord('S') + 1)]
    for col in wide_cols:
        ws1.column_dimensions[col].width = 24

    ws1.merge_cells("A1:S1")
    ws1["A1"] = f"Profil — {nom or ''} {prenom or ''}".strip()
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = LEFT

    ws1.merge_cells("A2:S2")
    ws1["A2"] = f"Matricule : {matricule or '-'}    |    Statut : {statut or '-'}"
    ws1["A2"].font = SUB_FONT
    ws1["A2"].alignment = LEFT

    ws1["A4"] = "Informations de base"
    ws1["A4"].font = Font(bold=True)
    ws1["A4"].fill = SEC_FILL
    ws1["A4"].alignment = LEFT

    ws1.freeze_panes = "A5"

    # =========================================================
    # 2) FEUILLE INFORMATIONS
    # =========================================================
    ws2 = wb.create_sheet("Informations")
    ws2.sheet_view.zoomScale = 120
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Informations complémentaires"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].alignment = LEFT

    ws2.append(["Catégorie", "Valeur"])
    style_header(ws2[2])

    if infos_data:
        for section_title, items in infos_data:
            ws2.append([section_title.upper(), ""])
            for label, value in items:
                if "Aucune" not in label:
                    ws2.append([f"  {label}", str(value)])

    style_table(ws2, 3, 1, ws2.max_row, 2, zebra=True)
    for cell in ws2["B"][2:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    autofit(ws2, min_w=18, max_w=80)
    ws2.freeze_panes = "A3"

    # =========================================================
    # 3) FEUILLE POLYVALENCES
    # =========================================================
    ws3 = wb.create_sheet("Polyvalences")
    ws3.sheet_view.zoomScale = 120
    ws3.merge_cells("A1:F1")
    ws3["A1"] = "Polyvalences & Évaluations"
    ws3["A1"].font = TITLE_FONT
    ws3["A1"].alignment = LEFT

    headers = ["Poste", "Niveau", "Date Évaluation", "Prochaine Éval.", "Ancienneté", "Statut"]
    ws3.append(headers)
    style_header(ws3[2])

    for p in polyvalences:
        ws3.append([
            str(p.get("poste", "")),
            str(p.get("niveau", "")),
            str(p.get("date_evaluation", "")),
            str(p.get("prochaine_evaluation", "")),
            str(p.get("anciennete", "")),
            str(p.get("statut_eval", "")),
        ])

    style_table(ws3, 3, 1, ws3.max_row, 6, zebra=True)
    for row in ws3.iter_rows(min_row=3, min_col=1, max_col=6, max_row=ws3.max_row):
        for c in row:
            c.alignment = CENTER
    for c in ws3["A"][2:]:
        c.alignment = Alignment(horizontal="left", vertical="center")
    autofit(ws3, min_w=12, max_w=28)
    ws3.freeze_panes = "A3"

    wb.save(file_path)
