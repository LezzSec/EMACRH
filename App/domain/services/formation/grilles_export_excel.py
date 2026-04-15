# -*- coding: utf-8 -*-
"""
Service d'export Excel pour la grille de polyvalence.
Aucune dépendance PyQt5 — pur Python / openpyxl.
"""

from typing import List, Optional

from infrastructure.logging.logging_config import get_logger

logger = get_logger(__name__)


def export_grille_excel(
    file_path: str,
    headers: List[str],
    row_labels: List[str],
    data: List[List[str]],
) -> None:
    """Exporte la grille dans un fichier Excel avec légende des niveaux.

    Args:
        file_path: Chemin complet du fichier .xlsx de destination.
        headers: Liste des en-têtes de colonnes (codes postes).
        row_labels: Liste des libellés de lignes (noms opérateurs + lignes synthèse).
        data: Matrice de données [ligne][colonne] en chaînes.

    Raises:
        Exception: toute erreur openpyxl est laissée remonter à l'appelant.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Grille Polyvalence"

    # En-tête colonnes
    ws.cell(row=1, column=1, value="Nom").font = Font(bold=True)
    for c, h in enumerate(headers, start=2):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)

    # Données
    for r, row_vals in enumerate(data, start=2):
        label = row_labels[r - 2] if (r - 2) < len(row_labels) else ""
        ws.cell(row=r, column=1, value=label)
        for c, val in enumerate(row_vals, start=2):
            ws.cell(row=r, column=c, value=val)

    # Mise en forme tableau
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for r in range(1, len(data) + 2):
        for c in range(1, len(headers) + 2):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin

    # Largeurs colonnes
    ws.column_dimensions[get_column_letter(1)].width = 28
    for c in range(2, len(headers) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 6

    # Légende "Niveaux" sous le tableau
    start_legend_row = len(data) + 3
    bullet = "\u2022"

    ws.cell(row=start_legend_row, column=1, value="Niveaux :").font = Font(bold=True)
    ws.merge_cells(
        start_row=start_legend_row, start_column=1,
        end_row=start_legend_row, end_column=len(headers) + 1,
    )

    legend_lines = [
        f"{bullet} Niveau 1 : Opérateur nouveau à encadrer par titulaire N3/4 ou absent du poste depuis plus de 12 mois. (< 80%)",
        f"{bullet} Niveau 2 : Opérateur formé et apte à conduire le poste seul. Certaines notions sont en cours d'acquisition. (> 80%)",
        f"{bullet} Niveau 3 : Opérateur titulaire, formé, apte à conduire le poste et apte à former. (> 90%)",
        f"{bullet} Niveau 4 : N3 + Leader ou Polyvalent (maîtrise 3 postes d'une ligne : mél. interne, cylindres, conditionnement) (> 90%)",
    ]

    for i, line in enumerate(legend_lines, start=1):
        r = start_legend_row + i
        ws.cell(row=r, column=1, value=line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers) + 1)
        cell = ws.cell(row=r, column=1)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r in range(start_legend_row, start_legend_row + len(legend_lines) + 1):
        ws.row_dimensions[r].height = 18

    # Zone d'impression
    last_row = start_legend_row + len(legend_lines)
    last_col_letter = get_column_letter(len(headers) + 1)
    ws.print_area = f"A1:{last_col_letter}{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    wb.save(file_path)
    logger.info(f"Export Excel grille polyvalence : {file_path}")
