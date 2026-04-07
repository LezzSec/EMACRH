# -*- coding: utf-8 -*-
"""
Service de génération documentaire pour les formations.

Génère les documents officiels pré-remplis à partir des données saisies dans l'application :
  1. Demande de formation (EQ 07 30)
  2. Feuille d'émargement / Attestation de présence (EQ 07 17)
  3. Fiche de suivi de formation (EQ 07 02 01)

Usage:
    from core.services.formation_export_service import FormationExportService

    success, msg, path = FormationExportService.generate_dossier_formation(formation_data)
    # path = chemin vers le fichier xlsx généré
"""

import os
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

from infrastructure.logging.logging_config import get_logger

logger = get_logger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl non disponible - génération xlsx impossible")


# Répertoire de sortie par défaut
_EXPORTS_DIR = Path(__file__).parent.parent.parent / "exports" / "formations"

# Couleurs
_BLUE_HEADER = "1F4E79"
_LIGHT_BLUE = "BDD7EE"
_GREY_LABEL = "F2F2F2"
_BORDER_COLOR = "4472C4"


def _thin_border() -> Border:
    side = Side(style='thin', color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font(size: int = 11) -> Font:
    return Font(name="Calibri", bold=True, size=size, color="FFFFFF")


def _label_font() -> Font:
    return Font(name="Calibri", bold=True, size=10)


def _value_font() -> Font:
    return Font(name="Calibri", size=10)


def _title_font(size: int = 14) -> Font:
    return Font(name="Calibri", bold=True, size=size, color=_BLUE_HEADER)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_BLUE_HEADER)


def _label_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_GREY_LABEL)


def _accent_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_LIGHT_BLUE)


def _set_cell(ws, coordinate: str, value, font=None, fill=None,
              alignment=None, border=None, number_format: str = None):
    cell = ws[coordinate]
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _fmt_date(d) -> str:
    if d is None:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")
    return str(d)


def _fmt_duree(heures) -> str:
    if not heures:
        return ""
    h = float(heures)
    jours = h / 7
    if h == int(h):
        return f"{int(h)} h ({jours:.1f} j)"
    return f"{h:.1f} h ({jours:.1f} j)"


# ============================  FEUILLE 1 : DEMANDE  ============================

def _build_demande(wb: "openpyxl.Workbook", data: Dict):
    ws = wb.create_sheet("Demande de formation (EQ 07 30)")

    # Largeurs de colonnes
    col_widths = [1, 1, 18, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Hauteurs de lignes par défaut
    for r in range(1, 50):
        ws.row_dimensions[r].height = 13

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_al = Alignment(horizontal="right", vertical="center")

    border = _thin_border()

    # --- Titre ---
    ws.merge_cells("C1:N2")
    _set_cell(ws, "C1", "DEMANDE DE FORMATION",
              font=_title_font(16), alignment=center)
    ws.merge_cells("O1:Q2")
    _set_cell(ws, "O1", "EQ 07 30 rev1",
              font=_value_font(), alignment=right_al)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 4

    # --- Ligne vide ---
    ws.row_dimensions[3].height = 2

    # --- Salarié ---
    ws.merge_cells("C4:Q4")
    _set_cell(ws, "C4", "INFORMATIONS SALARIÉ",
              font=_header_font(11), fill=_header_fill(), alignment=center)

    # Nom
    ws.merge_cells("C5:D5")
    _set_cell(ws, "C5", "Nom :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("E5:G5")
    _set_cell(ws, "E5", data.get('nom', '').upper(),
              font=_value_font(), alignment=left, border=border)

    # Prénom
    ws.merge_cells("H5:I5")
    _set_cell(ws, "H5", "Prénom :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("J5:L5")
    _set_cell(ws, "J5", data.get('prenom', ''),
              font=_value_font(), alignment=left, border=border)

    # Service
    ws.merge_cells("M5:N5")
    _set_cell(ws, "M5", "Service :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("O5:Q5")
    _set_cell(ws, "O5", data.get('service', ''),
              font=_value_font(), alignment=left, border=border)

    ws.row_dimensions[5].height = 18

    # --- Formation ---
    ws.row_dimensions[6].height = 2

    ws.merge_cells("C7:Q7")
    _set_cell(ws, "C7", "INFORMATIONS FORMATION",
              font=_header_font(11), fill=_header_fill(), alignment=center)

    # Intitulé + Type
    ws.merge_cells("C8:D8")
    _set_cell(ws, "C8", "Intitulé :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("E8:L8")
    _set_cell(ws, "E8", data.get('intitule', ''),
              font=_value_font(), alignment=left, border=border)
    ws.merge_cells("M8:N8")
    _set_cell(ws, "M8", "Type :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    type_str = data.get('type_formation', '')
    code = data.get('code_formation')
    libelle_tranche = data.get('libelle_tranche', '')
    if code and libelle_tranche:
        type_str = f"{code} – {libelle_tranche}"
    elif code:
        type_str = str(code)
    ws.merge_cells("O8:Q8")
    _set_cell(ws, "O8", type_str,
              font=_value_font(), alignment=left, border=border)
    ws.row_dimensions[8].height = 18

    # Objectif
    ws.merge_cells("C9:D10")
    _set_cell(ws, "C9", "Objectif :", font=_label_font(), fill=_label_fill(),
              alignment=Alignment(horizontal="left", vertical="top"))
    ws.merge_cells("E9:Q10")
    _set_cell(ws, "E9", data.get('objectif', ''),
              font=_value_font(),
              alignment=Alignment(horizontal="left", vertical="top", wrap_text=True))
    ws.row_dimensions[9].height = 16
    ws.row_dimensions[10].height = 16

    # Organisme + Lieu  (ligne 11)
    ws.merge_cells("C11:D11")
    _set_cell(ws, "C11", "Organisme :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("E11:H11")
    _set_cell(ws, "E11", data.get('organisme', ''),
              font=_value_font(), alignment=left, border=border)

    ws.merge_cells("I11:J11")
    _set_cell(ws, "I11", "Lieu :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("K11:Q11")
    _set_cell(ws, "K11", data.get('lieu', ''),
              font=_value_font(), alignment=left, border=border)

    ws.row_dimensions[11].height = 18

    # Coûts (ligne 12) : pédagogique + salarial calculé depuis le taux horaire
    ws.merge_cells("C12:D12")
    _set_cell(ws, "C12", "Coût péda. :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    cout = data.get('cout')
    ws.merge_cells("E12:G12")
    _set_cell(ws, "E12",
              f"{float(cout):.2f} €" if cout else "",
              font=_value_font(), alignment=left, border=border)

    ws.merge_cells("H12:J12")
    _set_cell(ws, "H12", "Coût salarial :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    cout_sal = data.get('cout_salarial')
    taux = data.get('taux_horaire')
    cout_sal_str = ""
    if cout_sal is not None:
        cout_sal_str = f"{float(cout_sal):.2f} €"
        if taux:
            cout_sal_str += f"  ({float(taux):.2f} €/h)"
    ws.merge_cells("K12:Q12")
    _set_cell(ws, "K12", cout_sal_str,
              font=_value_font(), alignment=left, border=border)

    ws.row_dimensions[12].height = 18

    # Dates + Durée (ligne 13)
    ws.merge_cells("C13:D13")
    _set_cell(ws, "C13", "Date début :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("E13:F13")
    _set_cell(ws, "E13", _fmt_date(data.get('date_debut')),
              font=_value_font(), alignment=center, border=border)

    ws.merge_cells("G13:H13")
    _set_cell(ws, "G13", "Date fin :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("I13:J13")
    _set_cell(ws, "I13", _fmt_date(data.get('date_fin')),
              font=_value_font(), alignment=center, border=border)

    ws.merge_cells("K13:L13")
    _set_cell(ws, "K13", "Durée :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("M13:Q13")
    _set_cell(ws, "M13", _fmt_duree(data.get('duree_heures')),
              font=_value_font(), alignment=left, border=border)

    ws.row_dimensions[13].height = 18

    # Formateur (ligne 14)
    ws.merge_cells("C14:D14")
    _set_cell(ws, "C14", "Formateur :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("E14:Q14")
    _set_cell(ws, "E14", data.get('formateur', ''),
              font=_value_font(), alignment=left, border=border)
    ws.row_dimensions[14].height = 18

    # Commentaire / Observations (lignes 15-16)
    ws.merge_cells("C15:D16")
    _set_cell(ws, "C15", "Observations :", font=_label_font(), fill=_label_fill(),
              alignment=Alignment(horizontal="left", vertical="top"))
    ws.merge_cells("E15:Q16")
    _set_cell(ws, "E15", data.get('commentaire', ''),
              font=_value_font(),
              alignment=Alignment(horizontal="left", vertical="top", wrap_text=True))
    ws.row_dimensions[15].height = 16
    ws.row_dimensions[16].height = 16

    # --- Signatures ---
    ws.row_dimensions[17].height = 2
    ws.merge_cells("C18:Q18")
    _set_cell(ws, "C18", "SIGNATURES",
              font=_header_font(11), fill=_header_fill(), alignment=center)

    ws.row_dimensions[19].height = 16
    headers_sig = [("C19:F19", "Demandeur"), ("G19:J19", "Date"),
                   ("K19:N19", "Responsable direct"), ("O19:Q19", "Visa")]
    for rng, label in headers_sig:
        ws.merge_cells(rng)
        _set_cell(ws, rng.split(":")[0], label,
                  font=_label_font(), fill=_label_fill(), alignment=center, border=border)

    ws.merge_cells("C20:F22")
    ws.merge_cells("G20:J22")
    ws.merge_cells("K20:N22")
    ws.merge_cells("O20:Q22")
    for cell_coord in ["C20", "G20", "K20", "O20"]:
        _set_cell(ws, cell_coord, "",
                  font=_value_font(), border=border,
                  alignment=Alignment(horizontal="center", vertical="center"))
    for r in [20, 21, 22]:
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[23].height = 4
    ws.merge_cells("C24:F24")
    _set_cell(ws, "C24", "Le demandeur",
              font=_label_font(), alignment=center)
    ws.merge_cells("K24:N24")
    _set_cell(ws, "K24", "Le responsable hiérarchique",
              font=_label_font(), alignment=center)

    # Note bas de page
    ws.row_dimensions[25].height = 4
    ws.merge_cells("B26:Q26")
    _set_cell(ws, "B26",
              "Cette demande devra être présentée au responsable hiérarchique "
              "et à la direction des Ressources Humaines.",
              font=Font(name="Calibri", size=9, italic=True),
              alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

    # Impression : 1 page
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:Q26"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


# ========================  FEUILLE 2 : ÉMARGEMENT  ========================

def _build_emargement(wb: "openpyxl.Workbook", data: Dict):
    ws = wb.create_sheet("Feuille d'émargement (EQ 07 17)")

    col_widths = [2, 4, 30, 16, 16, 16, 16, 16, 16, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = _thin_border()

    # --- Titre ---
    ws.merge_cells("C1:I2")
    _set_cell(ws, "C1", "FEUILLE D'ÉMARGEMENT / ATTESTATION DE PRÉSENCE",
              font=_title_font(14), alignment=center)
    ws.merge_cells("C3:I3")
    _set_cell(ws, "C3", "EQ 07 17 rév.2",
              font=_value_font(),
              alignment=Alignment(horizontal="right", vertical="center"))
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # --- Infos formation ---
    ws.row_dimensions[4].height = 8

    ws.merge_cells("C5:D5")
    _set_cell(ws, "C5", "Intitulé :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("E5:I5")
    _set_cell(ws, "E5", data.get('intitule', ''),
              font=_value_font(), alignment=left, border=border)
    ws.row_dimensions[5].height = 22

    ws.merge_cells("C6:D6")
    _set_cell(ws, "C6", "Lieu :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("E6:F6")
    _set_cell(ws, "E6", data.get('lieu', ''),
              font=_value_font(), alignment=left, border=border)

    ws.merge_cells("G6:H6")
    _set_cell(ws, "G6", "Durée :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    _set_cell(ws, "I6", _fmt_duree(data.get('duree_heures')),
              font=_value_font(), alignment=left, border=border)
    ws.row_dimensions[6].height = 22

    ws.merge_cells("C7:D7")
    _set_cell(ws, "C7", "Date(s) :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    dates_str = _fmt_date(data.get('date_debut'))
    if data.get('date_fin') and data.get('date_fin') != data.get('date_debut'):
        dates_str += f"  au  {_fmt_date(data.get('date_fin'))}"
    ws.merge_cells("E7:F7")
    _set_cell(ws, "E7", dates_str,
              font=_value_font(), alignment=left, border=border)

    ws.merge_cells("G7:H7")
    _set_cell(ws, "G7", "Organisme :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    _set_cell(ws, "I7", data.get('organisme', ''),
              font=_value_font(), alignment=left, border=border)
    ws.row_dimensions[7].height = 22

    ws.merge_cells("C8:D8")
    _set_cell(ws, "C8", "Formateur :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("E8:I8")
    _set_cell(ws, "E8", data.get('formateur', ''),
              font=_value_font(), alignment=left, border=border)
    ws.row_dimensions[8].height = 22

    # --- Tableau des stagiaires ---
    ws.row_dimensions[9].height = 8

    ws.merge_cells("C10:I10")
    _set_cell(ws, "C10", "LISTE DES STAGIAIRES",
              font=_header_font(11), fill=_header_fill(), alignment=center)
    ws.row_dimensions[10].height = 20

    # En-têtes
    headers = ["N°", "Nom et Prénom du stagiaire", "Signature Matin",
               "Signature Après-midi", "Nb heures", "Observations"]
    col_map = ["C", "D", "F", "G", "H", "I"]
    ws.merge_cells("C11:C11")
    ws.merge_cells("D11:E11")
    ws.merge_cells("F11:F11")
    ws.merge_cells("G11:G11")
    ws.merge_cells("H11:H11")
    ws.merge_cells("I11:I11")
    header_cells = ["C11", "D11", "F11", "G11", "H11", "I11"]
    for cell_c, label in zip(header_cells, headers):
        _set_cell(ws, cell_c, label,
                  font=_label_font(), fill=_accent_fill(),
                  alignment=center, border=border)
    ws.row_dimensions[11].height = 30

    # Pré-remplir la première ligne avec le stagiaire de la formation
    nom_prenom = f"{data.get('prenom', '')} {data.get('nom', '').upper()}".strip()
    row = 12
    _set_cell(ws, f"C{row}", "1", font=_value_font(), alignment=center, border=border)
    ws.merge_cells(f"D{row}:E{row}")
    _set_cell(ws, f"D{row}", nom_prenom,
              font=_value_font(), alignment=left, border=border)
    _set_cell(ws, f"F{row}", "", border=border)
    _set_cell(ws, f"G{row}", "", border=border)
    _set_cell(ws, f"H{row}", "", border=border)
    _set_cell(ws, f"I{row}", "", border=border)
    ws.row_dimensions[row].height = 22

    # Lignes vides supplémentaires (2 à 15)
    for i in range(2, 16):
        row = 12 + i - 1
        _set_cell(ws, f"C{row}", str(i), font=_value_font(), alignment=center, border=border)
        ws.merge_cells(f"D{row}:E{row}")
        _set_cell(ws, f"D{row}", "", border=border)
        _set_cell(ws, f"F{row}", "", border=border)
        _set_cell(ws, f"G{row}", "", border=border)
        _set_cell(ws, f"H{row}", "", border=border)
        _set_cell(ws, f"I{row}", "", border=border)
        ws.row_dimensions[row].height = 22

    # Signature formateur
    last = 12 + 15
    ws.row_dimensions[last].height = 8
    ws.merge_cells(f"C{last+1}:D{last+1}")
    _set_cell(ws, f"C{last+1}", "Signature du Formateur :",
              font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells(f"E{last+1}:I{last+1}")
    _set_cell(ws, f"E{last+1}", "",
              font=_value_font(), border=border)
    ws.row_dimensions[last + 1].height = 40

    note_row = last + 2
    ws.merge_cells(f"C{note_row}:I{note_row}")
    _set_cell(ws, f"C{note_row}",
              "* Par ma signature, j'atteste avoir reçu / dispensé la formation ci-dessus.",
              font=Font(name="Calibri", size=9, italic=True),
              alignment=Alignment(horizontal="center", vertical="center"))

    # Impression : 1 page
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:I{note_row}"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


# ======================  FEUILLE 3 : FICHE DE SUIVI  ======================

def _build_suivi(wb: "openpyxl.Workbook", data: Dict):
    ws = wb.create_sheet("Fiche de suivi (EQ 07 02 01)")

    # A=marge, B=marge, C..H=contenu, I=OUI, J=NON, K=marge
    col_widths = [2, 4, 30, 12, 12, 12, 12, 12, 8, 8, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = _thin_border()
    bot_border = Border(bottom=Side(style='thin', color="000000"))

    # --- Titre ---
    ws.merge_cells("C1:H1")
    _set_cell(ws, "C1", "FICHE DE SUIVI DE FORMATION",
              font=_title_font(14), alignment=center)
    ws.merge_cells("I1:K1")
    _set_cell(ws, "I1", "EQ 07 02 01 rev.6",
              font=_value_font(),
              alignment=Alignment(horizontal="right", vertical="center"))
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6

    # --- Dates / Durée ---
    ws.merge_cells("C3:D3")
    _set_cell(ws, "C3", "Formation réalisée du :",
              font=_label_font(), fill=_label_fill(), alignment=left, border=border)
    ws.merge_cells("E3:F3")
    _set_cell(ws, "E3", _fmt_date(data.get('date_debut')),
              font=_value_font(), alignment=center, border=border)
    ws.merge_cells("G3:H3")
    _set_cell(ws, "G3", "au", font=_label_font(), alignment=center)
    ws.merge_cells("I3:K3")
    _set_cell(ws, "I3", _fmt_date(data.get('date_fin')),
              font=_value_font(), alignment=center, border=border)
    ws.row_dimensions[3].height = 20

    ws.merge_cells("C4:D4")
    _set_cell(ws, "C4", "Durée de la formation :",
              font=_label_font(), fill=_label_fill(), alignment=left, border=border)
    ws.merge_cells("E4:F4")
    _set_cell(ws, "E4", _fmt_duree(data.get('duree_heures')),
              font=_value_font(), alignment=left, border=border)
    ws.merge_cells("G4:H4")
    _set_cell(ws, "G4", "Coût pédagogique :",
              font=_label_font(), fill=_label_fill(), alignment=left, border=border)
    cout = data.get('cout')
    cout_sal = data.get('cout_salarial')
    cout_str = f"{float(cout):.2f} €" if cout else "—"
    if cout_sal:
        cout_str += f"  |  Salarial : {float(cout_sal):.2f} €"
    ws.merge_cells("I4:K4")
    _set_cell(ws, "I4", cout_str,
              font=_value_font(), alignment=left, border=border)
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 6

    # --- Stagiaire / Organisme ---
    ws.merge_cells("C6:E6")
    _set_cell(ws, "C6", "Stagiaire :",
              font=Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
              fill=_header_fill(), alignment=left)
    ws.merge_cells("F6:K6")
    _set_cell(ws, "F6", "Organisme de formation :",
              font=Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
              fill=_header_fill(), alignment=left)
    ws.row_dimensions[6].height = 18

    ws.merge_cells("C7:D7")
    _set_cell(ws, "C7", "Nom :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    nom_affiche = data.get('nom', '').upper().replace('-', '-\n')
    _set_cell(ws, "E7", nom_affiche,
              font=Font(name="Calibri", bold=True, size=11),
              alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
              border=border)
    ws.merge_cells("F7:G7")
    _set_cell(ws, "F7", "Raison sociale :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("H7:K7")
    _set_cell(ws, "H7", data.get('organisme', ''),
              font=_value_font(), alignment=left, border=border)
    ws.row_dimensions[7].height = 36

    ws.merge_cells("C8:D8")
    _set_cell(ws, "C8", "Prénom :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    prenom_affiche = data.get('prenom', '').replace('-', '-\n')
    _set_cell(ws, "E8", prenom_affiche,
              font=Font(name="Calibri", bold=True, size=11),
              alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
              border=border)
    ws.merge_cells("F8:G8")
    _set_cell(ws, "F8", "Intitulé de la formation :", font=_label_font(), fill=_label_fill(),
              alignment=left, border=border)
    ws.merge_cells("H8:K8")
    _set_cell(ws, "H8", data.get('intitule', ''),
              font=_value_font(),
              alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
              border=border)
    ws.row_dimensions[8].height = 20

    # Texte description
    ws.merge_cells("C9:K9")
    _set_cell(ws, "C9",
              "Ce document permet d'évaluer la satisfaction du stagiaire en fin de formation "
              "ainsi que l'atteinte des objectifs pédagogiques.",
              font=Font(name="Calibri", size=9, italic=True),
              alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[9].height = 26

    # --- Section : Pour quel(s) raison(s) ---
    ws.merge_cells("C10:H10")
    _set_cell(ws, "C10",
              "Pour quel(s) raison(s) avez-vous suivi cette formation ?",
              font=Font(name="Calibri", bold=True, italic=True, size=10),
              alignment=left)
    _set_cell(ws, "I10", "OUI", font=_header_font(10), fill=_header_fill(),
              alignment=center, border=border)
    _set_cell(ws, "J10", "NON", font=_header_font(10), fill=_header_fill(),
              alignment=center, border=border)
    ws.row_dimensions[10].height = 20

    raisons = [
        "La formation est prévue par votre entreprise",
        "Utile pour renforcer vos compétences dans votre poste actuel",
        "Utile pour votre évaluation professionnelle",
    ]
    for i, r in enumerate(raisons):
        row = 11 + i
        ws.merge_cells(f"C{row}:H{row}")
        _set_cell(ws, f"C{row}", r, font=_value_font(), alignment=left, border=border)
        _set_cell(ws, f"I{row}", "", font=_value_font(), alignment=center, border=border)
        _set_cell(ws, f"J{row}", "", font=_value_font(), alignment=center, border=border)
        ws.row_dimensions[row].height = 18

    # --- Section : Votre évaluation ---
    ws.merge_cells("C14:H14")
    _set_cell(ws, "C14", "Votre évaluation de la formation",
              font=Font(name="Calibri", bold=True, italic=True, size=10),
              alignment=left)
    _set_cell(ws, "I14", "OUI", font=_header_font(10), fill=_header_fill(),
              alignment=center, border=border)
    _set_cell(ws, "J14", "NON", font=_header_font(10), fill=_header_fill(),
              alignment=center, border=border)
    ws.row_dimensions[14].height = 20

    ws.merge_cells("C15:K15")
    _set_cell(ws, "C15",
              "Cochez une case en fonction de votre appréciation de l'organisation et du contenu de la formation.",
              font=Font(name="Calibri", size=9, italic=True),
              alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[15].height = 18

    questions = [
        "1. Organisation et déroulement de la formation",
        "2. Le contenu est clair et adapté",
        "3. Conformité du contenu au programme",
        "4. Animation de la formation par le ou les intervenants\n( aptitude, motivation, compétence et disponibilité)",
        "5. Qualité des supports pédagogiques",
        "6. Qualité du matériel audiovisuel",
        "7. Progression de la formation (durée, rythme)",
        "8. Organisation matérielle : convocation, lieu, pauses",
        "9. L'objectif de la formation est atteint?",
        "10. La composition du groupe est satisfaisante (nombre de participants,\nniveaux homogènes)",
    ]

    for i, q in enumerate(questions):
        row = 16 + i
        ws.merge_cells(f"C{row}:H{row}")
        _set_cell(ws, f"C{row}", q,
                  font=_value_font(),
                  alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                  border=border)
        _set_cell(ws, f"I{row}", "", font=_value_font(), alignment=center, border=border)
        _set_cell(ws, f"J{row}", "", font=_value_font(), alignment=center, border=border)
        ws.row_dimensions[row].height = 18
    ws.row_dimensions[19].height = 30  # Q4 : deux lignes
    ws.row_dimensions[25].height = 30  # Q10 : deux lignes

    # Note + Commentaires (row 26)
    note_row = 26
    ws.merge_cells(f"C{note_row}:H{note_row}")
    _set_cell(ws, f"C{note_row}", "Note",
              font=_label_font(), fill=_label_fill(), alignment=left, border=border)
    ws.merge_cells(f"I{note_row}:J{note_row}")
    _set_cell(ws, f"I{note_row}", "0 / 10",
              font=_value_font(), alignment=center, border=border)
    ws.row_dimensions[note_row].height = 20

    ws.merge_cells("C27:K27")
    _set_cell(ws, "C27", "Commentaires : ___________________________________________________________",
              font=_value_font(), alignment=left)
    ws.row_dimensions[27].height = 18

    # --- Section : Votre satisfaction ---
    ws.row_dimensions[28].height = 6

    ws.merge_cells("C29:H29")
    _set_cell(ws, "C29", "Votre satisfaction :",
              font=Font(name="Calibri", bold=True, italic=True, size=10),
              alignment=left)
    _set_cell(ws, "I29", "OUI", font=_header_font(10), fill=_header_fill(),
              alignment=center, border=border)
    _set_cell(ws, "J29", "NON", font=_header_font(10), fill=_header_fill(),
              alignment=center, border=border)
    ws.row_dimensions[29].height = 20

    satisfactions = [
        "La formation a-t-elle répondu à vos attentes?",
        "Estimez-vous que la formation était en adéquation avec le métier ou\nles réalités du secteur?",
        "Recommanderiez-vous ce stage à une personne exerçant le même métier que\nvous?",
    ]
    for i, s in enumerate(satisfactions):
        row = 30 + i
        ws.merge_cells(f"C{row}:H{row}")
        _set_cell(ws, f"C{row}", s,
                  font=_value_font(),
                  alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                  border=border)
        _set_cell(ws, f"I{row}", "", font=_value_font(), alignment=center, border=border)
        _set_cell(ws, f"J{row}", "", font=_value_font(), alignment=center, border=border)
        ws.row_dimensions[row].height = 26

    # --- Section : Atteinte des objectifs ---
    ws.row_dimensions[33].height = 6

    ws.merge_cells("C34:K34")
    _set_cell(ws, "C34",
              "Evaluation de l'atteinte des objectifs fixés "
              "( à réaliser lors de l'entretien annuel d'appréciation et de progrès EG 07 61 01)",
              font=Font(name="Calibri", bold=True, italic=True, underline="single", size=10),
              alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[34].height = 28

    for r in [35, 36, 37]:
        ws.merge_cells(f"C{r}:K{r}")
        _set_cell(ws, f"C{r}", "", border=bot_border)
        ws.row_dimensions[r].height = 20

    # --- Signatures ---
    ws.row_dimensions[38].height = 6

    ws.merge_cells("C39:G39")
    _set_cell(ws, "C39", "Visa du Responsable hiérarchique (demandeur)",
              font=_label_font(), fill=_label_fill(), alignment=left, border=border)
    ws.merge_cells("H39:I39")
    _set_cell(ws, "H39", "A                              le",
              font=_value_font(), alignment=left, border=border)
    ws.merge_cells("J39:K39")
    _set_cell(ws, "J39", "", font=_value_font(), border=border)
    ws.row_dimensions[39].height = 20

    ws.merge_cells("C40:K40")
    _set_cell(ws, "C40", "Signature du stagiaire :",
              font=_label_font(), fill=_label_fill(), alignment=left, border=border)
    ws.row_dimensions[40].height = 36

    # Impression : 1 page (largeur et hauteur)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:K40"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


# ========================  POINT D'ENTRÉE PUBLIC  ========================

class FormationExportService:
    """Service de génération des documents officiels de formation pré-remplis."""

    @staticmethod
    def generate_dossier_formation(
        formation_data: Dict,
        output_dir: Optional[Path] = None
    ) -> Tuple[bool, str, Optional[str]]:
        """
        Génère un classeur xlsx avec les 3 documents officiels pré-remplis.

        Args:
            formation_data: Dict renvoyé par FormationServiceCRUD.get_formation_by_id().
                            Doit contenir : nom, prenom, intitule, organisme, lieu,
                            objectif, formateur, date_debut, date_fin, duree_heures,
                            cout, commentaire, service.
            output_dir: Répertoire de sortie (par défaut exports/formations/).

        Returns:
            (success, message, chemin_fichier)
        """
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl non installé (pip install openpyxl)", None

        try:
            out_dir = Path(output_dir) if output_dir else _EXPORTS_DIR
            out_dir.mkdir(parents=True, exist_ok=True)

            nom = formation_data.get('nom', 'INCONNU').upper()
            prenom = formation_data.get('prenom', '')
            intitule_court = formation_data.get('intitule', 'formation')[:30]
            date_str = datetime.now().strftime("%Y%m%d_%H%M")

            filename = f"Formation_{nom}_{prenom}_{intitule_court}_{date_str}.xlsx"
            filename = "".join(
                c if (c.isalnum() or c in "._- ") else "_" for c in filename
            ).strip()
            output_path = out_dir / filename

            wb = openpyxl.Workbook()
            # Supprimer la feuille par défaut
            del wb[wb.sheetnames[0]]

            _build_demande(wb, formation_data)
            _build_emargement(wb, formation_data)
            _build_suivi(wb, formation_data)

            wb.save(str(output_path))
            logger.info(f"Dossier formation généré : {output_path}")
            return True, "Dossier de formation généré avec succès.", str(output_path)

        except Exception as e:
            logger.exception(f"Erreur génération dossier formation: {e}")
            return False, f"Erreur lors de la génération : {e}", None

    @staticmethod
    def open_file(path: str):
        """Ouvre le fichier avec l'application par défaut du système."""
        import subprocess
        import sys
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(["open", path])
            else:
                subprocess.run(["xdg-open", path])
        except Exception as e:
            logger.error(f"Impossible d'ouvrir {path}: {e}")
