# -*- coding: utf-8 -*-
"""
Service de gestion des documents templates.

Ce service gere les templates de documents (formulaires d'evaluation, consignes, etc.)
qui peuvent etre pre-remplis avec les informations de l'operateur et ouverts pour impression.

Stockage : Les templates sont stockes en BLOB dans la base de donnees MySQL.
Les anciens templates sur filesystem (legacy) restent accessibles via fichier_source.
"""

import os
import sys
import json
import shutil
import tempfile
import mimetypes
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple

from core.db.query_executor import QueryExecutor
from core.db.configbd import get_connection
from core.services.logger import log_hist
from core.utils.logging_config import get_logger

logger = get_logger(__name__)


# Taille max recommandee pour un template (16 Mo)
MAX_TEMPLATE_SIZE_BYTES = 16 * 1024 * 1024


def get_templates_dir() -> Path:
    """
    Retourne le repertoire des templates (pour legacy filesystem).

    Priorite de resolution :
    1. Variable d'environnement EMAC_TEMPLATES_DIR
    2. En production (.exe): %APPDATA%/EMAC/templates/
    3. En developpement: App/templates/
    """
    custom_dir = os.environ.get('EMAC_TEMPLATES_DIR')
    if custom_dir:
        custom_path = Path(custom_dir)
        if custom_path.exists():
            return custom_path
        else:
            logger.warning(
                f"EMAC_TEMPLATES_DIR pointe vers un chemin inaccessible: {custom_dir}"
            )

    if getattr(sys, 'frozen', False):
        app_data = os.environ.get('APPDATA', os.path.expanduser('~'))
        return Path(app_data) / 'EMAC' / 'templates'

    return Path(__file__).parent.parent.parent / 'templates'


def get_temp_dir() -> Path:
    """Retourne le repertoire temporaire pour les fichiers generes."""
    temp_base = Path(tempfile.gettempdir()) / 'EMAC_templates'
    temp_base.mkdir(parents=True, exist_ok=True)
    return temp_base


def get_all_templates(actif_only: bool = True) -> List[Dict]:
    """
    Recupere tous les templates (metadonnees, sans le BLOB).

    Args:
        actif_only: Si True, ne retourne que les templates actifs

    Returns:
        Liste des templates avec leurs metadonnees
    """
    query = """
        SELECT id, nom, fichier_source, contexte, postes_associes,
               champ_operateur, champ_auditeur, champ_date,
               obligatoire, description, ordre_affichage, actif, stockage_type
        FROM documents_templates
    """
    if actif_only:
        query += " WHERE actif = TRUE"
    query += " ORDER BY contexte, ordre_affichage"

    templates = QueryExecutor.fetch_all(query, dictionary=True)

    for t in templates:
        if t['postes_associes']:
            try:
                t['postes_associes'] = json.loads(t['postes_associes'])
            except json.JSONDecodeError:
                t['postes_associes'] = []
        else:
            t['postes_associes'] = []

    return templates


def get_templates_by_contexte(contexte: str) -> List[Dict]:
    """
    Recupere les templates pour un contexte donne.

    Args:
        contexte: 'NOUVEL_OPERATEUR', 'NIVEAU_3', ou 'POSTE'
    """
    query = """
        SELECT id, nom, fichier_source, contexte, postes_associes,
               champ_operateur, champ_auditeur, champ_date,
               obligatoire, description, ordre_affichage, stockage_type
        FROM documents_templates
        WHERE contexte = %s AND actif = TRUE
        ORDER BY ordre_affichage
    """

    templates = QueryExecutor.fetch_all(query, (contexte,), dictionary=True)

    for t in templates:
        if t['postes_associes']:
            try:
                t['postes_associes'] = json.loads(t['postes_associes'])
            except json.JSONDecodeError:
                t['postes_associes'] = []
        else:
            t['postes_associes'] = []

    return templates


def get_templates_for_poste(code_poste: str) -> List[Dict]:
    """Recupere les templates associes a un poste specifique."""
    all_poste_templates = get_templates_by_contexte('POSTE')

    matching = []
    for t in all_poste_templates:
        if code_poste in t['postes_associes']:
            matching.append(t)

    return matching


def get_templates_for_nouvel_operateur() -> List[Dict]:
    """Recupere les templates pour un nouvel operateur."""
    return get_templates_by_contexte('NOUVEL_OPERATEUR')


def get_templates_for_niveau_3() -> List[Dict]:
    """Recupere les templates pour le passage au niveau 3."""
    return get_templates_by_contexte('NIVEAU_3')


def _get_template_content(template: Dict) -> Optional[Tuple[bytes, str]]:
    """
    Recupere le contenu binaire d'un template.

    Args:
        template: Dict avec au minimum 'id', 'stockage_type', 'fichier_source'

    Returns:
        Tuple (contenu_bytes, extension) ou None si introuvable
    """
    # BLOB: lire depuis la base
    if template.get('stockage_type') == 'BLOB':
        row = QueryExecutor.fetch_one(
            "SELECT contenu_fichier, fichier_source FROM documents_templates WHERE id = %s",
            (template['id'],),
            dictionary=True
        )
        if row and row['contenu_fichier']:
            ext = Path(row['fichier_source']).suffix
            return (row['contenu_fichier'], ext)

    # FILESYSTEM (legacy): lire depuis le disque
    fichier_source = template['fichier_source']
    templates_dir = get_templates_dir()

    # Nettoyer le chemin
    fichier_clean = fichier_source.replace('templates/', '').replace('\\', '/').strip('/')

    if '..' in fichier_clean or fichier_clean.startswith('/'):
        logger.warning(f"Tentative de path traversal detectee: {fichier_source}")
        return None

    source_path = (templates_dir / fichier_clean).resolve()

    try:
        source_path.relative_to(templates_dir.resolve())
    except ValueError:
        logger.warning(f"Path traversal bloque: {source_path} hors de {templates_dir}")
        return None

    if not source_path.exists():
        return None

    with open(source_path, 'rb') as f:
        return (f.read(), source_path.suffix)


def generate_filled_template(
    template_id: int,
    operateur_nom: str = "",
    operateur_prenom: str = "",
    auditeur_nom: str = "",
    date_str: str = None
) -> Tuple[bool, str, Optional[str]]:
    """
    Genere une copie du template pre-remplie avec les informations fournies.

    Fonctionne avec les templates BLOB (en base) et FILESYSTEM (legacy).

    Args:
        template_id: ID du template
        operateur_nom: Nom de l'operateur
        operateur_prenom: Prenom de l'operateur
        auditeur_nom: Nom de l'auditeur/evaluateur
        date_str: Date au format DD/MM/YYYY (par defaut: aujourd'hui)

    Returns:
        Tuple (succes, message, chemin_fichier_genere)
    """
    # Recuperer les infos du template
    template = QueryExecutor.fetch_one(
        "SELECT * FROM documents_templates WHERE id = %s",
        (template_id,),
        dictionary=True
    )

    if not template:
        return False, "Template non trouve", None

    # Recuperer le contenu du template
    content_result = _get_template_content(template)
    if not content_result:
        return False, "Fichier template non trouve ou inaccessible", None

    contenu, extension = content_result

    # Date par defaut
    if not date_str:
        date_str = datetime.now().strftime('%d/%m/%Y')

    # Nom complet de l'operateur
    operateur_complet = f"{operateur_prenom} {operateur_nom}".strip()

    # Creer le nom du fichier de sortie
    safe_name = f"{template['nom']}_{operateur_complet}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    safe_name = "".join(c for c in safe_name if c.isalnum() or c in (' ', '_', '-')).strip()
    output_filename = f"{safe_name}{extension}"

    # Ecrire dans le repertoire temporaire
    temp_dir = get_temp_dir()
    output_path = temp_dir / output_filename

    with open(output_path, 'wb') as f:
        f.write(contenu)

    # Pre-remplir si c'est un Excel
    if extension.lower() in ('.xlsx', '.xlsm', '.xls'):
        try:
            _fill_excel_template(
                output_path,
                template['champ_operateur'],
                template['champ_auditeur'],
                template['champ_date'],
                operateur_complet,
                auditeur_nom,
                date_str
            )
        except Exception as e:
            # Si le pre-remplissage echoue, on garde le fichier non rempli
            logger.warning(f"Pre-remplissage echoue pour {template['nom']}: {e}")

    # Logger l'action
    log_hist(
        "GENERATION_TEMPLATE",
        f"Generation du template '{template['nom']}' pour {operateur_complet}"
    )

    return True, "Fichier genere avec succes", str(output_path)


def upload_template(
    template_id: int,
    fichier_source: str
) -> Tuple[bool, str]:
    """
    Upload un fichier template dans la base de donnees (BLOB).

    Args:
        template_id: ID du template existant a mettre a jour
        fichier_source: Chemin du fichier a uploader

    Returns:
        (succes, message)
    """
    source_path = Path(fichier_source)
    if not source_path.exists():
        return False, f"Fichier introuvable: {fichier_source}"

    taille = source_path.stat().st_size
    if taille > MAX_TEMPLATE_SIZE_BYTES:
        taille_mo = taille / (1024 * 1024)
        return False, f"Fichier trop volumineux ({taille_mo:.1f} Mo). Maximum: 16 Mo"

    try:
        with open(source_path, 'rb') as f:
            contenu = f.read()

        type_mime, _ = mimetypes.guess_type(str(source_path))
        if type_mime is None:
            type_mime = "application/octet-stream"

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(
            """UPDATE documents_templates
               SET contenu_fichier = %s,
                   type_mime = %s,
                   taille_octets = %s,
                   stockage_type = 'BLOB',
                   fichier_source = %s
               WHERE id = %s""",
            (contenu, type_mime, taille, source_path.name, template_id)
        )
        conn.commit()
        cursor.close()
        conn.close()

        logger.info(f"Template ID {template_id} uploade en BLOB ({taille} octets)")
        return True, f"Template mis a jour avec succes ({source_path.name})"

    except Exception as e:
        logger.exception(f"Erreur upload template: {e}")
        return False, f"Erreur lors de l'upload: {str(e)}"


def add_template(
    nom: str,
    fichier_source: str,
    contexte: str,
    postes_associes: List[str] = None,
    champ_operateur: str = None,
    champ_auditeur: str = None,
    champ_date: str = None,
    obligatoire: bool = False,
    description: str = None,
    ordre_affichage: int = 0
) -> Tuple[bool, str, Optional[int]]:
    """
    Ajoute un nouveau template avec son fichier stocke en BLOB.

    Args:
        nom: Nom affiche du template
        fichier_source: Chemin du fichier a uploader
        contexte: 'NOUVEL_OPERATEUR', 'NIVEAU_3', ou 'POSTE'
        postes_associes: Liste des codes postes (pour contexte POSTE)
        champ_operateur: Cellule Excel pour le nom operateur (ex: "D7")
        champ_auditeur: Cellule Excel pour le nom auditeur
        champ_date: Cellule Excel pour la date
        obligatoire: Document obligatoire
        description: Description du template
        ordre_affichage: Ordre d'affichage

    Returns:
        (succes, message, template_id)
    """
    source_path = Path(fichier_source)
    if not source_path.exists():
        return False, f"Fichier introuvable: {fichier_source}", None

    taille = source_path.stat().st_size
    if taille > MAX_TEMPLATE_SIZE_BYTES:
        taille_mo = taille / (1024 * 1024)
        return False, f"Fichier trop volumineux ({taille_mo:.1f} Mo). Maximum: 16 Mo", None

    try:
        with open(source_path, 'rb') as f:
            contenu = f.read()

        type_mime, _ = mimetypes.guess_type(str(source_path))
        if type_mime is None:
            type_mime = "application/octet-stream"

        postes_json = json.dumps(postes_associes) if postes_associes else None

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(
            """INSERT INTO documents_templates (
                nom, fichier_source, contenu_fichier, type_mime, taille_octets,
                stockage_type, contexte, postes_associes,
                champ_operateur, champ_auditeur, champ_date,
                obligatoire, description, ordre_affichage
            ) VALUES (
                %s, %s, %s, %s, %s, 'BLOB', %s, %s, %s, %s, %s, %s, %s, %s
            )""",
            (
                nom, source_path.name, contenu, type_mime, taille,
                contexte, postes_json,
                champ_operateur, champ_auditeur, champ_date,
                obligatoire, description, ordre_affichage
            )
        )

        template_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()

        logger.info(f"Nouveau template '{nom}' ajoute en BLOB (ID: {template_id})")
        return True, f"Template '{nom}' ajoute avec succes", template_id

    except Exception as e:
        logger.exception(f"Erreur ajout template: {e}")
        return False, f"Erreur lors de l'ajout: {str(e)}", None


def _fill_excel_template(
    file_path: Path,
    cell_operateur: str,
    cell_auditeur: str,
    cell_date: str,
    operateur_nom: str,
    auditeur_nom: str,
    date_str: str
) -> Tuple[bool, str]:
    """
    Remplit les cellules d'un fichier Excel avec les donnees fournies.
    """
    file_path = Path(file_path)
    extension = file_path.suffix.lower()

    if extension == '.xls':
        return True, "Fichier .xls copie (pre-remplissage non supporte pour ce format)"

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ws = wb.active

        if cell_operateur and operateur_nom:
            ws[cell_operateur] = operateur_nom

        if cell_auditeur and auditeur_nom:
            ws[cell_auditeur] = auditeur_nom

        if cell_date and date_str:
            ws[cell_date] = date_str

        wb.save(file_path)
        wb.close()

        return True, "Fichier rempli avec succes"

    except Exception as e:
        return False, f"Erreur lors du remplissage: {str(e)}"


def open_template_file(file_path: str) -> Tuple[bool, str]:
    """
    Ouvre un fichier template avec l'application par defaut du systeme.

    SECURITE: Valide le chemin avant ouverture.
    """
    import subprocess
    import platform

    try:
        path = Path(file_path).resolve()

        if not path.exists():
            return False, "Fichier non trouve"

        if not path.is_file():
            return False, "Chemin invalide"

        # SECURITE: Verifier que le chemin est dans un repertoire autorise
        temp_dir = get_temp_dir().resolve()
        templates_dir = get_templates_dir().resolve()

        is_in_temp = False
        is_in_templates = False
        try:
            path.relative_to(temp_dir)
            is_in_temp = True
        except ValueError:
            pass
        try:
            path.relative_to(templates_dir)
            is_in_templates = True
        except ValueError:
            pass

        if not (is_in_temp or is_in_templates):
            logger.warning(f"Tentative d'ouverture de fichier hors zone autorisee: {path}")
            return False, "Acces au fichier refuse"

        file_str = str(path)

        if platform.system() == 'Windows':
            os.startfile(file_str)
        elif platform.system() == 'Darwin':
            subprocess.run(['open', file_str], check=True)
        else:
            subprocess.run(['xdg-open', file_str], check=True)

        return True, "Fichier ouvert"

    except Exception as e:
        logger.error(f"Erreur ouverture fichier: {e}")
        return False, "Impossible d'ouvrir le fichier"


def print_template_file(file_path: str, printer_name: str) -> Tuple[bool, str]:
    """
    Imprime un fichier vers une imprimante specifique.

    Pour les fichiers Excel (.xlsx / .xlsm), tous les onglets sont imprimes
    via l'automatisation COM d'Excel (PowerShell).
    Pour les autres formats, ShellExecute 'printto' est utilise.

    SECURITE: Valide le chemin avant impression (meme regles que open_template_file).

    Args:
        file_path: Chemin absolu du fichier a imprimer
        printer_name: Nom de l'imprimante Windows cible

    Returns:
        Tuple (succes, message)
    """
    import subprocess

    try:
        path = Path(file_path).resolve()

        if not path.exists() or not path.is_file():
            return False, "Fichier non trouve"

        # SECURITE: chemin dans zone autorisee uniquement
        temp_dir = get_temp_dir().resolve()
        templates_dir = get_templates_dir().resolve()
        in_allowed = False
        for allowed in (temp_dir, templates_dir):
            try:
                path.relative_to(allowed)
                in_allowed = True
                break
            except ValueError:
                pass
        if not in_allowed:
            logger.warning(f"Impression refusee hors zone autorisee: {path}")
            return False, "Acces au fichier refuse"

        ext = path.suffix.lower()

        if ext in ('.xlsx', '.xlsm') and sys.platform == 'win32':
            # Excel: COM via PowerShell → imprime TOUS les onglets du classeur
            path_esc = str(path).replace("'", "''")
            printer_esc = printer_name.replace("'", "''")

            ps_cmd = (
                f"$ErrorActionPreference = 'Stop'; "
                f"$xl = New-Object -ComObject Excel.Application; "
                f"$xl.Visible = $false; $xl.DisplayAlerts = $false; "
                f"try {{ "
                f"  $wb = $xl.Workbooks.Open('{path_esc}'); "
                f"  $wb.PrintOut([Type]::Missing, [Type]::Missing, 1, $false, '{printer_esc}'); "
                f"  $wb.Close($false) "
                f"}} finally {{ "
                f"  $xl.Quit(); "
                f"  [System.Runtime.Interopservices.Marshal]::ReleaseComObject($xl) | Out-Null; "
                f"  [System.GC]::Collect() "
                f"}}"
            )
            result = subprocess.run(
                ['powershell', '-NoProfile', '-NonInteractive', '-Command', ps_cmd],
                capture_output=True, text=True, timeout=120
            )
            if result.returncode != 0:
                err = (result.stderr or result.stdout or "").strip()
                logger.warning(f"Impression Excel echouee (PowerShell): {err}")
                return False, f"Impression echouee: {err}"
            return True, "Impression lancee (tous les onglets)"

        else:
            # Autres formats (PDF, Word, .xls…): ShellExecute printto
            import ctypes
            ret = ctypes.windll.shell32.ShellExecuteW(
                None, 'printto', str(path), f'"{printer_name}"', None, 0
            )
            if ret <= 32:
                logger.warning(f"ShellExecute printto echoue (code {ret}) pour {path}")
                return False, f"Impression echouee (code {ret})"
            return True, "Impression lancee"

    except Exception as e:
        logger.exception(f"Erreur impression fichier: {e}")
        return False, "Impossible de lancer l'impression"


def check_templates_table_exists() -> bool:
    """Verifie si la table documents_templates existe."""
    try:
        count = QueryExecutor.fetch_scalar("""
            SELECT COUNT(*) FROM information_schema.tables
            WHERE table_schema = DATABASE()
            AND table_name = 'documents_templates'
        """, default=0)
        return count > 0
    except Exception:
        return False


def get_template_by_id(template_id: int) -> Optional[Dict]:
    """Recupere un template par son ID."""
    template = QueryExecutor.fetch_one(
        """SELECT id, nom, fichier_source, contexte, postes_associes,
                  champ_operateur, champ_auditeur, champ_date,
                  obligatoire, description, ordre_affichage, actif, stockage_type
           FROM documents_templates WHERE id = %s""",
        (template_id,),
        dictionary=True
    )

    if template and template['postes_associes']:
        try:
            template['postes_associes'] = json.loads(template['postes_associes'])
        except json.JSONDecodeError:
            template['postes_associes'] = []

    return template


def get_postes_for_operateur(operateur_id: int) -> List[str]:
    """Recupere les codes de tous les postes auxquels un operateur est affecte."""
    results = QueryExecutor.fetch_all("""
        SELECT DISTINCT p.numposte
        FROM polyvalence pv
        JOIN postes p ON pv.poste_id = p.id
        WHERE pv.operateur_id = %s
    """, (operateur_id,), dictionary=True)

    return [r['numposte'] for r in results if r['numposte']]


def get_all_templates_for_operateur(operateur_id: int) -> Dict[str, List[Dict]]:
    """Recupere tous les templates applicables a un operateur, groupes par contexte."""
    result = {
        'NOUVEL_OPERATEUR': get_templates_for_nouvel_operateur(),
        'NIVEAU_3': get_templates_for_niveau_3(),
        'POSTE': []
    }

    postes = get_postes_for_operateur(operateur_id)

    seen_ids = set()
    for code_poste in postes:
        templates = get_templates_for_poste(code_poste)
        for t in templates:
            if t['id'] not in seen_ids:
                seen_ids.add(t['id'])
                result['POSTE'].append(t)

    return result
