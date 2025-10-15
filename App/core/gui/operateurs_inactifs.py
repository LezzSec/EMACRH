# operateurs_inactifs.py – Consultation des opérateurs inactifs et leurs historiques
# Affiche tous les opérateurs INACTIF avec leurs polyvalences (postes, niveaux, dates)

from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QLineEdit, QComboBox, QGroupBox,
    QMessageBox, QAbstractItemView, QWidget
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor

from core.db.configbd import get_connection as get_db_connection

import datetime as dt


# -------- Helpers DB --------
def _cursor(conn):
    """Retourne (cursor, dict_mode)."""
    try:
        cur = conn.cursor(dictionary=True)
        return cur, True
    except TypeError:
        cur = conn.cursor()
        return cur, False


def _rows(cur, dict_mode):
    """Retourne une liste de dicts quelle que soit la lib DB."""
    if dict_mode:
        return cur.fetchall()
    names = [d[0] for d in cur.description]
    return [dict(zip(names, r)) for r in cur.fetchall()]


class OperateursInactifsDialog(QDialog):
    """
    Fenêtre d'affichage des opérateurs inactifs avec leurs statistiques complètes :
    - Liste des opérateurs INACTIF
    - Détails de toutes leurs polyvalences (postes, niveaux, dates)
    - Filtres : recherche par nom, filtre par poste
    - Export possible des données
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Historique des Opérateurs Inactifs")
        self.setGeometry(100, 100, 1200, 700)
        
        self.operateurs_inactifs = []  # Liste complète des opérateurs inactifs
        self.current_operateur_id = None
        
        layout = QVBoxLayout(self)
        
        # === Header ===
        header = QLabel("📋 Opérateurs Inactifs - Historique Complet")
        header.setFont(QFont("Arial", 16, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        # === Section Filtres ===
        filters_group = QGroupBox("Filtres")
        filters_layout = QHBoxLayout()
        
        filters_layout.addWidget(QLabel("Recherche :"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Nom ou prénom de l'opérateur...")
        self.search_input.textChanged.connect(self.filter_operateurs)
        filters_layout.addWidget(self.search_input, 1)
        
        filters_layout.addWidget(QLabel("Poste occupé :"))
        self.poste_filter = QComboBox()
        self.poste_filter.addItem("(Tous)", None)
        self.poste_filter.currentIndexChanged.connect(self.filter_operateurs)
        filters_layout.addWidget(self.poste_filter)
        
        self.refresh_btn = QPushButton("Actualiser")
        self.refresh_btn.clicked.connect(self.load_data)
        filters_layout.addWidget(self.refresh_btn)
        
        filters_group.setLayout(filters_layout)
        layout.addWidget(filters_group)
        
        # === Section Principale (2 colonnes) ===
        main_content = QHBoxLayout()
        
        # Colonne gauche : Liste des opérateurs inactifs
        left_panel = QVBoxLayout()
        
        left_header = QLabel("Opérateurs Inactifs")
        left_header.setFont(QFont("Arial", 12, QFont.Bold))
        left_panel.addWidget(left_header)
        
        self.operateurs_table = QTableWidget()
        self.operateurs_table.setColumnCount(3)
        self.operateurs_table.setHorizontalHeaderLabels(["Nom", "Prénom", "Nb Postes"])
        self.operateurs_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.operateurs_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.operateurs_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.operateurs_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.operateurs_table.itemSelectionChanged.connect(self.on_operateur_selected)
        left_panel.addWidget(self.operateurs_table)
        
        # Stats rapides
        self.stats_label = QLabel("Total : 0 opérateur(s) inactif(s)")
        self.stats_label.setStyleSheet("color: #6b7280; font-style: italic;")
        left_panel.addWidget(self.stats_label)
        
        main_content.addLayout(left_panel, 1)
        
        # Colonne droite : Détails des polyvalences
        right_panel = QVBoxLayout()
        
        self.detail_header = QLabel("Sélectionnez un opérateur")
        self.detail_header.setFont(QFont("Arial", 12, QFont.Bold))
        right_panel.addWidget(self.detail_header)
        
        self.polyvalences_table = QTableWidget()
        self.polyvalences_table.setColumnCount(5)
        self.polyvalences_table.setHorizontalHeaderLabels([
            "Poste", "Niveau", "Date Évaluation", "Prochaine Éval.", "Ancienneté"
        ])
        self.polyvalences_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.polyvalences_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        right_panel.addWidget(self.polyvalences_table)
        
        # Statistiques détaillées
        stats_panel = QHBoxLayout()
        
        self.stat_n1 = self._create_stat_card("Niveau 1", "0", "#ef4444")
        self.stat_n2 = self._create_stat_card("Niveau 2", "0", "#f59e0b")
        self.stat_n3 = self._create_stat_card("Niveau 3", "0", "#10b981")
        self.stat_n4 = self._create_stat_card("Niveau 4", "0", "#3b82f6")
        
        stats_panel.addWidget(self.stat_n1)
        stats_panel.addWidget(self.stat_n2)
        stats_panel.addWidget(self.stat_n3)
        stats_panel.addWidget(self.stat_n4)
        
        right_panel.addLayout(stats_panel)
        
        main_content.addLayout(right_panel, 2)
        
        layout.addLayout(main_content, 1)
        
        # === Boutons d'action ===
        actions = QHBoxLayout()
        actions.addStretch()
        
        self.export_btn = QPushButton("Exporter en Excel")
        self.export_btn.clicked.connect(self.export_data)
        actions.addWidget(self.export_btn)
        
        self.close_btn = QPushButton("Fermer")
        self.close_btn.clicked.connect(self.close)
        actions.addWidget(self.close_btn)
        
        layout.addLayout(actions)
        
        # Chargement initial
        self.load_postes_filter()
        self.load_data()
    
    def _create_stat_card(self, label, value, color):
        """Crée une carte statistique colorée."""
        card = QWidget()
        card.setStyleSheet(f"""
            QWidget {{
                background: {color};
                border-radius: 8px;
                padding: 10px;
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(8, 8, 8, 8)
        
        value_label = QLabel(value)
        value_label.setFont(QFont("Arial", 20, QFont.Bold))
        value_label.setStyleSheet("color: white;")
        value_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(value_label)
        
        text_label = QLabel(label)
        text_label.setStyleSheet("color: white; font-size: 11px;")
        text_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(text_label)
        
        # Stocker le label de valeur pour mise à jour
        card.value_label = value_label
        
        return card
    
    def load_postes_filter(self):
        """Charge les postes dans le filtre."""
        try:
            connection = get_db_connection()
            cursor, dict_mode = _cursor(connection)
            
            cursor.execute(
                "SELECT DISTINCT poste_code FROM postes "
                "WHERE COALESCE(visible, 1) = 1 "
                "ORDER BY poste_code"
            )
            rows = _rows(cursor, dict_mode)
            
            self.poste_filter.blockSignals(True)
            self.poste_filter.clear()
            self.poste_filter.addItem("(Tous)", None)
            
            for r in rows:
                poste_code = r.get("poste_code", "")
                self.poste_filter.addItem(poste_code, poste_code)
            
            self.poste_filter.blockSignals(False)
            
            cursor.close()
            connection.close()
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de charger les postes :\n{e}")
    
    def load_data(self):
        """Charge tous les opérateurs inactifs avec leurs statistiques."""
        try:
            connection = get_db_connection()
            cursor, dict_mode = _cursor(connection)
            
            # Requête pour récupérer les opérateurs inactifs avec nombre de postes
            query = """
                SELECT 
                    o.id,
                    o.nom,
                    o.prenom,
                    COUNT(p.id) as nb_postes
                FROM operateurs o
                LEFT JOIN polyvalence p ON o.id = p.operateur_id
                WHERE UPPER(o.statut) = 'INACTIF'
                GROUP BY o.id, o.nom, o.prenom
                ORDER BY o.nom, o.prenom
            """
            
            cursor.execute(query)
            rows = _rows(cursor, dict_mode)
            
            self.operateurs_inactifs = rows
            
            cursor.close()
            connection.close()
            
            # Afficher dans la table
            self.filter_operateurs()
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de charger les données :\n{e}")
    
    def filter_operateurs(self):
        """Filtre les opérateurs selon les critères."""
        search_text = self.search_input.text().lower()
        poste_filter = self.poste_filter.currentData()
        
        self.operateurs_table.setRowCount(0)
        
        filtered_count = 0
        
        for op in self.operateurs_inactifs:
            nom = op.get("nom", "").lower()
            prenom = op.get("prenom", "").lower()
            
            # Filtre recherche
            if search_text and search_text not in nom and search_text not in prenom:
                continue
            
            # Filtre poste (vérifier si l'opérateur a occupé ce poste)
            if poste_filter:
                if not self._operateur_has_poste(op.get("id"), poste_filter):
                    continue
            
            # Ajouter à la table
            row = self.operateurs_table.rowCount()
            self.operateurs_table.insertRow(row)
            
            item_nom = QTableWidgetItem(op.get("nom", ""))
            item_nom.setData(Qt.UserRole, op.get("id"))  # Stocker l'ID
            self.operateurs_table.setItem(row, 0, item_nom)
            
            self.operateurs_table.setItem(row, 1, QTableWidgetItem(op.get("prenom", "")))
            self.operateurs_table.setItem(row, 2, QTableWidgetItem(str(op.get("nb_postes", 0))))
            
            filtered_count += 1
        
        self.stats_label.setText(f"Total : {filtered_count} opérateur(s) inactif(s)")
    
    def _operateur_has_poste(self, operateur_id, poste_code):
        """Vérifie si un opérateur a occupé un poste spécifique."""
        try:
            connection = get_db_connection()
            cursor, dict_mode = _cursor(connection)
            
            cursor.execute(
                """SELECT COUNT(*) as cnt FROM polyvalence p
                   JOIN postes ps ON p.poste_id = ps.id
                   WHERE p.operateur_id = %s AND ps.poste_code = %s""",
                (operateur_id, poste_code)
            )
            
            result = cursor.fetchone()
            count = result["cnt"] if dict_mode else result[0]
            
            cursor.close()
            connection.close()
            
            return count > 0
            
        except Exception:
            return False
    
    def on_operateur_selected(self):
        """Charge les polyvalences de l'opérateur sélectionné."""
        selected = self.operateurs_table.selectedItems()
        if not selected:
            return
        
        # Récupérer l'ID de l'opérateur
        operateur_id = selected[0].data(Qt.UserRole)
        nom = selected[0].text()
        prenom = selected[1].text()
        
        self.current_operateur_id = operateur_id
        self.detail_header.setText(f"Polyvalences de {nom} {prenom}")
        
        self.load_polyvalences(operateur_id)
    
    def load_polyvalences(self, operateur_id):
        """Charge toutes les polyvalences d'un opérateur."""
        try:
            connection = get_db_connection()
            cursor, dict_mode = _cursor(connection)
            
            query = """
                SELECT 
                    ps.poste_code,
                    p.niveau,
                    p.date_evaluation,
                    p.prochaine_evaluation
                FROM polyvalence p
                JOIN postes ps ON p.poste_id = ps.id
                WHERE p.operateur_id = %s
                ORDER BY ps.poste_code
            """
            
            cursor.execute(query, (operateur_id,))
            rows = _rows(cursor, dict_mode)
            
            cursor.close()
            connection.close()
            
            # Réinitialiser la table
            self.polyvalences_table.setRowCount(0)
            
            # Compteurs pour stats
            niveaux = {1: 0, 2: 0, 3: 0, 4: 0}
            
            for r in rows:
                row = self.polyvalences_table.rowCount()
                self.polyvalences_table.insertRow(row)
                
                poste = r.get("poste_code", "")
                niveau = r.get("niveau")
                date_eval = r.get("date_evaluation")
                date_next = r.get("prochaine_evaluation")
                
                # Poste
                self.polyvalences_table.setItem(row, 0, QTableWidgetItem(str(poste)))
                
                # Niveau (avec couleur)
                niveau_item = QTableWidgetItem(str(niveau) if niveau else "N/A")
                if niveau:
                    niveaux[int(niveau)] += 1
                    if niveau == 1:
                        niveau_item.setBackground(QColor("#fef2f2"))
                        niveau_item.setForeground(QColor("#dc2626"))
                    elif niveau == 2:
                        niveau_item.setBackground(QColor("#fffbeb"))
                        niveau_item.setForeground(QColor("#d97706"))
                    elif niveau == 3:
                        niveau_item.setBackground(QColor("#f0fdf4"))
                        niveau_item.setForeground(QColor("#059669"))
                    elif niveau == 4:
                        niveau_item.setBackground(QColor("#eff6ff"))
                        niveau_item.setForeground(QColor("#2563eb"))
                niveau_item.setTextAlignment(Qt.AlignCenter)
                self.polyvalences_table.setItem(row, 1, niveau_item)
                
                # Date évaluation
                date_eval_str = self._format_date(date_eval)
                self.polyvalences_table.setItem(row, 2, QTableWidgetItem(date_eval_str))
                
                # Prochaine évaluation
                date_next_str = self._format_date(date_next)
                self.polyvalences_table.setItem(row, 3, QTableWidgetItem(date_next_str))
                
                # Ancienneté (calcul depuis date_evaluation)
                anciennete = self._calculate_anciennete(date_eval)
                self.polyvalences_table.setItem(row, 4, QTableWidgetItem(anciennete))
            
            # Mettre à jour les statistiques
            self.stat_n1.value_label.setText(str(niveaux[1]))
            self.stat_n2.value_label.setText(str(niveaux[2]))
            self.stat_n3.value_label.setText(str(niveaux[3]))
            self.stat_n4.value_label.setText(str(niveaux[4]))
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de charger les polyvalences :\n{e}")
    
    def _format_date(self, date_val):
        """Formate une date pour l'affichage."""
        if date_val is None:
            return "N/A"
        
        if isinstance(date_val, str):
            try:
                date_obj = dt.datetime.strptime(date_val, "%Y-%m-%d")
                return date_obj.strftime("%d/%m/%Y")
            except Exception:
                return date_val
        
        if hasattr(date_val, "strftime"):
            return date_val.strftime("%d/%m/%Y")
        
        return str(date_val)
    
    def _calculate_anciennete(self, date_eval):
        """Calcule l'ancienneté depuis la date d'évaluation."""
        if date_eval is None:
            return "N/A"
        
        try:
            if isinstance(date_eval, str):
                date_obj = dt.datetime.strptime(date_eval, "%Y-%m-%d").date()
            elif hasattr(date_eval, "date"):
                date_obj = date_eval.date()
            else:
                date_obj = date_eval
            
            today = dt.date.today()
            delta = today - date_obj
            
            years = delta.days // 365
            months = (delta.days % 365) // 30
            
            if years > 0:
                return f"{years} an(s) {months} mois"
            elif months > 0:
                return f"{months} mois"
            else:
                return f"{delta.days} jour(s)"
                
        except Exception:
            return "N/A"
    
    def export_data(self):
        """Exporte les données en Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from PyQt5.QtWidgets import QFileDialog
            
            # Demander le fichier de destination
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Exporter les données", 
                f"operateurs_inactifs_{dt.date.today().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            wb = Workbook()
            
            # Feuille 1 : Liste des opérateurs
            ws1 = wb.active
            ws1.title = "Opérateurs Inactifs"
            
            # En-têtes
            headers1 = ["Nom", "Prénom", "Nombre de Postes"]
            ws1.append(headers1)
            
            # Style en-têtes
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Données
            for op in self.operateurs_inactifs:
                ws1.append([
                    op.get("nom", ""),
                    op.get("prenom", ""),
                    op.get("nb_postes", 0)
                ])
            
            # Feuille 2 : Détails des polyvalences
            ws2 = wb.create_sheet("Détails Polyvalences")
            
            headers2 = ["Nom", "Prénom", "Poste", "Niveau", "Date Évaluation", "Prochaine Évaluation"]
            ws2.append(headers2)
            
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Récupérer toutes les polyvalences
            connection = get_db_connection()
            cursor, dict_mode = _cursor(connection)
            
            query = """
                SELECT 
                    o.nom,
                    o.prenom,
                    ps.poste_code,
                    p.niveau,
                    p.date_evaluation,
                    p.prochaine_evaluation
                FROM operateurs o
                JOIN polyvalence p ON o.id = p.operateur_id
                JOIN postes ps ON p.poste_id = ps.id
                WHERE UPPER(o.statut) = 'INACTIF'
                ORDER BY o.nom, o.prenom, ps.poste_code
            """
            
            cursor.execute(query)
            rows = _rows(cursor, dict_mode)
            
            for r in rows:
                ws2.append([
                    r.get("nom", ""),
                    r.get("prenom", ""),
                    r.get("poste_code", ""),
                    r.get("niveau", ""),
                    self._format_date(r.get("date_evaluation")),
                    self._format_date(r.get("prochaine_evaluation"))
                ])
            
            cursor.close()
            connection.close()
            
            # Ajuster les largeurs
            for ws in [ws1, ws2]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder
            wb.save(file_path)
            
            QMessageBox.information(
                self, "Export réussi",
                f"✅ Les données ont été exportées avec succès !\n\n{file_path}"
            )
            
        except ImportError:
            QMessageBox.warning(
                self, "Module manquant",
                "Le module 'openpyxl' est requis pour l'export Excel.\n\n"
                "Installez-le avec : pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Impossible d'exporter les données :\n{e}")


# Test autonome
if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication
    
    app = QApplication(sys.argv)
    dialog = OperateursInactifsDialog()
    dialog.show()
    sys.exit(app.exec_())