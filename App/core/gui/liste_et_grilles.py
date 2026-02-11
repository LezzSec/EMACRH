from PyQt5.QtWidgets import (
    QApplication, QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem,
    QLabel, QMessageBox, QInputDialog, QDialogButtonBox, QListWidget, QListWidgetItem,
    QLineEdit, QFileDialog, QAbstractItemView, QComboBox, QWidget, QHeaderView
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont
import sys

# ✅ OPTIMISATION : Import paresseux de pandas (uniquement si export Excel nécessaire)
# import pandas as pd  # Déplacé dans les fonctions qui l'utilisent
from datetime import datetime, timedelta
from core.gui.historique import HistoriqueDialog
from core.db.query_executor import QueryExecutor
from .besoin_poste_dialog import BesoinPosteDialog
from core.services.logger import log_hist
from core.utils.logging_config import get_logger
logger = get_logger(__name__)
try:
    from core.gui.ui_theme import EmacButton, get_current_theme
    from core.gui.emac_ui_kit import add_custom_title_bar, show_error_message
    THEME_AVAILABLE = True
except ImportError:
    THEME_AVAILABLE = False
    show_error_message = None


class GrillesDialog(QDialog):

    SUMMARY_ROWS = [
        "Niveau 1",
        "Niveau 2",
        "Niveau 3",
        "Niveau 4",
        "Nb total d'opérateurs au poste",
        "Total des niveaux 3 et 4",
        "Besoins par poste",
    ]

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Grilles de Polyvalence")
        self.setGeometry(100, 80, 1400, 800)

        self.is_editable = False
        self.modified_cells = set()
        self._is_processing_change = False

        self._setup_theme_colors()

        # Layout principal avec marges nulles
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Barre de titre personnalisée
        title_bar = add_custom_title_bar(self, "Grilles de Polyvalence")
        main_layout.addWidget(title_bar)

        # Widget de contenu
        content_widget = QWidget()
        self.layout = QVBoxLayout(content_widget)
        self.layout.setContentsMargins(12, 12, 12, 12)
        self.layout.setSpacing(8)

        # En-tête compact
        if THEME_AVAILABLE:
            header_label = QLabel("Grilles de Polyvalence")
            header_label.setStyleSheet("font-size: 18px; font-weight: bold;")
            self.layout.addWidget(header_label)
        else:
            header = QLabel("Grilles de Polyvalence")
            header.setStyleSheet("font-size: 18px; font-weight: bold;")
            self.layout.addWidget(header)

        # Barre d'outils unifiée (tout en une ligne)
        self.add_unified_toolbar()

        # Table principale (sans carte pour économiser l'espace)
        self.main_table = QTableWidget()
        self.main_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.main_table.cellChanged.connect(self.on_cell_changed)
        if THEME_AVAILABLE:
            self._style_table()
        self.layout.addWidget(self.main_table, 1)  # stretch factor = 1 pour occuper tout l'espace

        # Tri désactivé (on utilise des boutons)
        self.main_table.setSortingEnabled(False)
        self.sort_column = None
        self.sort_order = Qt.AscendingOrder

        # Infos niveaux compactes en bas
        self.add_compact_level_info()

        # Ajouter le widget de contenu au layout principal
        main_layout.addWidget(content_widget)

        # Données
        self.operateurs = []
        self.postes = []
        self.load_data()

    def _setup_theme_colors(self):
        """Définit les couleurs adaptées au thème actuel (clair ou sombre)."""
        if THEME_AVAILABLE:
            ThemeCls = get_current_theme()
            # Couleurs pour lignes de synthèse (non éditables)
            self.color_synthesis_bg = QColor(ThemeCls.BG_CARD)
            self.color_synthesis_text = QColor(ThemeCls.TXT)
            # Couleurs pour ligne "Besoins par poste" (éditable)
            self.color_besoins_bg = QColor(ThemeCls.BG_ELEV)
            self.color_besoins_text = QColor(ThemeCls.TXT)
        else:
            # Fallback si thème non disponible (mode clair par défaut)
            self.color_synthesis_bg = QColor(211, 211, 211)  # lightGray
            self.color_synthesis_text = QColor(17, 24, 39)   # texte sombre
            self.color_besoins_bg = QColor(255, 255, 255)    # white
            self.color_besoins_text = QColor(17, 24, 39)

    # ----------------- Filtrage -----------------
    def show_filter_dialog(self):
        """Affiche le dialogue de sélection rapide pour filtrer"""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QListWidget, QListWidgetItem, QLineEdit

        dialog = QDialog(self)
        dialog.setWindowTitle("Filtrer la grille")
        dialog.setMinimumWidth(700)
        dialog.setMinimumHeight(600)

        layout = QVBoxLayout(dialog)

        # Deux colonnes : Opérateurs et Postes
        content_layout = QHBoxLayout()

        # === COLONNE GAUCHE : OPÉRATEURS ===
        ops_widget = QWidget()
        ops_layout = QVBoxLayout(ops_widget)

        ops_header = QLabel("👥 Opérateurs")
        ops_header.setStyleSheet("font-weight: bold; font-size: 14px;")
        ops_layout.addWidget(ops_header)

        # Barre de recherche opérateurs
        self.ops_search = QLineEdit()
        self.ops_search.setPlaceholderText("🔍 Rechercher un opérateur...")
        ops_layout.addWidget(self.ops_search)

        # Liste multi-sélection des opérateurs
        self.ops_list = QListWidget()
        self.ops_list.setSelectionMode(QListWidget.MultiSelection)

        for idx, (op_id, op_name) in enumerate(self.operateurs):
            item = QListWidgetItem(op_name)
            item.setData(Qt.UserRole, idx)  # Stocker l'index de ligne
            # Pré-sélectionner si visible
            if not self.main_table.isRowHidden(idx):
                item.setSelected(True)
            self.ops_list.addItem(item)

        ops_layout.addWidget(self.ops_list)


        content_layout.addWidget(ops_widget)

        # === COLONNE DROITE : POSTES ===
        postes_widget = QWidget()
        postes_layout = QVBoxLayout(postes_widget)

        postes_header = QLabel("📍 Postes")
        postes_header.setStyleSheet("font-weight: bold; font-size: 14px;")
        postes_layout.addWidget(postes_header)

        # Barre de recherche postes
        self.postes_search = QLineEdit()
        self.postes_search.setPlaceholderText("🔍 Rechercher un poste...")
        postes_layout.addWidget(self.postes_search)

        # Liste multi-sélection des postes
        self.postes_list = QListWidget()
        self.postes_list.setSelectionMode(QListWidget.MultiSelection)

        for idx, (poste_id, poste_code) in enumerate(self.postes):
            item = QListWidgetItem(poste_code)
            item.setData(Qt.UserRole, idx)  # Stocker l'index de colonne
            # Pré-sélectionner si visible
            if not self.main_table.isColumnHidden(idx):
                item.setSelected(True)
            self.postes_list.addItem(item)

        postes_layout.addWidget(self.postes_list)

        content_layout.addWidget(postes_widget)

        layout.addLayout(content_layout)

        # Boutons de validation
        buttons_layout = QHBoxLayout()
        btn_apply = QPushButton("Appliquer")
        btn_apply.setStyleSheet("background-color: #27ae60; color: white; padding: 8px; font-weight: bold;")
        btn_cancel = QPushButton("Annuler")
        btn_cancel.setStyleSheet("padding: 8px;")
        buttons_layout.addStretch()
        buttons_layout.addWidget(btn_apply)
        buttons_layout.addWidget(btn_cancel)
        layout.addLayout(buttons_layout)


        # Note : pas de sélection automatique croisée pour éviter la confusion
        # L'utilisateur sélectionne manuellement ce qu'il veut voir

        # Connexions recherche
        self.ops_search.textChanged.connect(lambda text: self.filter_list(self.ops_list, text))
        self.postes_search.textChanged.connect(lambda text: self.filter_list(self.postes_list, text))

        # Connexions validation
        btn_apply.clicked.connect(lambda: self.apply_filters_quick(dialog))
        btn_cancel.clicked.connect(dialog.reject)

        dialog.exec_()

    def filter_list(self, list_widget, search_text):
        """Filtre une liste selon le texte de recherche"""
        for i in range(list_widget.count()):
            item = list_widget.item(i)
            item.setHidden(search_text.lower() not in item.text().lower())

    def inverse_selection(self, list_widget):
        """Inverse la sélection d'une liste"""
        for i in range(list_widget.count()):
            item = list_widget.item(i)
            if not item.isHidden():  # Ne pas toucher aux éléments cachés par la recherche
                item.setSelected(not item.isSelected())

    def auto_select_postes_for_operators(self):
        """Sélectionne automatiquement les postes où les opérateurs sélectionnés ont des compétences"""
        if not hasattr(self, 'ops_list') or not hasattr(self, 'postes_list'):
            return

        # Récupérer les indices des opérateurs sélectionnés
        selected_op_rows = [item.data(Qt.UserRole) for item in self.ops_list.selectedItems()]

        if not selected_op_rows:
            return

        # Trouver tous les postes où au moins un opérateur sélectionné a une compétence
        postes_to_select = set()

        for row_idx in selected_op_rows:
            for col_idx in range(self.main_table.columnCount()):
                # Vérifier si l'opérateur a un niveau dans ce poste
                item = self.main_table.item(row_idx, col_idx)
                if item and item.text().strip() and item.text().strip().isdigit():
                    postes_to_select.add(col_idx)

        # Sélectionner ces postes dans la liste
        for i in range(self.postes_list.count()):
            item = self.postes_list.item(i)
            col_idx = item.data(Qt.UserRole)
            if col_idx in postes_to_select:
                item.setSelected(True)

    def auto_select_operators_for_postes(self):
        """Sélectionne automatiquement les opérateurs qui ont des compétences dans les postes sélectionnés"""
        if not hasattr(self, 'ops_list') or not hasattr(self, 'postes_list'):
            return

        # Récupérer les indices des postes sélectionnés
        selected_poste_cols = [item.data(Qt.UserRole) for item in self.postes_list.selectedItems()]

        if not selected_poste_cols:
            return

        # Trouver tous les opérateurs qui ont une compétence dans au moins un des postes sélectionnés
        operators_to_select = set()

        for col_idx in selected_poste_cols:
            for row_idx in range(len(self.operateurs)):
                # Vérifier si l'opérateur a un niveau dans ce poste
                item = self.main_table.item(row_idx, col_idx)
                if item and item.text().strip() and item.text().strip().isdigit():
                    operators_to_select.add(row_idx)

        # Sélectionner ces opérateurs dans la liste
        for i in range(self.ops_list.count()):
            item = self.ops_list.item(i)
            row_idx = item.data(Qt.UserRole)
            if row_idx in operators_to_select:
                item.setSelected(True)

    def apply_filters_quick(self, dialog):
        """Applique les filtres depuis les listes de sélection"""
        n_ops = len(self.operateurs)

        # Récupérer les sélections
        selected_op_rows = [item.data(Qt.UserRole) for item in self.ops_list.selectedItems()]
        selected_poste_cols = [item.data(Qt.UserRole) for item in self.postes_list.selectedItems()]

        # Cas 1 : Si AUCUN opérateur n'est sélectionné mais des postes sont sélectionnés
        # → Afficher automatiquement UNIQUEMENT les opérateurs qui ont des compétences dans ces postes
        if not selected_op_rows and selected_poste_cols:
            operators_with_skills = set()
            for col_idx in selected_poste_cols:
                for row_idx in range(n_ops):
                    item = self.main_table.item(row_idx, col_idx)
                    if item and item.text().strip() and item.text().strip().isdigit():
                        operators_with_skills.add(row_idx)

            # Masquer tous les opérateurs
            for row in range(n_ops):
                self.main_table.setRowHidden(row, True)

            # Afficher uniquement ceux avec compétences
            for row_idx in operators_with_skills:
                self.main_table.setRowHidden(row_idx, False)

        # Cas 2 : Des opérateurs sont sélectionnés (avec ou sans postes)
        # → Afficher uniquement les opérateurs sélectionnés
        else:
            # Masquer tous les opérateurs
            for row in range(n_ops):
                self.main_table.setRowHidden(row, True)

            # Afficher les opérateurs sélectionnés
            for row_idx in selected_op_rows:
                self.main_table.setRowHidden(row_idx, False)

        # Gestion des colonnes de postes
        if selected_poste_cols:
            # Si des postes sont sélectionnés, afficher uniquement ceux-là
            for col in range(len(self.postes)):
                self.main_table.setColumnHidden(col, True)
            for col_idx in selected_poste_cols:
                self.main_table.setColumnHidden(col_idx, False)
        elif selected_op_rows:
            # Si des opérateurs sont sélectionnés sans postes spécifiques,
            # afficher uniquement les postes où ces opérateurs ont des compétences
            postes_with_skills = set()
            for row_idx in selected_op_rows:
                for col_idx in range(len(self.postes)):
                    item = self.main_table.item(row_idx, col_idx)
                    if item and item.text().strip() and item.text().strip().isdigit():
                        postes_with_skills.add(col_idx)

            # Masquer tous les postes
            for col in range(len(self.postes)):
                self.main_table.setColumnHidden(col, True)

            # Afficher uniquement les postes avec compétences
            for col_idx in postes_with_skills:
                self.main_table.setColumnHidden(col_idx, False)
        else:
            # Si rien n'est sélectionné, afficher TOUS les postes
            for col in range(len(self.postes)):
                self.main_table.setColumnHidden(col, False)

        # Recalculer les statistiques
        self.update_statistics()

        dialog.accept()

    def reset_filters(self):
        """Réaffiche toutes les lignes et colonnes cachées."""
        for row in range(self.main_table.rowCount()):
            self.main_table.setRowHidden(row, False)
        for col in range(self.main_table.columnCount()):
            self.main_table.setColumnHidden(col, False)

        self.update_statistics()

    # ----------------- UI haut -----------------
    def add_unified_toolbar(self):
        """Barre d'outils unifiée avec tous les boutons sur une seule ligne"""
        toolbar_layout = QHBoxLayout()
        toolbar_layout.setSpacing(6)

        if THEME_AVAILABLE:
            # Groupe Édition
            add_button = EmacButton("+ Ajouter", variant='primary')
            add_button.setToolTip("Ajouter une colonne (poste) ou une ligne (opérateur)")
            add_button.clicked.connect(self.add_data)
            toolbar_layout.addWidget(add_button)

            remove_button = EmacButton("− Masquer", variant='ghost')
            remove_button.setToolTip("Masquer une colonne ou une ligne")
            remove_button.clicked.connect(self.remove_data)
            toolbar_layout.addWidget(remove_button)

            self.edit_mode_button = EmacButton("✏️ Éditer", variant='ghost')
            self.edit_mode_button.setToolTip("Activer/Désactiver le mode édition")
            self.edit_mode_button.clicked.connect(self.toggle_edit_mode)
            toolbar_layout.addWidget(self.edit_mode_button)

            duplicate_button = EmacButton("📄 Dupliquer", variant='ghost')
            duplicate_button.setToolTip("Dupliquer une colonne ou une ligne")
            duplicate_button.clicked.connect(self.duplicate_data)
            toolbar_layout.addWidget(duplicate_button)

            # Séparateur visuel
            separator = QLabel("|")
            separator.setStyleSheet("color: #d1d5db; font-size: 18px; padding: 0 8px;")
            toolbar_layout.addWidget(separator)

            # Groupe Filtres
            filter_button = EmacButton("🔍 Filtrer", variant='ghost')
            filter_button.setToolTip("Sélectionner les opérateurs et/ou postes à afficher")
            filter_button.clicked.connect(self.show_filter_dialog)
            toolbar_layout.addWidget(filter_button)

            reset_filter_button = EmacButton("↻ Réinitialiser", variant='ghost')
            reset_filter_button.setToolTip("Réinitialiser les filtres et afficher toutes les données")
            reset_filter_button.clicked.connect(self.reset_filters)
            toolbar_layout.addWidget(reset_filter_button)

            toolbar_layout.addStretch()

            # Groupe Actions
            export_button = EmacButton("📤 Exporter", variant='ghost')
            export_button.setToolTip("Exporter l'état visuel actuel (avec filtres appliqués)")
            export_button.clicked.connect(self.export_data)
            toolbar_layout.addWidget(export_button)
        else:
            # Version sans thème
            add_button = QPushButton("+ Ajouter")
            add_button.clicked.connect(self.add_data)
            toolbar_layout.addWidget(add_button)

            remove_button = QPushButton("− Masquer")
            remove_button.clicked.connect(self.remove_data)
            toolbar_layout.addWidget(remove_button)

            self.edit_mode_button = QPushButton("✏️ Éditer")
            self.edit_mode_button.clicked.connect(self.toggle_edit_mode)
            toolbar_layout.addWidget(self.edit_mode_button)

            duplicate_button = QPushButton("📄 Dupliquer")
            duplicate_button.clicked.connect(self.duplicate_data)
            toolbar_layout.addWidget(duplicate_button)

            toolbar_layout.addWidget(QLabel("|"))

            filter_button = QPushButton("🔍 Filtrer")
            filter_button.clicked.connect(self.show_filter_dialog)
            toolbar_layout.addWidget(filter_button)

            reset_filter_button = QPushButton("↻ Réinitialiser")
            reset_filter_button.clicked.connect(self.reset_filters)
            toolbar_layout.addWidget(reset_filter_button)

            toolbar_layout.addStretch()

            export_button = QPushButton("📤 Exporter")
            export_button.clicked.connect(self.export_data)
            toolbar_layout.addWidget(export_button)

        self.layout.addLayout(toolbar_layout)

    def add_compact_level_info(self):
        """Section compacte d'informations sur les niveaux de compétence"""
        info_layout = QHBoxLayout()
        info_layout.setSpacing(12)

        info_label = QLabel("<b>Niveaux :</b>")
        info_layout.addWidget(info_label)

        # Niveaux en ligne
        level_descriptions = [
            ("<b>N1</b> : &lt;80% (nouveau/absent 12 mois)", "#dc2626"),
            ("<b>N2</b> : ≥80% (formé, autonome)", "#d97706"),
            ("<b>N3</b> : &gt;90% (formateur)", "#059669"),
            ("<b>N4</b> : &gt;90% (leader/polyvalent ≥3 postes)", "#0369a1"),
        ]

        for desc, color in level_descriptions:
            label = QLabel(desc)
            label.setStyleSheet(f"color: {color}; font-size: 11px;")
            info_layout.addWidget(label)

        info_layout.addStretch()
        self.layout.addLayout(info_layout)

    def _style_table(self):
        """Applique un style moderne à la table"""
        if not THEME_AVAILABLE:
            return

        ThemeCls = get_current_theme()

        # Style de la table
        self.main_table.setStyleSheet(f"""
            QTableWidget {{
                background: {ThemeCls.BG_TABLE};
                border: 1px solid {ThemeCls.BDR};
                border-radius: 10px;
                gridline-color: {ThemeCls.BDR};
            }}
            QTableWidget::item {{
                padding: 6px;
                border: none;
            }}
            QTableWidget::item:selected {{
                background: {ThemeCls.PRI};
                color: white;
            }}
            QHeaderView::section {{
                background: {ThemeCls.BG_ELEV};
                color: {ThemeCls.TXT};
                padding: 8px;
                border: 1px solid {ThemeCls.BDR};
                font-weight: 600;
                font-size: 13px;
            }}
            QHeaderView::section:hover {{
                background: {ThemeCls.BDR};
            }}
        """)

        # Ajuster les en-têtes
        self.main_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.main_table.verticalHeader().setDefaultSectionSize(32)
        self.main_table.horizontalHeader().setDefaultSectionSize(70)

    # ----------------- Filtres -----------------
    def show_multiselect_dialog(self, title, items):
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setMinimumWidth(300)

        layout = QVBoxLayout()
        list_widget = QListWidget()
        list_widget.setSelectionMode(QListWidget.MultiSelection)
        for item in items:
            list_item = QListWidgetItem(item)
            list_widget.addItem(list_item)

        layout.addWidget(list_widget)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        dialog.setLayout(layout)

        if dialog.exec_() == QDialog.Accepted:
            return [item.text() for item in list_widget.selectedItems()]
        return []


    # ----------------- Édition -----------------
    def on_cell_changed(self, row, column):
        """
        Version finale propre, sans erreurs d'indentation
        """
        if not self.is_editable:
            return

        item = self.main_table.item(row, column)
        if not item:
            return

        # Déconnecter le signal pour éviter la boucle infinie
        try:
            self.main_table.cellChanged.disconnect(self.on_cell_changed)
        except Exception:
            pass

        try:
            import json
            from datetime import datetime, timedelta

            n_ops = len(self.operateurs)

            # Ignorer les lignes de total
            if row >= n_ops:
                return

            operateur_id = self.operateurs[row][0]
            operateur_nom = self.operateurs[row][1]
            poste_id = self.postes[column][0]
            poste_code = self.postes[column][1]
            value = item.text().strip()

            if value and not value.isdigit():
                QMessageBox.warning(self, "Erreur", "Veuillez entrer un nombre valide")
                return

            # ==============================
            # 1️⃣ LECTURE ancienne polyvalence
            # ==============================
            old_niveau = None
            old_date_eval = None
            old_date_next = None

            try:
                result = QueryExecutor.fetch_one("""
                    SELECT niveau, date_evaluation, prochaine_evaluation
                    FROM polyvalence
                    WHERE operateur_id = %s AND poste_id = %s
                """, (operateur_id, poste_id))
                if result:
                    old_niveau, old_date_eval, old_date_next = result

            except Exception as e:
                logger.error(f"Erreur lecture ancienne valeur : {e}")

            # ==============================
            # 2️⃣ MODIFICATION de la polyvalence
            # ==============================
            action = None
            new_niveau_int = None

            try:
                if value == "":
                    # Suppression
                    QueryExecutor.execute_write(
                        "DELETE FROM polyvalence WHERE operateur_id = %s AND poste_id = %s",
                        (operateur_id, poste_id)
                    )
                    action = 'DELETE'

                else:
                    new_niveau_int = int(value)

                    # Archivage si modification
                    if old_niveau is not None and old_niveau != new_niveau_int:
                        try:
                            QueryExecutor.execute_write("""
                                INSERT INTO historique_polyvalence
                                (operateur_id, poste_id, action_type,
                                 ancien_niveau, nouveau_niveau,
                                 ancienne_date_evaluation, nouvelle_date_evaluation,
                                 commentaire, date_action)
                                VALUES (%s, %s, 'IMPORT_MANUEL',
                                        %s, %s,
                                        %s, NULL,
                                        'Ancienne polyvalence archivée lors de modification depuis la grille',
                                        NOW())
                            """, (
                                operateur_id,
                                poste_id,
                                old_niveau,
                                new_niveau_int,
                                old_date_eval
                            ))
                        except Exception as arch_err:
                            logger.warning(f"Erreur archivage : {arch_err}")

                    # Calculer la prochaine évaluation selon le niveau
                    from datetime import date, timedelta
                    if new_niveau_int == 1:
                        jours = 30  # 1 mois
                    elif new_niveau_int == 2:
                        jours = 30  # 1 mois
                    elif new_niveau_int in [3, 4]:
                        jours = 3650  # 10 ans
                    else:
                        jours = 30  # Par défaut 1 mois

                    prochaine_eval = date.today() + timedelta(days=jours)

                    # Insert ou update
                    if old_niveau is None:
                        QueryExecutor.execute_write("""
                            INSERT INTO polyvalence (operateur_id, poste_id, niveau, date_evaluation, prochaine_evaluation)
                            VALUES (%s, %s, %s, CURDATE(), %s)
                        """, (operateur_id, poste_id, new_niveau_int, prochaine_eval))
                        action = 'INSERT'
                    else:
                        QueryExecutor.execute_write("""
                            UPDATE polyvalence SET niveau = %s, date_evaluation = CURDATE(), prochaine_evaluation = %s
                            WHERE operateur_id = %s AND poste_id = %s
                        """, (new_niveau_int, prochaine_eval, operateur_id, poste_id))
                        action = 'UPDATE'

            except Exception as e:
                logger.exception(f"Erreur modification: {e}")
                if show_error_message:
                    show_error_message(self, "Erreur", "Erreur lors de la modification", e)
                else:
                    QMessageBox.critical(self, "Erreur", "Erreur lors de la modification. Contactez l'administrateur.")
                return

            # ==============================
            # 3️⃣ LOGGING dans l'historique
            # ==============================
            try:
                # Récupérer des informations supplémentaires
                matricule = None
                atelier_nom = None
                utilisateur = None

                try:
                    # Matricule de l'opérateur
                    result = QueryExecutor.fetch_one("SELECT matricule FROM personnel WHERE id = %s", (operateur_id,), dictionary=True)
                    if result:
                        matricule = result.get('matricule')

                    # Atelier du poste
                    result = QueryExecutor.fetch_one("""
                        SELECT a.nom as atelier_nom
                        FROM postes p
                        LEFT JOIN atelier a ON p.atelier_id = a.id
                        WHERE p.id = %s
                    """, (poste_id,), dictionary=True)
                    if result:
                        atelier_nom = result.get('atelier_nom')

                    # Utilisateur connecté
                    from core.services.auth_service import get_current_user
                    current_user = get_current_user()
                    if current_user:
                        utilisateur = current_user.get('username') or f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip()
                except Exception as e:
                    logger.warning(f"Erreur récupération infos supplémentaires: {e}")

                # Construction du JSON enrichi
                base_info = {
                    "operateur": operateur_nom,
                    "matricule": matricule,
                    "poste": poste_code,
                    "atelier": atelier_nom,
                    "source": "Grille de polyvalence"
                }

                if action == 'DELETE':
                    description = json.dumps({
                        **base_info,
                        "niveau_supprime": old_niveau,
                        "date_eval_supprimee": str(old_date_eval) if old_date_eval else None,
                        "type": "suppression"
                    }, ensure_ascii=False)

                elif action == 'INSERT':
                    description = json.dumps({
                        **base_info,
                        "niveau": new_niveau_int,
                        "date_evaluation": str(date.today()),
                        "prochaine_evaluation": str(prochaine_eval),
                        "type": "ajout"
                    }, ensure_ascii=False)

                else:  # UPDATE
                    description = json.dumps({
                        **base_info,
                        "changes": {
                            "niveau": {"old": old_niveau, "new": new_niveau_int}
                        },
                        "ancienne_date_eval": str(old_date_eval) if old_date_eval else None,
                        "nouvelle_date_eval": str(date.today()),
                        "prochaine_evaluation": str(prochaine_eval),
                        "type": "modification"
                    }, ensure_ascii=False)

                # Insert dans l'historique
                QueryExecutor.execute_write("""
                    INSERT INTO historique (date_time, action, operateur_id, poste_id, description, utilisateur)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (datetime.now(), action, operateur_id, poste_id, description, utilisateur))

                # Message final
                if action == 'DELETE':
                    QMessageBox.information(self, "Supprimé",
                        f"Compétence supprimée\n\n{operateur_nom} - {poste_code}")

                elif action == 'INSERT':
                    QMessageBox.information(self, "Ajouté",
                        f"Nouvelle compétence ajoutée\n\n{operateur_nom} - {poste_code}\nNiveau : {new_niveau_int}")

                else:
                    QMessageBox.information(self, "Mis à jour",
                        f"Compétence modifiée\n\n{operateur_nom} - {poste_code}\n"
                        f"{old_niveau} → {new_niveau_int}")

            except Exception as e:
                logger.warning(f"Erreur logging : {e}")
                QMessageBox.warning(self, "Attention",
                    f"Modification OK mais erreur dans l'historique :\n{e}")

        finally:
            # Reconnecter le signal
            try:
                self.main_table.cellChanged.connect(self.on_cell_changed)
            except Exception:
                pass


    def reload_cell(self, row, column):
        """Recharge une cellule opérateur spécifique avec sa valeur depuis la base de données."""
        try:
            operateur_id = self.operateurs[row][0]
            poste_id = self.postes[column][0]

            result = QueryExecutor.fetch_one(
                "SELECT niveau FROM polyvalence WHERE operateur_id = %s AND poste_id = %s",
                (operateur_id, poste_id),
                dictionary=True
            )

            self.main_table.blockSignals(True)
            if result:
                value = result.get("niveau")
                self.main_table.setItem(row, column, QTableWidgetItem(str(value)))
            else:
                self.main_table.setItem(row, column, QTableWidgetItem(""))
            self.main_table.blockSignals(False)

        except Exception as e:
            logger.error(f"Erreur lors du rechargement de la cellule : {e}")

    # ----------------- Statistiques -----------------
    def update_statistics(self):
        """Remplit les 7 lignes de synthèse en bas, alignées aux colonnes des postes (sans toucher aux besoins)."""
        try:
            self.main_table.blockSignals(True)

            n_ops = len(self.operateurs)
            row_lvl1 = n_ops + 0
            row_lvl2 = n_ops + 1
            row_lvl3 = n_ops + 2
            row_lvl4 = n_ops + 3
            row_total_ops = n_ops + 4
            row_total_34  = n_ops + 5
            row_besoins   = n_ops + 6  # éditable, on NE l'écrase PAS

            for col in range(self.main_table.columnCount()):
                niveaux = []
                any_value = 0

                for row in range(n_ops):
                    item = self.main_table.item(row, col)
                    txt = item.text().strip() if item else ""
                    if txt.isdigit():
                        niveaux.append(int(txt))
                        any_value += 1

                c1 = niveaux.count(1)
                c2 = niveaux.count(2)
                c3 = niveaux.count(3)
                c4 = niveaux.count(4)
                total_34 = c3 + c4

                def _set(r, val):
                    it = self.main_table.item(r, col)
                    if not it:
                        it = QTableWidgetItem("")
                        self.main_table.setItem(r, col, it)
                    it.setText("" if (val is None or val == 0) else str(val))
                    it.setTextAlignment(Qt.AlignCenter)

                # Option styling initial (col 0 pour coupure visuelle – conservé)
                if col == 0:
                    _set(row_lvl1, None)
                else:
                    _set(row_lvl1, c1)

                _set(row_lvl2, c2)
                _set(row_lvl3, c3)
                _set(row_lvl4, c4)
                _set(row_total_ops, any_value)
                _set(row_total_34, total_34)
                # row_besoins: ne pas modifier

            self.main_table.blockSignals(False)

        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour des statistiques : {e}")
            self.main_table.blockSignals(False)

    # ----------------- Chargement -----------------
    def load_data(self):
        """Charge opérateurs/postes + prépare 7 lignes de synthèse en bas + remplit la ligne Besoins.

        ✅ OPTIMISÉ : Utilise 3 requêtes séparées au lieu d'un CROSS JOIN massif
        """
        try:
            # ✅ REQUÊTE 1 : Liste des opérateurs actifs (RAPIDE)
            operateurs_rows = QueryExecutor.fetch_all("""
                SELECT id, nom, prenom
                FROM personnel
                WHERE statut = 'ACTIF'
                  AND matricule IS NOT NULL
                  AND matricule != ''
                ORDER BY nom, prenom
            """, dictionary=True)

            # ✅ REQUÊTE 2 : Liste des postes visibles (RAPIDE)
            # Exclure la colonne "PRODUCTION" qui ne contient pas de vraies données de polyvalence
            postes_rows = QueryExecutor.fetch_all("""
                SELECT id, poste_code
                FROM postes
                WHERE visible = 1
                  AND poste_code != 'PRODUCTION'
                ORDER BY poste_code
            """, dictionary=True)

            # ✅ REQUÊTE 3 : Seulement les polyvalences existantes (RAPIDE - pas de CROSS JOIN)
            polyvalences_rows = QueryExecutor.fetch_all("""
                SELECT operateur_id, poste_id, niveau
                FROM polyvalence
                WHERE operateur_id IN (
                    SELECT id FROM personnel
                    WHERE statut = 'ACTIF'
                      AND matricule IS NOT NULL
                      AND matricule != ''
                )
                AND poste_id IN (
                    SELECT id FROM postes WHERE visible = 1
                )
            """, dictionary=True)

            # --- Récupération des besoins_postes ---
            besoins_rows = QueryExecutor.fetch_all("SELECT id, besoins_postes FROM postes WHERE visible = 1", dictionary=True)
            besoins_by_id = {r["id"]: r.get("besoins_postes", "") for r in besoins_rows}

            # ✅ OPTIMISATION : Construire les structures de données depuis les 3 requêtes
            # Créer un dictionnaire de polyvalences indexé par (operateur_id, poste_id)
            polyvalences_dict = {}
            for pv in polyvalences_rows:
                key = (pv['operateur_id'], pv['poste_id'])
                polyvalences_dict[key] = pv['niveau']

            # Construire la liste des opérateurs
            # ⚠️ IMPORTANT: On ne construit pas encore self.operateurs ici
            # Elle sera construite plus tard dans l'ordre TRIÉ pour correspondre aux lignes du tableau
            operateurs_dict = {}
            for op in operateurs_rows:
                nom_complet = f"{op['nom']} {op['prenom']}".strip()
                operateurs_dict[nom_complet] = {'id': op['id'], 'postes': {}}

            # Construire la liste des postes
            self.postes = [(p['id'], p['poste_code']) for p in postes_rows]

            # Remplir les niveaux de polyvalence pour chaque opérateur
            for nom_complet, data in operateurs_dict.items():
                op_id = data['id']
                for poste_id, poste_code in self.postes:
                    key = (op_id, poste_id)
                    niveau = polyvalences_dict.get(key, '')
                    data['postes'][poste_code] = niveau

            # --- dimensions : #ops + 7 lignes de synthèse ---
            n_ops = len(operateurs_dict)
            SUMMARY_ROWS = [
                "Niveau 1",
                "Niveau 2",
                "Niveau 3",
                "Niveau 4",
                "Nb total d'opérateurs au poste",
                "Total des niveaux 3 et 4",
                "Besoins par poste",
            ]

            self.main_table.blockSignals(True)
            self.main_table.setRowCount(n_ops + len(self.SUMMARY_ROWS))
            self.main_table.setColumnCount(len(self.postes))

            # En-têtes colonnes / lignes
            self.main_table.setHorizontalHeaderLabels([poste[1] for poste in self.postes])
            self.main_table.setVerticalHeaderLabels(
                sorted(operateurs_dict.keys()) + self.SUMMARY_ROWS
            )

            # ✅ CORRECTION BUG DÉCALAGE: Construire self.operateurs dans l'ordre TRIÉ
            # pour que les index correspondent aux lignes du tableau
            sorted_operateurs = sorted(operateurs_dict.items())  # Trié par nom_complet
            self.operateurs = [(data['id'], nom_complet) for nom_complet, data in sorted_operateurs]

            # Remplissage des cellules opérateurs
            for row_idx, (nom_complet, data) in enumerate(sorted_operateurs):
                for col_idx, (_, poste_code) in enumerate(self.postes):
                    niveau = data['postes'].get(poste_code, '')
                    item = QTableWidgetItem(str(niveau))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.main_table.setItem(row_idx, col_idx, item)

            # 💥 Style : 6 premières lignes de synthèse non éditables (avec couleurs thème)
            start_row = n_ops  # première ligne de synthèse
            for r in range(start_row, start_row + len(self.SUMMARY_ROWS) - 1):
                for c in range(self.main_table.columnCount()):
                    it = self.main_table.item(r, c)
                    if not it:
                        it = QTableWidgetItem("")
                        self.main_table.setItem(r, c, it)
                    it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                    # 💥 Couleurs adaptées au thème
                    it.setBackground(self.color_synthesis_bg)
                    it.setForeground(self.color_synthesis_text)

            # 💥 Ligne "Besoins par poste" éditable (avec couleurs thème)
            besoins_row = start_row + len(self.SUMMARY_ROWS) - 1
            for c in range(self.main_table.columnCount()):
                it = self.main_table.item(besoins_row, c)
                if not it:
                    it = QTableWidgetItem("")
                    self.main_table.setItem(besoins_row, c, it)
                it.setFlags((it.flags() | Qt.ItemIsEditable))
                # 💥 Couleurs adaptées au thème pour la ligne éditable
                it.setBackground(self.color_besoins_bg)
                it.setForeground(self.color_besoins_text)
                it.setTextAlignment(Qt.AlignCenter)

            # --- Affichage des besoins_postes alignés aux colonnes ---
            for col_idx, (poste_id, _code) in enumerate(self.postes):
                val = besoins_by_id.get(poste_id, "")
                it = self.main_table.item(besoins_row, col_idx)
                it.setText("" if val in (None, "") else str(val))

            self.main_table.blockSignals(False)

            # Calculs
            self.update_statistics()

        except Exception as e:
            logger.exception(f"Erreur chargement donnees: {e}")
            if show_error_message:
                show_error_message(self, "Erreur", "Erreur lors du chargement des données", e)
            else:
                QMessageBox.critical(self, "Erreur", "Erreur lors du chargement des données. Contactez l'administrateur.")

    # ----------------- Mode édition -----------------
    def toggle_edit_mode(self):
        """Active/désactive le mode édition avec sauvegarde automatique."""
        self.is_editable = not self.is_editable

        if self.is_editable:
            # Mode édition activé
            self.main_table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)

            # Changer l'apparence du bouton
            if THEME_AVAILABLE and hasattr(self, 'edit_mode_button'):
                self.edit_mode_button.setProperty('class', 'primary')
                self.edit_mode_button.style().unpolish(self.edit_mode_button)
                self.edit_mode_button.style().polish(self.edit_mode_button)
                self.edit_mode_button.setText("✏️ Mode Édition (Actif)")

            QMessageBox.information(self, "Mode Édition",
                "✅ Mode édition activé\n\n"
                "💡 Les modifications seront sauvegardées automatiquement\n"
                "   lorsque vous désactiverez le mode édition.")
        else:
            # Mode édition désactivé - SAUVEGARDER AUTOMATIQUEMENT
            # Restaurer l'apparence du bouton
            if THEME_AVAILABLE and hasattr(self, 'edit_mode_button'):
                self.edit_mode_button.setProperty('class', 'ghost')
                self.edit_mode_button.style().unpolish(self.edit_mode_button)
                self.edit_mode_button.style().polish(self.edit_mode_button)
                self.edit_mode_button.setText("✏️ Mode Édition")

            if self.modified_cells:
                self.save_changes()
            else:
                self.main_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
                QMessageBox.information(self, "Mode Édition",
                    "Mode édition désactivé.\n\nAucune modification à sauvegarder.")
                
    def save_changes(self):
        """Enregistre les modifications ET les log dans l'historique."""
        import json
        from datetime import datetime

        try:
            modifications_count = 0

            for row, col in self.modified_cells:
                if row >= len(self.operateurs) or col >= len(self.postes):
                    continue

                operateur_id = self.operateurs[row][0]
                operateur_nom = self.operateurs[row][1]  # "Nom Prenom"
                poste_id = self.postes[col][0]
                poste_code = self.postes[col][1]  # "0515"

                item = self.main_table.item(row, col)
                new_niveau = item.text() if item else None

                if new_niveau is None or new_niveau.strip() == "":
                    new_niveau = None
                elif not new_niveau.isdigit():
                    QMessageBox.critical(self, "Erreur",
                        f"Valeur incorrecte : '{new_niveau}' dans la cellule ({row + 1}, {col + 1}).")
                    continue

                new_niveau_int = int(new_niveau) if new_niveau else None

                # ✅ NOUVEAU : Vérifier l'ancien niveau
                existing = QueryExecutor.fetch_one("""
                    SELECT niveau FROM polyvalence
                    WHERE operateur_id = %s AND poste_id = %s
                """, (operateur_id, poste_id), dictionary=True)

                if existing:
                    old_niveau = existing.get('niveau')
                    action = 'UPDATE'
                else:
                    old_niveau = None
                    action = 'INSERT'

                # Enregistrer la modification dans polyvalence
                QueryExecutor.execute_write(
                    "REPLACE INTO polyvalence (operateur_id, poste_id, niveau) VALUES (%s, %s, %s)",
                    (operateur_id, poste_id, new_niveau_int)
                )

                # ✅ NOUVEAU : Construire la description JSON pour l'historique
                if action == 'INSERT':
                    description = json.dumps({
                        "operateur": operateur_nom,
                        "poste": poste_code,
                        "niveau": new_niveau_int,
                        "type": "ajout"
                    }, ensure_ascii=False)
                else:  # UPDATE
                    changes = {}
                    if old_niveau != new_niveau_int:
                        changes["niveau"] = {"old": old_niveau, "new": new_niveau_int}

                    description = json.dumps({
                        "operateur": operateur_nom,
                        "poste": poste_code,
                        "changes": changes,
                        "type": "modification"
                    }, ensure_ascii=False)

                # ✅ NOUVEAU : Enregistrer dans l'historique
                try:
                    # Récupérer l'utilisateur connecté
                    utilisateur = None
                    try:
                        from core.services.auth_service import get_current_user
                        current_user = get_current_user()
                        if current_user:
                            utilisateur = current_user.get('username') or f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip()
                    except Exception:
                        pass

                    QueryExecutor.execute_write("""
                        INSERT INTO historique (date_time, action, operateur_id, poste_id, description, utilisateur)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (datetime.now(), action, operateur_id, poste_id, description, utilisateur))
                    modifications_count += 1
                except Exception as e:
                    logger.warning(f"Erreur logging historique : {e}")
                    # On continue même si le log échoue

            self.modified_cells.clear()

            QMessageBox.information(self, "Succès",
                f"{modifications_count} modification(s) enregistrée(s) dans l'historique !")

        except Exception as e:
            QMessageBox.critical(self, "Erreur",
                f"Erreur lors de l'enregistrement : {e}")

    # ----------------- Ajout / Suppression / Duplication -----------------
    def add_data(self):
        from core.db.configbd import DatabaseConnection
        choice, ok = QInputDialog.getItem(self, "Ajouter", "Que voulez-vous ajouter ?", ["Colonne", "Ligne"], 0, False)
        if ok and choice:
            with DatabaseConnection() as connection:
                cursor = connection.cursor(dictionary=True)
                if choice == "Colonne":
                    col_name, name_ok = QInputDialog.getText(self, "Nom de Colonne", "Entrez un nom pour la nouvelle colonne :")
                    if name_ok and col_name:
                        # Vérifier doublon
                        if QueryExecutor.exists("postes", {"poste_code": col_name}):
                            QMessageBox.warning(self, "Attention", f"Le poste '{col_name}' existe déjà.")
                        else:
                            # INSERT poste
                            cursor.execute("INSERT INTO postes (poste_code, visible) VALUES (%s, 1)", (col_name,))

                            # Pop-up besoin (obligatoire en création)
                            dlg = BesoinPosteDialog(parent=self, titre_poste=col_name)
                            if dlg.exec_() != dlg.Accepted:
                                connection.rollback()
                                QMessageBox.information(self, "Création annulée", "Le poste n'a pas été créé.")
                            else:
                                besoin_val = dlg.get_besoin_int_or_none()
                                cursor.execute(
                                    "UPDATE postes SET besoins_postes = %s WHERE poste_code = %s",
                                    (besoin_val, col_name)
                                )
                                connection.commit()
                                self.load_data()
                elif choice == "Ligne":
                    row_name, name_ok = QInputDialog.getText(self, "Nom de Ligne", "Entrez le nom pour la nouvelle ligne (Nom) :")
                    if name_ok and row_name:
                        prenom, ok = QInputDialog.getText(self, "Prénom", "Entrez le prénom de l'opérateur :")
                        if ok:
                            cursor.execute("INSERT INTO personnel (nom, prenom, statut) VALUES (%s, %s, 'ACTIF')", (row_name, prenom))
                            connection.commit()
                            self.load_data()
                cursor.close()

    def remove_data(self):
        from core.db.configbd import DatabaseConnection
        choice, ok = QInputDialog.getItem(self, "Supprimer", "Que voulez-vous supprimer ?", ["Colonne", "Ligne"], 0, False)
        if ok and choice:
            with DatabaseConnection() as connection:
                cursor = connection.cursor(dictionary=True)
                if choice == "Colonne":
                    col_name, name_ok = QInputDialog.getText(self, "Nom de la Colonne", "Entrez le nom du poste à masquer :")
                    if name_ok and col_name:
                        poste = QueryExecutor.fetch_one("SELECT id FROM postes WHERE poste_code = %s AND visible = 1", (col_name,), dictionary=True)
                        if poste:
                            poste_id = poste["id"]
                            cursor.execute("UPDATE postes SET visible = 0 WHERE id = %s", (poste_id,))
                            connection.commit()
                            self.load_data()
                        else:
                            QMessageBox.critical(self, "Erreur", f"Le poste '{col_name}' n'existe pas ou est déjà masqué.")
                elif choice == "Ligne":
                    # Récupérer la liste des opérateurs actifs avec matricule
                    operateurs = QueryExecutor.fetch_all("""
                        SELECT id, nom, prenom, matricule
                        FROM personnel
                        WHERE statut = 'ACTIF'
                        AND matricule IS NOT NULL
                        AND matricule != ''
                        ORDER BY nom, prenom
                    """, dictionary=True)

                    if not operateurs:
                        QMessageBox.information(self, "Information", "Aucun opérateur actif à masquer.")
                        cursor.close()
                        return

                    # Créer une liste de noms pour la sélection
                    operateur_names = []
                    operateur_map = {}
                    for op in operateurs:
                        op_id = op["id"]
                        nom = op["nom"]
                        prenom = op["prenom"]
                        matricule = op["matricule"]

                        display_name = f"{nom} {prenom} ({matricule})"
                        operateur_names.append(display_name)
                        operateur_map[display_name] = op_id

                    # Afficher le dialogue de sélection
                    selected_name, name_ok = QInputDialog.getItem(
                        self,
                        "Masquer un opérateur",
                        "Sélectionnez l'opérateur à masquer :",
                        operateur_names,
                        0,
                        False
                    )

                    if name_ok and selected_name:
                        operateur_id = operateur_map[selected_name]
                        cursor.execute("UPDATE personnel SET statut = 'INACTIF' WHERE id = %s", (operateur_id,))
                        connection.commit()
                        QMessageBox.information(self, "Succès", f"L'opérateur {selected_name} a été masqué.")
                        self.load_data()
                cursor.close()

    def duplicate_data(self):
        from core.db.configbd import DatabaseConnection
        choice, ok = QInputDialog.getItem(self, "Dupliquer", "Que voulez-vous dupliquer ?", ["Colonne", "Ligne"], 0, False)
        if ok and choice:
            with DatabaseConnection() as connection:
                cursor = connection.cursor(dictionary=True)
                if choice == "Colonne":
                    col_name, name_ok = QInputDialog.getText(self, "Nom de la Colonne", "Entrez le nom du poste à dupliquer :")
                    if name_ok and col_name:
                        poste = QueryExecutor.fetch_one("SELECT id, poste_code FROM postes WHERE poste_code = %s AND visible = 1", (col_name,), dictionary=True)
                        if poste:
                            poste_id = poste["id"]
                            poste_code = poste["poste_code"]

                            # Création du nouveau poste (MySQL)
                            cursor.execute(
                                "INSERT INTO postes (poste_code, visible) VALUES (%s, 1)",
                                (f"{poste_code}_copy",)
                            )
                            new_poste_id = cursor.lastrowid

                            # Copier la polyvalence
                            cursor.execute(
                                """
                                INSERT INTO polyvalence (operateur_id, poste_id, niveau)
                                SELECT operateur_id, %s, niveau FROM polyvalence WHERE poste_id = %s
                                """,
                                (new_poste_id, poste_id)
                            )
                            connection.commit()
                            self.load_data()
                        else:
                            QMessageBox.critical(self, "Erreur", f"Le poste '{col_name}' n'existe pas ou est masqué.")

                elif choice == "Ligne":
                    row_name, name_ok = QInputDialog.getText(self, "Nom de la Ligne", "Entrez le nom complet de l'opérateur à dupliquer :")
                    if name_ok and row_name:
                        operateur = QueryExecutor.fetch_one("SELECT id, nom, prenom FROM personnel WHERE CONCAT(nom, ' ', prenom) = %s AND statut = 'ACTIF'", (row_name,), dictionary=True)
                        if operateur:
                            operateur_id = operateur["id"]
                            nom = operateur["nom"]
                            prenom = operateur["prenom"]

                            # Créer le nouvel opérateur (MySQL)
                            cursor.execute(
                                "INSERT INTO personnel (nom, prenom, statut) VALUES (%s, %s, 'ACTIF')",
                                (f"{nom}_copy", prenom)
                            )
                            new_operateur_id = cursor.lastrowid

                            # Copier la polyvalence
                            cursor.execute(
                                """
                                INSERT INTO polyvalence (operateur_id, poste_id, niveau)
                                SELECT %s, poste_id, niveau FROM polyvalence WHERE operateur_id = %s
                                """,
                                (new_operateur_id, operateur_id)
                            )
                            connection.commit()
                            self.load_data()
                        else:
                            QMessageBox.critical(self, "Erreur", f"L'opérateur '{row_name}' n'existe pas ou est inactif.")
                cursor.close()

    # ----------------- Reload & Tri -----------------
    def reload_data(self):
        """ Remet la table dans son état initial, recharge et retrie. """
        self.reset_filters()
        self.load_data()
        self._sort_columns_az()
        self._apply_sort_state()

    def _remember_sort_state(self, column, order):
        self.sort_column = column
        self.sort_order = order

    def _apply_sort_state(self):
        if self.sort_column is not None:
            self.main_table.blockSignals(True)
            self.main_table.sortItems(self.sort_column, self.sort_order)
            self.main_table.blockSignals(False)

    def _sort_columns_az(self):
        header = self.main_table.horizontalHeader()
        col_count = self.main_table.columnCount()
        if col_count <= 1:
            return

        target = sorted(
            range(col_count),
            key=lambda i: self.main_table.horizontalHeaderItem(i).text()
                          if self.main_table.horizontalHeaderItem(i) else ""
        )
        for new_pos, logical_index in enumerate(target):
            current_visual = header.visualIndex(logical_index)
            if current_visual != new_pos:
                header.moveSection(current_visual, new_pos)

    # ----------------- Export -----------------
    def export_data(self):
        """Dialogue moderne d'export avec sélection de l'état et du format."""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QRadioButton, QButtonGroup, QGroupBox
        from PyQt5.QtCore import Qt

        # Créer un dialogue moderne
        dialog = QDialog(self)
        dialog.setWindowTitle("Exporter les données")
        dialog.setMinimumWidth(500)

        if THEME_AVAILABLE:
            ThemeCls = get_current_theme()
            dialog.setStyleSheet(f"""
                QDialog {{
                    background: {ThemeCls.BG};
                    color: {ThemeCls.TXT};
                }}
                QLabel {{
                    color: {ThemeCls.TXT};
                }}
                QGroupBox {{
                    border: 2px solid {ThemeCls.BDR};
                    border-radius: 8px;
                    margin-top: 12px;
                    padding: 16px;
                    font-weight: bold;
                    background: {ThemeCls.BG_CARD};
                }}
                QGroupBox::title {{
                    subcontrol-origin: margin;
                    subcontrol-position: top left;
                    padding: 0 8px;
                    color: {ThemeCls.PRI};
                }}
                QRadioButton {{
                    color: {ThemeCls.TXT};
                    padding: 8px;
                    spacing: 8px;
                }}
                QRadioButton::indicator {{
                    width: 18px;
                    height: 18px;
                }}
            """)

        layout = QVBoxLayout(dialog)
        layout.setSpacing(20)
        layout.setContentsMargins(24, 24, 24, 24)

        # En-tête avec icône
        header = QLabel("Exporter la grille de polyvalence")
        header.setStyleSheet("font-size: 18px; font-weight: bold; padding-bottom: 8px;")
        layout.addWidget(header)

        desc = QLabel("Choisissez les options d'exportation :")
        desc.setStyleSheet("font-size: 13px; color: #6b7280; padding-bottom: 12px;")
        layout.addWidget(desc)

        # Groupe 1 : Choix de l'état
        state_group = QGroupBox("Données à exporter")
        state_layout = QVBoxLayout()
        state_layout.setSpacing(12)

        state_btn_group = QButtonGroup(dialog)

        current_radio = QRadioButton("État actuel (avec filtres et modifications)")
        current_radio.setChecked(True)
        current_desc = QLabel("   → Exporte uniquement les lignes et colonnes visibles")
        current_desc.setStyleSheet("font-size: 11px; color: #6b7280; font-weight: normal; padding-left: 30px;")

        full_radio = QRadioButton("Grille complète (toutes les données)")
        full_desc = QLabel("   → Exporte l'intégralité de la grille sans filtres")
        full_desc.setStyleSheet("font-size: 11px; color: #6b7280; font-weight: normal; padding-left: 30px;")

        state_btn_group.addButton(current_radio, 0)
        state_btn_group.addButton(full_radio, 1)

        state_layout.addWidget(current_radio)
        state_layout.addWidget(current_desc)
        state_layout.addSpacing(8)
        state_layout.addWidget(full_radio)
        state_layout.addWidget(full_desc)
        state_group.setLayout(state_layout)
        layout.addWidget(state_group)

        # Groupe 2 : Choix du format
        format_group = QGroupBox("Format d'exportation")
        format_layout = QVBoxLayout()
        format_layout.setSpacing(12)

        format_btn_group = QButtonGroup(dialog)

        excel_radio = QRadioButton("Excel (.xlsx)")
        excel_radio.setChecked(True)
        excel_desc = QLabel("   → Tableau modifiable avec légende des niveaux")
        excel_desc.setStyleSheet("font-size: 11px; color: #6b7280; font-weight: normal; padding-left: 30px;")

        pdf_radio = QRadioButton("PDF (.pdf)")
        pdf_desc = QLabel("   → Document imprimable avec résumé par opérateur")
        pdf_desc.setStyleSheet("font-size: 11px; color: #6b7280; font-weight: normal; padding-left: 30px;")

        format_btn_group.addButton(excel_radio, 0)
        format_btn_group.addButton(pdf_radio, 1)

        format_layout.addWidget(excel_radio)
        format_layout.addWidget(excel_desc)
        format_layout.addSpacing(8)
        format_layout.addWidget(pdf_radio)
        format_layout.addWidget(pdf_desc)
        format_group.setLayout(format_layout)
        layout.addWidget(format_group)

        # Boutons d'action
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(12)
        buttons_layout.addStretch()

        if THEME_AVAILABLE:
            cancel_btn = EmacButton("Annuler", variant='ghost')
            export_btn = EmacButton("Exporter", variant='primary')
        else:
            cancel_btn = QPushButton("Annuler")
            export_btn = QPushButton("Exporter")
            export_btn.setStyleSheet("background-color: #2563eb; color: white; padding: 10px 24px; font-weight: bold; border-radius: 6px;")
            cancel_btn.setStyleSheet("padding: 10px 24px; border-radius: 6px;")

        cancel_btn.clicked.connect(dialog.reject)
        export_btn.clicked.connect(dialog.accept)

        buttons_layout.addWidget(cancel_btn)
        buttons_layout.addWidget(export_btn)
        layout.addLayout(buttons_layout)

        # Afficher le dialogue
        if dialog.exec_() != QDialog.Accepted:
            return

        # Récupérer les choix
        export_current_state = (state_btn_group.checkedId() == 0)
        format_choice = "Excel" if format_btn_group.checkedId() == 0 else "PDF"

        # Demander le nom du fichier
        file_extension = {"Excel": "xlsx", "PDF": "pdf"}[format_choice]
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le fichier", "", f"{format_choice} Files (*.{file_extension})"
        )
        if not file_name:
            return

        # Effectuer l'export
        if format_choice == "Excel":
            self.export_to_excel(file_name, export_current_state)
        elif format_choice == "PDF":
            self.export_to_pdf(file_name, export_current_state)

    def export_to_excel(self, file_name, export_current_state):
        """Exporte les données du tableau en fichier Excel + ajoute la légende 'Niveaux' pour impression."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.utils.dataframe import dataframe_to_rows

            # Récupère l'état visible (ou complet) de la grille -> DataFrame
            data = []
            headers = []
            for col in range(self.main_table.columnCount()):
                if export_current_state and self.main_table.isColumnHidden(col):
                    continue
                headers.append(self.main_table.horizontalHeaderItem(col).text())

            for row in range(self.main_table.rowCount()):
                if export_current_state and self.main_table.isRowHidden(row):
                    continue
                row_data = [self.main_table.verticalHeaderItem(row).text()]
                for col in range(self.main_table.columnCount()):
                    if export_current_state and self.main_table.isColumnHidden(col):
                        continue
                    item = self.main_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            if not data:
                QMessageBox.warning(self, "Exportation annulée", "Aucune donnée visible à exporter.")
                return

            # ====== Création du classeur ======
            wb = Workbook()
            ws = wb.active
            ws.title = "Grille Polyvalence"

            # Écrit l’en-tête
            ws.cell(row=1, column=1, value="Nom").font = Font(bold=True)
            for c, h in enumerate(headers, start=2):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True)

            # Écrit les données
            for r, row_vals in enumerate(data, start=2):
                for c, val in enumerate(row_vals, start=1):
                    ws.cell(row=r, column=c, value=val)

            # Mise en forme tableau
            thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))
            for r in range(1, len(data) + 1 + 1):
                for c in range(1, len(headers) + 1 + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin

            # Largeurs de colonnes (Nom plus large)
            ws.column_dimensions[get_column_letter(1)].width = 28
            for c in range(2, len(headers) + 2):
                ws.column_dimensions[get_column_letter(c)].width = 6

            # ====== Légende "Niveaux" sous le tableau ======
            start_legend_row = len(data) + 3  # 1 ligne vide + 1 marge
            bullet = "\u2022"  # •

            # Titre
            ws.cell(row=start_legend_row, column=1, value="Niveaux :").font = Font(bold=True)
            ws.merge_cells(start_row=start_legend_row, start_column=1,
                           end_row=start_legend_row, end_column=len(headers) + 1)

            # Lignes de légende
            legend_lines = [
                f"{bullet} Niveau 1 : Opérateur nouveau à encadrer par titulaire N3/4 ou absent du poste depuis plus de 12 mois. (< 80%)",
                f"{bullet} Niveau 2 : Opérateur formé et apte à conduire le poste seul. Certaines notions sont en cours d'acquisition. (> 80%)",
                f"{bullet} Niveau 3 : Opérateur titulaire, formé, apte à conduire le poste et apte à former. (> 90%)",
                f"{bullet} Niveau 4 : N3 + Leader ou Polyvalent (maîtrise 3 postes d'une ligne : mél. interne, cylindres, conditionnement) (> 90%)",
            ]

            for i, line in enumerate(legend_lines, start=1):
                r = start_legend_row + i
                ws.cell(row=r, column=1, value=line)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers) + 1)
                c = ws.cell(row=r, column=1)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Ajuste automatiquement la hauteur des lignes de légende (un peu plus hautes pour l’impression)
            for r in range(start_legend_row, start_legend_row + len(legend_lines) + 1):
                ws.row_dimensions[r].height = 18

            # Zone d’impression incluant la légende
            last_row = start_legend_row + len(legend_lines)
            last_col_letter = get_column_letter(len(headers) + 1)
            ws.print_area = f"A1:{last_col_letter}{last_row}"

            # Orientation paysage conseillée pour impression (facultatif)
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

            # Sauvegarde
            wb.save(file_name)
            QMessageBox.information(self, "Exportation réussie", f"Les données ont été exportées dans {file_name}")

        except Exception as e:
            logger.exception(f"Erreur exportation: {e}")
            if show_error_message:
                show_error_message(self, "Erreur", "Erreur lors de l'exportation", e)
            else:
                QMessageBox.critical(self, "Erreur", "Erreur lors de l'exportation. Contactez l'administrateur.")


    def export_to_pdf(self, file_name, export_current_state):
        """Export PDF ultra-compact sur une seule page avec résumé par opérateur"""
        try:
            from datetime import datetime
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import Flowable

            # Récupération des données visibles
            headers = []
            for c in range(self.main_table.columnCount()):
                if export_current_state and self.main_table.isColumnHidden(c):
                    continue
                hi = self.main_table.horizontalHeaderItem(c)
                headers.append(hi.text() if hi else "")

            data_rows = []
            row_headers = []
            n_ops = len(self.operateurs)

            for r in range(self.main_table.rowCount()):
                if export_current_state and self.main_table.isRowHidden(r):
                    continue
                vh = self.main_table.verticalHeaderItem(r)
                row_name = vh.text() if vh else ""
                row_headers.append(row_name)

                row = []
                for c in range(self.main_table.columnCount()):
                    if export_current_state and self.main_table.isColumnHidden(c):
                        continue
                    it = self.main_table.item(r, c)
                    row.append("" if it is None else it.text())
                data_rows.append(row)

            if not data_rows:
                QMessageBox.warning(self, "Exportation annulée", "Aucune donnée à exporter.")
                return

            synthesis_start = len(data_rows) - 7
            operator_rows = data_rows[:synthesis_start]
            operator_headers = row_headers[:synthesis_start]
            synthesis_rows = data_rows[synthesis_start:]
            synthesis_headers = row_headers[synthesis_start:]

            # Configuration page A4 portrait - marges ultra-minimales
            page = A4
            LM, RM, TM, BM = 3*mm, 3*mm, 8*mm, 3*mm  # Marges minimales (TM réduit)
            usable_w = page[0] - LM - RM
            usable_h = page[1] - TM - BM

            # Tailles de police augmentées avec l'espace maximal
            title_size = 11
            th_size = 6.5
            td_size = 6

            # Calcul largeurs colonnes avec espace maximal
            n_poste_cols = len(headers)
            first_col_w = 30*mm  # Colonne noms élargie
            summary_col_w = 12*mm  # Colonne résumé élargie
            remaining_w = usable_w - first_col_w - summary_col_w
            poste_col_w = max(4.2*mm, remaining_w / n_poste_cols)  # Colonnes postes élargies

            col_widths = [first_col_w] + [poste_col_w] * n_poste_cols + [summary_col_w]

            # Texte vertical pour les en-têtes de postes
            class RotatedText(Flowable):
                def __init__(self, text, fontSize=5):
                    Flowable.__init__(self)
                    self.text = text
                    self.fontSize = fontSize
                    self.width = fontSize * 1.2
                    self.height = min(22, len(text) * fontSize * 0.6)

                def draw(self):
                    canvas = self.canv
                    canvas.saveState()
                    canvas.translate(self.width / 2, 0)
                    canvas.rotate(90)
                    canvas.setFont("Helvetica-Bold", self.fontSize)
                    canvas.drawString(1, -self.fontSize/3, self.text)
                    canvas.restoreState()

            # Construction du tableau principal
            styles = getSampleStyleSheet()
            td_c_style = ParagraphStyle("TDc", fontName="Helvetica", fontSize=td_size, leading=td_size+1, alignment=1)
            td_l_style = ParagraphStyle("TDl", fontName="Helvetica", fontSize=td_size, leading=td_size+1, alignment=0)
            td_b_style = ParagraphStyle("TDb", fontName="Helvetica-Bold", fontSize=td_size, leading=td_size+1, alignment=1)

            # En-tête avec résumé
            header_row = [""] + [RotatedText(h, th_size) for h in headers] + [RotatedText("Résumé", th_size)]

            data = [header_row]

            # Lignes opérateurs avec calcul du résumé
            for i, row in enumerate(operator_rows):
                # Calcul résumé : compter les niveaux
                levels = [val.strip() for val in row if val.strip().isdigit()]
                n1 = levels.count('1')
                n2 = levels.count('2')
                n3 = levels.count('3')
                n4 = levels.count('4')
                total = len(levels)

                # Format détaillé : afficher tous les niveaux présents
                summary_parts = []
                if n1 > 0:
                    summary_parts.append(f"{n1}xN1")
                if n2 > 0:
                    summary_parts.append(f"{n2}xN2")
                if n3 > 0:
                    summary_parts.append(f"{n3}xN3")
                if n4 > 0:
                    summary_parts.append(f"{n4}xN4")

                # Afficher tous les niveaux ou rien si aucun poste
                summary = "<br/>".join(summary_parts) if summary_parts else ""

                data.append(
                    [Paragraph(operator_headers[i], td_l_style)] +
                    [Paragraph(str(v) if v else "", td_c_style) for v in row] +
                    [Paragraph(f"<b>{summary}</b>", td_b_style)]
                )

            # Lignes de synthèse
            for i, row in enumerate(synthesis_rows):
                data.append(
                    [Paragraph(f"<b>{synthesis_headers[i]}</b>", td_l_style)] +
                    [Paragraph(str(v) if v else "", td_c_style) for v in row] +
                    [Paragraph("", td_c_style)]  # Pas de résumé pour les lignes de synthèse
                )

            table = Table(data, colWidths=col_widths)

            n_op_rows = len(operator_rows) + 1

            table_style = [
                ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#BFBFBF")),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (1, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("BACKGROUND", (0, n_op_rows), (-1, -1), colors.HexColor("#E0E0E0")),
                ("BACKGROUND", (-1, 1), (-1, n_op_rows-1), colors.HexColor("#FFF9C4")),  # Colonne résumé en jaune
                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                ("RIGHTPADDING", (0, 0), (-1, -1), 1),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]

            table.setStyle(TableStyle(table_style))

            # Légende compacte en bas
            legend_style = ParagraphStyle("Leg", fontName="Helvetica", fontSize=6, leading=7, alignment=0)
            legend_text = (
                "<b>Niveaux :</b> "
                "<b>N1</b>: &lt;80% (nouveau/absent 12m) | "
                "<b>N2</b>: ≥80% (formé, autonome) | "
                "<b>N3</b>: &gt;90% (formateur) | "
                "<b>N4</b>: &gt;90% (leader/polyvalent) | "
                "<b>Résumé :</b> affiche le nombre de postes N3 et N4 (ex: 5xN3 = 5 postes niveau 3)"
            )
            legend = Paragraph(legend_text, legend_style)

            # En-tête page
            def on_page(canvas, doc):
                canvas.saveState()
                canvas.setFont("Helvetica-Bold", title_size)
                canvas.drawCentredString(page[0]/2, page[1] - TM + 3*mm,
                                        f"Grille de Polyvalence au {datetime.now().strftime('%d/%m/%Y')}")
                canvas.setFont("Helvetica", 6)
                canvas.drawString(LM, TM - 3*mm, "LQ 07 02 02 rév.1")
                canvas.restoreState()

            doc = SimpleDocTemplate(
                file_name, pagesize=page,
                leftMargin=LM, rightMargin=RM, topMargin=TM + 5*mm, bottomMargin=BM + 3*mm,
                title="Grille de Polyvalence"
            )

            story = [
                table,
                Spacer(1, 2*mm),
                legend
            ]

            doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
            QMessageBox.information(self, "Exportation réussie", f"PDF généré : {file_name}")

        except Exception as e:
            logger.exception(f"Erreur export PDF: {e}")
            if show_error_message:
                show_error_message(self, "Erreur", "Erreur lors de l'export PDF", e)
            else:
                QMessageBox.critical(self, "Erreur", "Erreur lors de l'export PDF. Contactez l'administrateur.")



if __name__ == "__main__":
    app = QApplication(sys.argv)
    dialog = GrillesDialog()
    dialog.show()
    sys.exit(app.exec_())
