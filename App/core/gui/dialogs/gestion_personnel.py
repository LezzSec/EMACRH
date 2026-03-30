# gestion_personnel.py вҖ“ Gestion complГЁte du personnel (actifs et inactifs)

from core.utils.logging_config import get_logger
logger = get_logger(__name__)

from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QLineEdit, QGroupBox,
    QMessageBox, QAbstractItemView, QWidget, QTabWidget, QCheckBox,
    QRadioButton, QDateEdit, QScrollArea, QFrame, QGridLayout, QInputDialog
)
from PyQt5.QtCore import Qt, QDate, QTimer, pyqtSignal
from PyQt5.QtGui import QFont, QColor

from core.repositories.personnel_repo import PersonnelRepository
from core.repositories.polyvalence_repo import PolyvalenceRepository
from core.services.contrat_service_crud import ContratServiceCRUD
from core.services.formation_service_crud import FormationServiceCRUD
from core.services import medical_service
from core.gui.workers.db_worker import DbWorker, DbThreadPool
from core.services.optimized_db_logger import log_hist
from core.gui.dialogs.historique_personnel import HistoriquePersonnelTab
from core.gui.components.emac_ui_kit import add_custom_title_bar, show_error_message
from core.services.auth_service import get_current_user
from core.services.permission_manager import require, can
from core.services.personnel_service import PersonnelService

import datetime as dt
from core.utils.date_format import format_date


class DetailOperateurDialog(QDialog):
    """FenГӘtre modale affichant tous les dГ©tails d'un opГ©rateur."""
    
    operateur_status_changed = pyqtSignal(int)  # Signal Г©mis si changement de statut
    
    def __init__(self, operateur_id, nom, prenom, statut, parent=None, is_production=True):
        super().__init__(parent)
        self.operateur_id = operateur_id
        self.operateur_nom = nom
        self.operateur_prenom = prenom
        self.is_production = is_production
        self.current_statut = statut.upper()
        self.date_entree = self._load_date_entree()  # Charger la date d'entrГ©e
        self.setWindowTitle(f"DГ©tails - {nom} {prenom}")
        self.setGeometry(200, 150, 900, 600)

        # CrГ©er le layout principal avec marges nulles pour la barre de titre
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Ajouter la barre de titre personnalisГ©e
        title_bar = add_custom_title_bar(self, f"DГ©tails - {nom} {prenom}")
        main_layout.addWidget(title_bar)

        # CrГ©er le widget de contenu avec marges normales
        content_widget = QWidget()
        layout = QVBoxLayout(content_widget)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        header_content = QVBoxLayout()
        header = QLabel(f"{nom} {prenom}")
        header.setFont(QFont("Segoe UI", 18, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        header_content.addWidget(header)

        status_label = QLabel(f"Statut : {self.current_statut}")
        if self.current_statut == "ACTIF":
            status_label.setStyleSheet("color: #10b981; font-weight: bold; font-size: 14px;")
        else:
            status_label.setStyleSheet("color: #dc2626; font-weight: bold; font-size: 14px;")
        status_label.setAlignment(Qt.AlignCenter)
        header_content.addWidget(status_label)

        layout.addLayout(header_content)
        
        tabs = QTabWidget()
        
        poly_tab = QWidget()
        poly_layout = QVBoxLayout(poly_tab)
        
        poly_label = QLabel("Polyvalences et CompГ©tences")
        poly_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        poly_layout.addWidget(poly_label)
        
        self.poly_table = QTableWidget()
        self.poly_table.setColumnCount(6)
        self.poly_table.setHorizontalHeaderLabels([
            "Poste", "Niveau", "Date Гүvaluation", "Prochaine Гүval.", "AnciennetГ©", "Statut"
        ])
        self.poly_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.poly_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        poly_layout.addWidget(self.poly_table)
        
        stats_box = QHBoxLayout()
        self.stat_n1 = self._create_mini_stat("N1", "0", "#ef4444")
        self.stat_n2 = self._create_mini_stat("N2", "0", "#f59e0b")
        self.stat_n3 = self._create_mini_stat("N3", "0", "#10b981")
        self.stat_n4 = self._create_mini_stat("N4", "0", "#3b82f6")
        self.stat_total = self._create_mini_stat("Total", "0", "#6b7280")
        
        stats_box.addWidget(self.stat_n1)
        stats_box.addWidget(self.stat_n2)
        stats_box.addWidget(self.stat_n3)
        stats_box.addWidget(self.stat_n4)
        stats_box.addWidget(self.stat_total)
        poly_layout.addLayout(stats_box)
        
        self._poly_tab_ref = poly_tab  # EmpГӘche le GC de dГ©truire poly_table quand is_production=False
        if self.is_production:
            tabs.addTab(poly_tab, "Polyvalences")

        # Onglet Infos ComplГ©mentaires (disposition en cartes)
        infos_tab = QWidget()
        infos_layout = QVBoxLayout(infos_tab)
        infos_layout.setSpacing(0)
        infos_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("""
            QScrollArea {
                background-color: #f8fafc;
                border: none;
            }
            QScrollBar:vertical {
                background: #f1f5f9;
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #cbd5e1;
                border-radius: 4px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background: #94a3b8;
            }
        """)

        self.infos_container = QWidget()
        self.infos_container.setStyleSheet("background-color: #f8fafc;")
        self.infos_cards_layout = QVBoxLayout(self.infos_container)
        self.infos_cards_layout.setSpacing(16)
        self.infos_cards_layout.setContentsMargins(20, 20, 20, 20)
        self.infos_cards_layout.setAlignment(Qt.AlignTop)

        scroll.setWidget(self.infos_container)
        infos_layout.addWidget(scroll)

        tabs.addTab(infos_tab, "Infos ComplГ©mentaires")

        # Onglet Historique polyvalences (production uniquement)
        if self.is_production:
            self.history_tab = HistoriquePersonnelTab(
                operateur_id=self.operateur_id,
                operateur_nom=nom,
                operateur_prenom=prenom,
                parent=self
            )
            tabs.addTab(self.history_tab, "Historique Polyvalences")

        self.tabs_widget = tabs
        layout.addWidget(tabs)
        
        actions = QHBoxLayout()

        self.toggle_status_btn = QPushButton()
        self.update_status_button()
        self.toggle_status_btn.clicked.connect(self.toggle_operateur_status)
        actions.addWidget(self.toggle_status_btn)

        if can("rh.personnel.delete"):
            self.delete_btn = QPushButton("Supprimer definitivement")
            self.delete_btn.setStyleSheet("""
                QPushButton {
                    background: #7f1d1d;
                    color: white;
                    font-weight: bold;
                    padding: 10px 20px;
                    border-radius: 8px;
                    border: 2px solid #991b1b;
                }
                QPushButton:hover { background: #991b1b; }
                QPushButton:pressed { background: #450a0a; }
            """)
            self.delete_btn.clicked.connect(self.delete_operateur)
            actions.addWidget(self.delete_btn)

        actions.addStretch()

        self.export_btn = QPushButton("Exporter le profil")
        self.export_btn.clicked.connect(self.export_profile)
        actions.addWidget(self.export_btn)

        self.close_btn = QPushButton("Fermer")
        self.close_btn.clicked.connect(self.close)
        actions.addWidget(self.close_btn)

        layout.addLayout(actions)

        # Ajouter le widget de contenu au layout principal
        main_layout.addWidget(content_widget)

        self.load_data()
    
    def update_status_button(self):
        """Met Г  jour le texte et le style du bouton de changement de statut."""
        if self.current_statut == "ACTIF":
            self.toggle_status_btn.setText("DГ©sactiver l'opГ©rateur")
            self.toggle_status_btn.setStyleSheet("""
                QPushButton {
                    background: #dc2626;
                    color: white;
                    font-weight: bold;
                    padding: 10px 20px;
                    border-radius: 8px;
                }
                QPushButton:hover {
                    background: #b91c1c;
                }
            """)
        else:
            self.toggle_status_btn.setText("RГ©activer l'opГ©rateur")
            self.toggle_status_btn.setStyleSheet("""
                QPushButton {
                    background: #10b981;
                    color: white;
                    font-weight: bold;
                    padding: 10px 20px;
                    border-radius: 8px;
                }
                QPushButton:hover {
                    background: #059669;
                }
            """)
    
    def _create_mini_stat(self, label, value, color):
        """CrГ©e une mini-carte de statistique."""
        widget = QWidget()
        widget.setStyleSheet(f"""
            QWidget {{
                background: {color};
                border-radius: 6px;
                padding: 8px;
            }}
        """)
        
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(2)
        
        val_label = QLabel(value)
        val_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        val_label.setStyleSheet("color: white;")
        val_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(val_label)
        
        text_label = QLabel(label)
        text_label.setStyleSheet("color: white; font-size: 10px;")
        text_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(text_label)
        
        widget.value_label = val_label
        return widget
    
    def load_data(self):
        """Charge toutes les donnГ©es de l'opГ©rateur."""
        if self.is_production:
            self.load_polyvalences()
        self.load_additional_infos()
        # L'historique polyvalences se charge automatiquement via HistoriquePersonnelTab

    
    
    def load_additional_infos(self):
        """Charge les infos complГ©mentaires avec une disposition en cartes modernes."""
        try:
            # Nettoyer les cartes existantes
            self._clear_infos_cards()

            # Stocker les donnГ©es pour l'export
            self._infos_data = []

            _pi = PersonnelRepository.get_personnel_infos(self.operateur_id)
            row_data = [_pi] if _pi else []

            personal_items = []
            if row_data:
                data = row_data[0]
                if data.get('date_entree'):
                    date_entree_str = data['date_entree'].strftime("%d/%m/%Y") if isinstance(data['date_entree'], dt.date) else str(data['date_entree'])
                    personal_items.append(("Date d'entrГ©e", date_entree_str))
                for key, val in data.items():
                    if key in ("personnel_id", "date_entree"):
                        continue
                    label = self._format_column_name(key)
                    value = (
                        val.strftime("%d/%m/%Y") if isinstance(val, dt.date)
                        else (str(val) if val not in (None, "",) else None)
                    )
                    if value is not None:
                        personal_items.append((label, value))

            if not personal_items:
                personal_items.append(("Information", "Aucune information complГ©mentaire enregistrГ©e"))

            self._add_info_card(
                "Informations Personnelles",
                personal_items,
                icon_color="#3b82f6",
                icon="рҹ‘Ө"
            )

            contr = ContratServiceCRUD.get_by_operateur(self.operateur_id, actif_only=True)[:1]

            contract_items = []
            if contr:
                c = contr[0]
                if c.get("type_contrat"):
                    contract_items.append(("Type", c.get("type_contrat")))
                if c.get("etp") is not None:
                    contract_items.append(("ETP", str(c.get("etp"))))
                if c.get("categorie"):
                    contract_items.append(("CatГ©gorie", c.get("categorie")))
                d1, d2 = c.get("date_debut"), c.get("date_fin")
                if d1:
                    deb = d1.strftime("%d/%m/%Y") if isinstance(d1, dt.date) else str(d1)
                    contract_items.append(("DГ©but", deb))
                if d2:
                    fin = d2.strftime("%d/%m/%Y") if isinstance(d2, dt.date) else str(d2)
                    contract_items.append(("Fin", fin))
            else:
                contract_items.append(("Statut", "Aucun contrat actif"))

            self._add_info_card(
                "Contrat Actuel",
                contract_items,
                icon_color="#f59e0b",
                icon="рҹ“Ӣ",
                on_click=self._open_contract_management if contr else None
            )

            formations = FormationServiceCRUD.get_by_operateur(self.operateur_id)[:5]

            formation_items = []
            if formations:
                for f in formations:
                    d1 = self._format_date(f.get("date_debut"))
                    d2 = self._format_date(f.get("date_fin"))
                    cert = " вң“" if f.get("certificat_obtenu") else ""
                    intitule = f.get("intitule", "(formation)")
                    formation_items.append((intitule, f"{d1} вҶ’ {d2}{cert}"))
            else:
                formation_items.append(("Aucune formation", "Aucune formation renseignГ©e"))

            self._add_info_card(
                "Formations",
                formation_items,
                icon_color="#10b981",
                icon="рҹҺ“",
                on_click=self._open_formation_rh if formations else None
            )

            validites = medical_service.get_validites_operateur(self.operateur_id)

            validite_items = []
            if validites:
                for v in validites:
                    d1 = self._format_date(v.get("date_debut"))
                    d2 = self._format_date(v.get("date_fin")) if v.get("date_fin") else "вҖ”"
                    tc = f" ({v['taux_incapacite']}%)" if v.get("taux_incapacite") is not None else ""
                    type_val = v.get("type_validite", "(type)")
                    validite_items.append((type_val, f"{d1} вҶ’ {d2}{tc}"))
            else:
                validite_items.append(("Aucune validitГ©", "Aucune validitГ© enregistrГ©e"))

            self._add_info_card(
                "ValiditГ©s",
                validite_items,
                icon_color="#8b5cf6",
                icon="вң…"
            )

            # Ajouter un stretch Г  la fin pour pousser les cartes vers le haut
            self.infos_cards_layout.addStretch()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur de chargement",
                f"Impossible de charger les infos dГ©taillГ©es :\n{e}"
            )

    def _clear_infos_cards(self):
        """Supprime toutes les cartes existantes du layout."""
        while self.infos_cards_layout.count():
            item = self.infos_cards_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

    def _add_info_card(self, title: str, items: list, icon_color: str = "#3b82f6", icon: str = "рҹ“„", on_click=None):
        """CrГ©e et ajoute une carte d'information au layout.

        Args:
            on_click: Fonction callback appelГ©e lors du clic sur la carte (optionnel)
        """
        # Stocker les donnГ©es pour l'export
        if hasattr(self, '_infos_data'):
            self._infos_data.append((title, items))

        card = QFrame()

        # Style diffГ©rent si cliquable
        if on_click:
            card.setCursor(Qt.PointingHandCursor)
            card.setStyleSheet(f"""
                QFrame {{
                    background-color: white;
                    border-radius: 12px;
                    border: 1px solid #e2e8f0;
                }}
                QFrame:hover {{
                    border: 2px solid {icon_color};
                    background-color: #fafafa;
                }}
            """)
            card.mousePressEvent = lambda event: on_click()
        else:
            card.setStyleSheet(f"""
                QFrame {{
                    background-color: white;
                    border-radius: 12px;
                    border: 1px solid #e2e8f0;
                }}
                QFrame:hover {{
                    border: 1px solid {icon_color};
                }}
            """)

        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(20, 16, 20, 16)
        card_layout.setSpacing(12)

        # En-tГӘte de la carte avec icГҙne
        header = QHBoxLayout()
        header.setSpacing(10)

        # IcГҙne dans un cercle colorГ©
        icon_label = QLabel(icon)
        icon_label.setFixedSize(36, 36)
        icon_label.setAlignment(Qt.AlignCenter)
        icon_label.setStyleSheet(f"""
            background-color: {icon_color}20;
            border-radius: 18px;
            font-size: 16px;
        """)
        header.addWidget(icon_label)

        # Titre
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        title_label.setStyleSheet(f"color: #1e293b; background: transparent;")
        header.addWidget(title_label)
        header.addStretch()

        card_layout.addLayout(header)

        # SГ©parateur
        separator = QFrame()
        separator.setFixedHeight(1)
        separator.setStyleSheet("background-color: #e2e8f0;")
        card_layout.addWidget(separator)

        # Contenu : grille de clГ©-valeur
        content_layout = QGridLayout()
        content_layout.setSpacing(8)
        content_layout.setColumnStretch(1, 1)

        for row, (label, value) in enumerate(items):
            # Label
            lbl = QLabel(label)
            lbl.setFont(QFont("Segoe UI", 10))
            lbl.setStyleSheet("color: #64748b; background: transparent;")
            content_layout.addWidget(lbl, row, 0, Qt.AlignTop)

            # Valeur
            val = QLabel(str(value) if value else "вҖ”")
            val.setFont(QFont("Segoe UI", 10, QFont.DemiBold))
            val.setStyleSheet("color: #1e293b; background: transparent;")
            val.setWordWrap(True)
            content_layout.addWidget(val, row, 1, Qt.AlignTop)

        card_layout.addLayout(content_layout)

        # Ajouter un indicateur "Voir dГ©tails" si cliquable
        if on_click:
            link_label = QLabel("Voir dГ©tails вҶ’")
            link_label.setFont(QFont("Segoe UI", 9))
            link_label.setStyleSheet(f"color: {icon_color}; background: transparent;")
            link_label.setAlignment(Qt.AlignRight)
            card_layout.addWidget(link_label)

        self.infos_cards_layout.addWidget(card)

    def _open_contract_management(self):
        """Ouvre la Gestion RH directement sur l'onglet Contrat pour cet opГ©rateur."""
        from core.gui.dialogs.gestion_rh import GestionRHDialog
        from core.services.rh_service import DomaineRH
        dialog = GestionRHDialog(self)
        dialog._selectionner_operateur_par_id(self.operateur_id)
        dialog._on_domaine_change(DomaineRH.CONTRAT.value)
        dialog.exec_()
        # Recharger les infos aprГЁs fermeture
        self.load_additional_infos()

    def _open_formation_rh(self):
        """Ouvre la Gestion RH directement sur l'onglet Formation pour cet opГ©rateur."""
        from core.gui.dialogs.gestion_rh import GestionRHDialog
        from core.services.rh_service import DomaineRH
        dialog = GestionRHDialog(self)
        dialog._selectionner_operateur_par_id(self.operateur_id)
        dialog._on_domaine_change(DomaineRH.FORMATION.value)
        dialog.exec_()
        # Recharger les infos aprГЁs fermeture
        self.load_additional_infos()

    def _format_column_name(self, col_name: str) -> str:
        """
        Formate un nom de colonne SQL en texte lisible.
        Exemple: 'date_naissance' вҶ’ 'Date Naissance'
        """
        if not col_name:
            return ""

        formatted = col_name.replace("_", " ")
        formatted = " ".join(word.capitalize() for word in formatted.split())

        replacements = {
            "Cp": "CP",
            "Rtt": "RTT",
            "Id": "ID",
            "Nir": "NIR",
            "Etp": "ETP",
            "Operateur": "OpГ©rateur",
        }
        for old, new in replacements.items():
            formatted = formatted.replace(old, new)

        return formatted


    
    def load_polyvalences(self):
        """Charge les polyvalences."""
        try:
            rows = PolyvalenceRepository.get_by_operateur_dict(self.operateur_id)
            
            self.poly_table.setRowCount(0)
            
            niveaux = {1: 0, 2: 0, 3: 0, 4: 0}
            
            for r in rows:
                row = self.poly_table.rowCount()
                self.poly_table.insertRow(row)
                
                poste = r.get("poste_code", "")
                niveau = r.get("niveau")
                date_eval = r.get("date_evaluation")
                date_next = r.get("prochaine_evaluation")
                
                # Poste
                self.poly_table.setItem(row, 0, QTableWidgetItem(str(poste)))
                
                # Niveau
                niveau_item = QTableWidgetItem(str(niveau) if niveau else "N/A")
                if niveau:
                    niveaux[int(niveau)] += 1
                    if niveau == 1:
                        niveau_item.setBackground(QColor("#fef2f2"))
                        niveau_item.setForeground(QColor("#dc2626"))
                    elif niveau == 2:
                        niveau_item.setBackground(QColor("#fffbeb"))
                        niveau_item.setForeground(QColor("#d97706"))
                    elif niveau == 3:
                        niveau_item.setBackground(QColor("#f0fdf4"))
                        niveau_item.setForeground(QColor("#059669"))
                    elif niveau == 4:
                        niveau_item.setBackground(QColor("#eff6ff"))
                        niveau_item.setForeground(QColor("#2563eb"))
                niveau_item.setTextAlignment(Qt.AlignCenter)
                self.poly_table.setItem(row, 1, niveau_item)
                
                # Dates
                self.poly_table.setItem(row, 2, QTableWidgetItem(self._format_date(date_eval)))
                self.poly_table.setItem(row, 3, QTableWidgetItem(self._format_date(date_next)))
                
                # AnciennetГ© (basГ©e sur la date d'entrГ©e)
                anciennete = self._calculate_anciennete()
                self.poly_table.setItem(row, 4, QTableWidgetItem(anciennete))
                
                # Statut (ГҖ jour / En retard)
                statut_item = QTableWidgetItem()
                if date_next:
                    today = dt.date.today()
                    if isinstance(date_next, str):
                        next_date = dt.datetime.strptime(date_next, "%Y-%m-%d").date()
                    elif hasattr(date_next, "date"):
                        next_date = date_next.date()
                    else:
                        next_date = date_next
                    
                    if next_date < today:
                        statut_item.setText("En retard")
                        statut_item.setForeground(QColor("#dc2626"))
                    else:
                        statut_item.setText("ГҖ jour")
                        statut_item.setForeground(QColor("#059669"))
                else:
                        statut_item.setText("ГҖ planifier")
                        statut_item.setForeground(QColor("#d97706"))  # orange
                
                statut_item.setTextAlignment(Qt.AlignCenter)
                self.poly_table.setItem(row, 5, statut_item)
            
            # Mise Г  jour des stats
            self.stat_n1.value_label.setText(str(niveaux[1]))
            self.stat_n2.value_label.setText(str(niveaux[2]))
            self.stat_n3.value_label.setText(str(niveaux[3]))
            self.stat_n4.value_label.setText(str(niveaux[4]))
            self.stat_total.value_label.setText(str(sum(niveaux.values())))
            
        except Exception as e:
            logger.exception(f"Erreur chargement polyvalences: {e}")
            show_error_message(self, "Erreur", "Impossible de charger les polyvalences", e)

    # MГ©thode load_summary() supprimГ©e - l'onglet RГ©sumГ© a Г©tГ© retirГ©
    
    # MГ©thode load_history() supprimГ©e - remplacГ©e par le widget HistoriquePersonnelTab

    def toggle_operateur_status(self):
        """Change le statut de l'opГ©rateur (ACTIF <-> INACTIF)."""
        try:
            require('rh.personnel.edit')
        except PermissionError:
            QMessageBox.warning(self, "AccГЁs refusГ©", "Vous n'avez pas les droits pour modifier le statut du personnel.")
            return

        new_statut = "INACTIF" if self.current_statut == "ACTIF" else "ACTIF"
        action = "dГ©sactiver" if new_statut == "INACTIF" else "rГ©activer"

        reply = QMessageBox.question(
            self, f"Confirmer {action}",
            f"ГҠtes-vous sГ»r de vouloir {action} cet opГ©rateur ?\n\n"
            f"Son statut passera Г  {new_statut}.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        try:
            # RГ©cupГ©rer nom/prГ©nom pour le log
            pers_info = PersonnelRepository.get_info_basique(self.operateur_id)

            # Mettre Г  jour le statut
            if new_statut == "INACTIF":
                PersonnelRepository.desactiver(self.operateur_id)
            else:
                PersonnelRepository.reactiver(self.operateur_id)

            # Logger le changement de statut
            if pers_info:
                import json
                log_hist(
                    action="UPDATE",
                    table_name="personnel",
                    record_id=self.operateur_id,
                    operateur_id=self.operateur_id,
                    description=json.dumps({
                        "operateur": f"{pers_info['prenom']} {pers_info['nom']}",
                        "old_statut": self.current_statut,
                        "new_statut": new_statut,
                        "type": "changement_statut"
                    }, ensure_ascii=False),
                    source="GUI/gestion_personnel"
                )

            self.current_statut = new_statut
            self.update_status_button()

            QMessageBox.information(
                self, "Statut modifiГ©",
                f"Le statut de l'opГ©rateur a Г©tГ© changГ© Г  {new_statut} avec succГЁs !"
            )

            self.operateur_status_changed.emit(self.operateur_id)

        except Exception as e:
            logger.exception(f"Erreur modification statut: {e}")
            show_error_message(self, "Erreur", "Impossible de modifier le statut", e)

    def _get_delete_counts(self, personnel_id: int) -> dict:
        """RГ©cupГЁre les comptages des donnГ©es liГ©es Г  un personnel."""
        from core.db.query_executor import QueryExecutor
        counts = {}
        queries = {
            "documents":    ("SELECT COUNT(*) FROM documents WHERE personnel_id = %s", personnel_id),
            "contrats":     ("SELECT COUNT(*) FROM contrat WHERE personnel_id = %s", personnel_id),
            "formations":   ("SELECT COUNT(*) FROM formation WHERE personnel_id = %s", personnel_id),
            "absences":     ("SELECT COUNT(*) FROM declaration WHERE personnel_id = %s", personnel_id),
            "polyvalences": ("SELECT COUNT(*) FROM polyvalence WHERE personnel_id = %s", personnel_id),
            "historique":   ("SELECT COUNT(*) FROM historique WHERE operateur_id = %s", personnel_id),
        }
        for key, (sql, param) in queries.items():
            try:
                counts[key] = QueryExecutor.fetch_scalar(sql, (param,), default=0) or 0
            except Exception:
                counts[key] = "?"
        return counts

    def delete_operateur(self):
        """Supprime dГ©finitivement l'opГ©rateur avec double confirmation."""
        try:
            require("rh.personnel.delete")
        except PermissionError:
            QMessageBox.warning(self, "AccГЁs refusГ©", "Vous n'avez pas les droits pour supprimer du personnel.")
            return

        nom_complet = f"{self.operateur_nom} {self.operateur_prenom}"

        counts = self._get_delete_counts(self.operateur_id)

        def _fmt(n):
            return f"<b>{n}</b>" if isinstance(n, int) and n > 0 else str(n)

        # --- 1re confirmation ---
        reply = QMessageBox.warning(
            self,
            "Suppression dГ©finitive",
            f"<b>Cette action est irrГ©versible.</b><br><br>"
            f"Supprimer <b>{nom_complet}</b> supprimera Г©galement :<br>"
            f"&nbsp;&nbsp;вҖў {_fmt(counts['documents'])} document(s) RH<br>"
            f"&nbsp;&nbsp;вҖў {_fmt(counts['contrats'])} contrat(s)<br>"
            f"&nbsp;&nbsp;вҖў {_fmt(counts['formations'])} formation(s)<br>"
            f"&nbsp;&nbsp;вҖў {_fmt(counts['absences'])} dГ©claration(s) d'absence<br>"
            f"&nbsp;&nbsp;вҖў {_fmt(counts['polyvalences'])} Г©valuation(s) / polyvalence(s)<br>"
            f"&nbsp;&nbsp;вҖў {_fmt(counts['historique'])} entrГ©e(s) d'historique<br><br>"
            f"Voulez-vous continuer ?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        # --- 2e confirmation : saisie du nom exact ---
        texte, ok = QInputDialog.getText(
            self,
            "Confirmation finale",
            f"Pour confirmer, saisissez exactement :\n{nom_complet}",
        )
        if not ok:
            return
        if texte.strip() != nom_complet:
            QMessageBox.warning(
                self, "AnnulГ©",
                "Le nom saisi ne correspond pas. Suppression annulГ©e."
            )
            return

        try:
            success, msg = PersonnelService.delete(record_id=self.operateur_id)
            if not success:
                raise RuntimeError(msg)

            QMessageBox.information(
                self, "SupprimГ©",
                f"{nom_complet} a Г©tГ© supprimГ© dГ©finitivement."
            )
            self.operateur_status_changed.emit(self.operateur_id)
            self.close()

        except Exception as e:
            logger.exception(f"Erreur suppression personnel {self.operateur_id}: {e}")
            show_error_message(self, "Erreur", "La suppression a Г©chouГ©", e)

    def export_profile(self):
        """Demande le format d'export puis lance PDF ou Excel."""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton

        dlg = QDialog(self)
        dlg.setWindowTitle("Exporter le profil")
        dlg.setMinimumWidth(360)

        v = QVBoxLayout(dlg)
        v.addWidget(QLabel("Choisir le format d'export :"))

        btns = QHBoxLayout()
        b_pdf = QPushButton("PDF")
        b_xlsx = QPushButton("Excel")
        b_cancel = QPushButton("Annuler")
        b_pdf.clicked.connect(lambda: (setattr(dlg, "_choice", "pdf"), dlg.accept()))
        b_xlsx.clicked.connect(lambda: (setattr(dlg, "_choice", "xlsx"), dlg.accept()))
        b_cancel.clicked.connect(dlg.reject)

        btns.addStretch(1)
        btns.addWidget(b_pdf)
        btns.addWidget(b_xlsx)
        btns.addWidget(b_cancel)
        v.addLayout(btns)

        if dlg.exec_() != QDialog.Accepted or getattr(dlg, "_choice", None) is None:
            return

        if dlg._choice == "pdf":
            self.export_profile_pdf()
        else:
            self.export_profile_excel()

    def export_profile_pdf(self):
        """
        Export PDF professionnel avec mise en page amГ©liorГ©e.
        Analyse et structure le contenu en sections claires.
        """
        try:
            from PyQt5.QtWidgets import QFileDialog, QMessageBox
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
            )
            from reportlab.lib.units import cm
            from reportlab.pdfgen.canvas import Canvas
            import datetime as _dt
            import re
    
            # ---------- Helpers ----------
            def _clean_text(text: str) -> str:
                """Nettoie le texte des caractГЁres de formatage."""
                if not text:
                    return ""
                # Supprimer les barres dГ©coratives
                text = re.sub(r'[в– в–Ҳв•”в•—в•ҡв•қв•җв”Җв”Ӯв”җв””в”ҳв”ңв”Өв”¬в”ҙв”ј]', '', text)
                # Supprimer les lignes vides multiples
                text = re.sub(r'\n{3,}', '\n\n', text)
                # Normaliser les espaces autour des :
                text = re.sub(r'\s*:\s*', ' : ', text)
                return text.strip()
    
            def _parse_section(text: str) -> dict:
                """Parse une section et extrait les donnГ©es structurГ©es."""
                lines = [l.strip() for l in text.split('\n') if l.strip()]
                sections = {}
                current_section = None
                current_items = []
                
                for line in lines:
                    # DГ©tecter les titres de section (tout en majuscules)
                    if line.isupper() and len(line) > 3:
                        if current_section:
                            sections[current_section] = current_items
                        current_section = line
                        current_items = []
                    elif line.startswith('вҖў') or line.startswith('-'):
                        # Item de liste
                        item = re.sub(r'^[вҖў\-]\s*', '', line)
                        current_items.append(item)
                    elif ':' in line and not line.startswith(' '):
                        # Ligne clГ© : valeur
                        current_items.append(line)
                
                if current_section:
                    sections[current_section] = current_items
                
                return sections
    
            # ---------- DonnГ©es opГ©rateur ----------
            nom = prenom = matricule = statut = "-"
            try:
                row = PersonnelRepository.get_info_basique(self.operateur_id)
                if row:
                    nom, prenom, matricule, statut = row['nom'], row['prenom'], row['matricule'], row['statut']
            except Exception:
                pass

            # ---------- Fichier ----------
            default_name = f"profil_{(matricule or 'operateur')}_{_dt.date.today():%Y%m%d}.pdf"
            file_path, _ = QFileDialog.getSaveFileName(self, "Exporter en PDF", default_name, "PDF (*.pdf)")
            if not file_path:
                return
    
            # ---------- Styles amГ©liorГ©s ----------
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=18,
                textColor=colors.HexColor("#1e40af"),
                spaceAfter=8,
                fontName="Helvetica-Bold"
            )
            
            section_style = ParagraphStyle(
                "CustomSection",
                parent=styles["Heading2"],
                fontSize=13,
                textColor=colors.HexColor("#1e40af"),
                spaceBefore=14,
                spaceAfter=8,
                fontName="Helvetica-Bold",
                borderWidth=0,
                borderColor=colors.HexColor("#3b82f6"),
                borderPadding=4
            )
            
            subsection_style = ParagraphStyle(
                "CustomSubsection",
                parent=styles["Heading3"],
                fontSize=11,
                textColor=colors.HexColor("#475569"),
                spaceBefore=8,
                spaceAfter=4,
                fontName="Helvetica-Bold"
            )
            
            body_style = ParagraphStyle(
                "CustomBody",
                parent=styles["Normal"],
                fontSize=10,
                leading=14,
                textColor=colors.HexColor("#1f2937")
            )
            
            bullet_style = ParagraphStyle(
                "CustomBullet",
                parent=body_style,
                leftIndent=20,
                bulletIndent=10,
                spaceBefore=2,
                spaceAfter=2
            )
    
            # ---------- En-tГӘte / Pied ----------
            def _header_footer(canvas: Canvas, doc):
                canvas.saveState()
                
                # Bandeau supГ©rieur
                canvas.setFillColor(colors.HexColor("#1e40af"))
                canvas.rect(0, A4[1]-40, A4[0], 40, fill=1, stroke=0)
                
                canvas.setFillColor(colors.white)
                canvas.setFont("Helvetica-Bold", 12)
                canvas.drawString(doc.leftMargin, A4[1]-25, f"{nom} {prenom}")
                
                canvas.setFont("Helvetica", 9)
                canvas.drawString(doc.leftMargin, A4[1]-38, 
                                f"Matricule : {matricule}  вҖў  Statut : {statut}")
                
                # Pied de page
                canvas.setFillColor(colors.HexColor("#64748b"))
                canvas.setFont("Helvetica", 8)
                canvas.drawCentredString(A4[0]/2, 15, f"Page {doc.page}")
                canvas.drawRightString(A4[0]-doc.rightMargin, 15, 
                                     f"GГ©nГ©rГ© le {format_date(_dt.date.today())}")
                
                canvas.restoreState()
    
            # ---------- Document ----------
            doc = SimpleDocTemplate(
                file_path, pagesize=A4,
                leftMargin=2*cm, rightMargin=2*cm, 
                topMargin=2.5*cm, bottomMargin=2*cm
            )
            flow = []
            
            # Titre principal
            flow.append(Paragraph("Fiche Personnel", title_style))
            flow.append(Spacer(1, 0.4*cm))

            # ---------- Table unifiГ©e : identitГ© + contrat + formations + validitГ©s ----------
            _sections = {t: items for t, items in (self._infos_data if hasattr(self, "_infos_data") and self._infos_data else [])}

            unified_rows = []
            header_style_idx = []  # indices des lignes d'en-tГӘte de section

            def _add_section_header(label):
                header_style_idx.append(len(unified_rows))
                unified_rows.append([label.upper(), ""])

            # IdentitГ©
            _add_section_header("IdentitГ©")
            unified_rows.append(["Nom", self.operateur_nom])
            unified_rows.append(["PrГ©nom", self.operateur_prenom])
            unified_rows.append(["Matricule", matricule or "вҖ”"])
            unified_rows.append(["Statut", statut])
            for label, value in _sections.get("Informations Personnelles", []):
                if value and "Aucune" not in str(value):
                    unified_rows.append([label, str(value)])

            # Contrat
            _add_section_header("Contrat actuel")
            contrat_items = [(l, v) for l, v in _sections.get("Contrat Actuel", []) if v and "Aucun" not in str(v)]
            if contrat_items:
                for l, v in contrat_items:
                    unified_rows.append([l, str(v)])
            else:
                unified_rows.append(["Statut", "Aucun contrat actif"])

            # Formations
            _add_section_header("Formations")
            formation_items = [(l, v) for l, v in _sections.get("Formations", []) if "Aucune" not in l]
            if formation_items:
                for l, v in formation_items:
                    unified_rows.append([l, str(v)])
            else:
                unified_rows.append(["", "Aucune formation renseignГ©e"])

            # ValiditГ©s
            _add_section_header("ValiditГ©s mГ©dicales")
            validite_items = [(l, v) for l, v in _sections.get("ValiditГ©s", []) if "Aucune" not in l]
            if validite_items:
                for l, v in validite_items:
                    unified_rows.append([l, str(v)])
            else:
                unified_rows.append(["", "Aucune validitГ© enregistrГ©e"])

            t_unified = Table(unified_rows, colWidths=[5*cm, 12*cm])
            base_style = [
                ("FONTSIZE", (0,0), (-1,-1), 9),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
                ("LEFTPADDING", (0,0), (-1,-1), 8),
                ("RIGHTPADDING", (0,0), (-1,-1), 8),
                ("TOPPADDING", (0,0), (-1,-1), 5),
                ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
                ("TEXTCOLOR", (0,0), (0,-1), colors.HexColor("#475569")),
                ("TEXTCOLOR", (1,0), (1,-1), colors.HexColor("#1f2937")),
                ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
            for idx in header_style_idx:
                base_style += [
                    ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#e2e8f0")),
                    ("TEXTCOLOR", (0, idx), (-1, idx), colors.HexColor("#334155")),
                    ("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"),
                    ("SPAN", (0, idx), (-1, idx)),
                ]
            t_unified.setStyle(TableStyle(base_style))
            flow.append(t_unified)
            flow.append(Spacer(1, 0.5*cm))

            # ---------- Polyvalences (production uniquement, si donnГ©es) ----------
            has_poly = (
                self.is_production
                and hasattr(self, "poly_table")
                and self.poly_table.rowCount() > 0
            )
            if has_poly:
                flow.append(Spacer(1, 0.5*cm))
                flow.append(Paragraph("Polyvalences et Гүvaluations", section_style))

                poly_data = [["Poste", "Niveau", "DerniГЁre\nГ©val.", "Prochaine\nГ©val.", "AnciennetГ©", "Statut"]]
                for r in range(self.poly_table.rowCount()):
                    row_data = []
                    for c in range(6):
                        item = self.poly_table.item(r, c)
                        row_data.append(item.text() if item else "")
                    poly_data.append(row_data)

                col_widths = [4*cm, 2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm]
                t_poly = Table(poly_data, colWidths=col_widths, repeatRows=1)
                t_poly.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1e40af")),
                    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                    ("FONTSIZE", (0,0), (-1,0), 9),
                    ("ALIGN", (0,0), (-1,0), "CENTER"),
                    ("VALIGN", (0,0), (-1,0), "MIDDLE"),
                    ("FONTSIZE", (0,1), (-1,-1), 9),
                    ("ALIGN", (1,1), (-1,-1), "CENTER"),
                    ("ALIGN", (0,1), (0,-1), "LEFT"),
                    ("VALIGN", (0,1), (-1,-1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
                    ("LEFTPADDING", (0,0), (-1,-1), 6),
                    ("RIGHTPADDING", (0,0), (-1,-1), 6),
                    ("TOPPADDING", (0,0), (-1,-1), 6),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                ]))
                flow.append(t_poly)
    
            # ---------- Build ----------
            doc.build(flow, onFirstPage=_header_footer, onLaterPages=_header_footer)
            QMessageBox.information(self, "Export rГ©ussi", 
                                  f"Le profil a Г©tГ© exportГ© avec succГЁs !\n\n{file_path}")
    
        except ImportError:
            QMessageBox.warning(self, "Module manquant",
                              "Le module 'reportlab' est requis pour l'export PDF.\n\n"
                              "Installez-le avec : pip install reportlab")
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", 
                               f"Une erreur s'est produite lors de l'export PDF :\n\n{str(e)}")



    def export_profile_excel(self):
        """Export Excel lisible : RГ©sumГ© (A..S, wrap + hauteur auto), Informations, Polyvalences."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from PyQt5.QtWidgets import QFileDialog, QMessageBox
            import datetime as _dt
    
            # ----- Styles -----
            THIN = Side(style="thin", color="DDDDDD")
            BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            HDR_FILL = PatternFill("solid", fgColor="4472C4")
            HDR_FONT = Font(bold=True, color="FFFFFF")
            SEC_FILL = PatternFill("solid", fgColor="EEF2FF")
            TITLE_FONT = Font(bold=True, size=16)
            SUB_FONT = Font(bold=True, size=12, color="555555")
            CENTER = Alignment(horizontal="center", vertical="center")
            LEFT = Alignment(horizontal="left", vertical="center")
            LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
            def style_header(row):
                for c in row:
                    c.fill = HDR_FILL
                    c.font = HDR_FONT
                    c.alignment = CENTER
                    c.border = BORDER
    
            def style_table(ws, start_row, start_col, end_row, end_col, zebra=False):
                for r in range(start_row, end_row + 1):
                    fill = PatternFill("solid", fgColor="F8FAFC") if zebra and r > start_row and (r - start_row) % 2 == 1 else None
                    for c in range(start_col, end_col + 1):
                        cell = ws.cell(r, c)
                        cell.border = BORDER
                        if fill:
                            cell.fill = fill
    
            def autofit(ws, min_w=10, max_w=80):
                for col in range(1, ws.max_column + 1):
                    letter = get_column_letter(col)
                    longest = 0
                    for row in ws.iter_rows(min_col=col, max_col=col):
                        v = row[0].value
                        if v is not None:
                            longest = max(longest, len(str(v)))
                    ws.column_dimensions[letter].width = max(min(longest + 2, max_w), min_w)
    
            # estimation de hauteur pour cellule fusionnГ©e (wrap)
            def set_wrapped_row_height(ws, row_idx, merged_cols, char_width=7):
                """Calcule et dГ©finit la hauteur de ligne pour du texte wrappГ©."""
                # Largeur totale disponible en pixels (approximation)
                total_width = 0
                for col in merged_cols:
                    col_width = ws.column_dimensions[col].width or 10
                    total_width += col_width * char_width  # Conversion largeur Excel -> pixels
                
                txt = str(ws.cell(row_idx, 1).value or "")
                if not txt:
                    ws.row_dimensions[row_idx].height = 20
                    return
                
                # Calculer le nombre de lignes en tenant compte des sauts de ligne
                lines_content = txt.split('\n')
                total_lines = 0
                
                for line in lines_content:
                    if not line.strip():
                        total_lines += 1  # Ligne vide
                    else:
                        # Nombre de lignes nГ©cessaires pour cette ligne de contenu
                        line_chars = len(line)
                        chars_per_line = max(1, total_width // char_width)
                        lines_needed = max(1, (line_chars + chars_per_line - 1) // chars_per_line)
                        total_lines += lines_needed
                
                # DГ©finir la hauteur (15 points par ligne + marge)
                ws.row_dimensions[row_idx].height = max(20, total_lines * 15 + 10)
    
            # ----- En-tГӘte opГ©rateur -----
            nom = prenom = matricule = statut = "-"
            try:
                row = PersonnelRepository.get_info_basique(self.operateur_id)
                if row:
                    nom, prenom, matricule, statut = row['nom'], row['prenom'], row['matricule'], row['statut']
            except Exception:
                pass
    
            default_name = f"profil_{(matricule or 'operateur')}_{_dt.date.today():%Y%m%d}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(self, "Exporter le profil", default_name, "Excel Files (*.xlsx)")
            if not file_path:
                return
    
            wb = Workbook()
    
            # =========================================================
            # 1) FEUILLE RГүSUMГү (grande largeur A..S)
            # =========================================================
            ws1 = wb.active; ws1.title = "RГ©sumГ©"
            ws1.sheet_view.zoomScale = 120  # zoom confortable
    
            # colonnes A..S bien larges
            wide_cols = [chr(c) for c in range(ord('A'), ord('S')+1)]  # A..S
            for col in wide_cols:
                ws1.column_dimensions[col].width = 24  # large et lisible
    
            # Titre & sous-titre fusionnГ©s sur A..S
            ws1.merge_cells("A1:S1")
            ws1["A1"] = f"Profil opГ©rateur вҖ” {nom or ''} {prenom or ''}".strip()
            ws1["A1"].font = TITLE_FONT
            ws1["A1"].alignment = LEFT
    
            ws1.merge_cells("A2:S2")
            ws1["A2"] = f"Matricule : {matricule or '-'}    |    Statut : {statut or '-'}"
            ws1["A2"].font = SUB_FONT
            ws1["A2"].alignment = LEFT
    
            # Note: L'onglet rГ©sumГ© a Г©tГ© supprimГ©, donc pas de contenu ici
            ws1["A4"] = "Informations de base"
            ws1["A4"].font = Font(bold=True)
            ws1["A4"].fill = SEC_FILL
            ws1["A4"].alignment = LEFT

            ws1.freeze_panes = "A5"
    
            # =========================================================
            # 2) FEUILLE INFORMATIONS
            # =========================================================
            ws2 = wb.create_sheet("Informations")
            ws2.sheet_view.zoomScale = 120
            ws2.merge_cells("A1:D1")
            ws2["A1"] = "Informations complГ©mentaires"
            ws2["A1"].font = TITLE_FONT
            ws2["A1"].alignment = LEFT
    
            ws2.append(["CatГ©gorie", "Valeur"])
            style_header(ws2[2])

            if hasattr(self, "_infos_data") and self._infos_data:
                for section_title, items in self._infos_data:
                    # Ajouter le titre de section
                    ws2.append([section_title.upper(), ""])
                    for label, value in items:
                        if "Aucune" not in label:
                            ws2.append([f"  {label}", str(value)])
    
            style_table(ws2, 3, 1, ws2.max_row, 2, zebra=True)
            for cell in ws2["B"][2:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            autofit(ws2, min_w=18, max_w=80)
            ws2.freeze_panes = "A3"
    
            # =========================================================
            # 3) FEUILLE POLYVALENCES
            # =========================================================
            ws3 = wb.create_sheet("Polyvalences")
            ws3.sheet_view.zoomScale = 120
            ws3.merge_cells("A1:F1")
            ws3["A1"] = "Polyvalences & Гүvaluations"
            ws3["A1"].font = TITLE_FONT
            ws3["A1"].alignment = LEFT
    
            headers = ["Poste", "Niveau", "Date Гүvaluation", "Prochaine Гүval.", "AnciennetГ©", "Statut"]
            ws3.append(headers)
            style_header(ws3[2])
    
            if hasattr(self, "poly_table"):
                for row in range(self.poly_table.rowCount()):
                    def t(c):
                        it = self.poly_table.item(row, c)
                        return it.text() if it else ""
                    ws3.append([t(0), t(1), t(2), t(3), t(4), t(5)])
    
            style_table(ws3, 3, 1, ws3.max_row, 6, zebra=True)
            for row in ws3.iter_rows(min_row=3, min_col=1, max_col=6, max_row=ws3.max_row):
                for c in row:
                    c.alignment = CENTER
            for c in ws3["A"][2:]:
                c.alignment = Alignment(horizontal="left", vertical="center")
            autofit(ws3, min_w=12, max_w=28)
            ws3.freeze_panes = "A3"
    
            # ----- Save -----
            wb.save(file_path)
            QMessageBox.information(self, "Export rГ©ussi", f"Profil exportГ© :\n{file_path}")
    
        except ImportError:
            QMessageBox.warning(self, "Module manquant",
                                "Le module 'openpyxl' est requis pour l'export Excel.\n\npip install openpyxl")
        except Exception as e:
            logger.exception(f"Erreur export profil: {e}")
            show_error_message(self, "Erreur", "Impossible d'exporter le profil", e)


    
    def _format_date(self, date_val):
        """Formate une date."""
        if date_val is None:
            return "N/A"
        if isinstance(date_val, str):
            try:
                return dt.datetime.strptime(date_val, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                return date_val
        if hasattr(date_val, "strftime"):
            return date_val.strftime("%d/%m/%Y")
        return str(date_val)
    
    def _load_date_entree(self):
        """Charge la date d'entrГ©e de l'opГ©rateur depuis personnel_infos."""
        try:
            return PersonnelRepository.get_date_entree(self.operateur_id)
        except Exception:
            return None

    def _calculate_anciennete(self):
        """Calcule l'anciennetГ© basГ©e sur la date d'entrГ©e de l'opГ©rateur."""
        if self.date_entree is None:
            return "N/A"
        try:
            if isinstance(self.date_entree, str):
                date_obj = dt.datetime.strptime(self.date_entree, "%Y-%m-%d").date()
            elif hasattr(self.date_entree, "date"):
                date_obj = self.date_entree.date()
            else:
                date_obj = self.date_entree

            delta = dt.date.today() - date_obj
            years = delta.days // 365
            months = (delta.days % 365) // 30

            if years > 0:
                return f"{years} an(s) {months} mois"
            elif months > 0:
                return f"{months} mois"
            else:
                return f"{delta.days} jour(s)"
        except Exception:
            return "N/A"


class GestionPersonnelDialog(QDialog):
    """
    FenГӘtre principale de gestion du personnel.
    Affiche tous les opГ©rateurs avec filtrage par statut.
    """

    # Signal Г©mis quand un opГ©rateur change de statut (pour rafraГ®chir le dashboard)
    data_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Gestion du Personnel")
        self.setGeometry(100, 100, 1000, 600)

        # --- Гүtat pagination ---
        self._page_offset = 0
        self._page_size = 50
        self._total_count = 0

        # Debounce recherche : attend 350 ms aprГЁs la derniГЁre frappe avant de requГӘter
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(350)
        self._search_timer.timeout.connect(self._reload_page)

        # Layout principal avec marges nulles
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Barre de titre personnalisГ©e
        title_bar = add_custom_title_bar(self, "Gestion du Personnel")
        main_layout.addWidget(title_bar)

        # Widget de contenu
        content_widget = QWidget()
        layout = QVBoxLayout(content_widget)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # === Header ===
        header = QLabel("Gestion du Personnel")
        header.setFont(QFont("Segoe UI", 18, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        subtitle = QLabel("Vue complГЁte du personnel")
        subtitle.setStyleSheet("color: #6b7280; font-style: italic;")
        subtitle.setAlignment(Qt.AlignCenter)
        layout.addWidget(subtitle)
        
        # === Filtres ===
        filters_group = QGroupBox("Filtres")
        filters_layout = QVBoxLayout()
        
        # Ligne 1 : Recherche
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Recherche :"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Nom ou prГ©nom...")
        self.search_input.textChanged.connect(self._search_timer.start)
        search_layout.addWidget(self.search_input, 1)
        filters_layout.addLayout(search_layout)
        
        # Ligne 2 : Statut (Radio buttons)
        statut_layout = QHBoxLayout()
        statut_layout.addWidget(QLabel("Statut :"))

        self.radio_tous = QRadioButton("Tous")
        self.radio_actifs = QRadioButton("Actifs")
        self.radio_inactifs = QRadioButton("Inactifs")

        self.radio_tous.setChecked(True)

        self.radio_tous.toggled.connect(self._reload_page)
        self.radio_actifs.toggled.connect(self._reload_page)
        self.radio_inactifs.toggled.connect(self._reload_page)

        statut_layout.addWidget(self.radio_tous)
        statut_layout.addWidget(self.radio_actifs)
        statut_layout.addWidget(self.radio_inactifs)
        statut_layout.addStretch()

        # Checkbox production uniquement (cochГ©e par dГ©faut pour gestion_production)
        self.production_only_check = QCheckBox("Production uniquement")
        current_user = get_current_user()
        is_gestion_production = (
            current_user and
            current_user.get('role_nom') == 'gestion_production'
        )
        self.production_only_check.setChecked(is_gestion_production)
        self.production_only_check.toggled.connect(self._on_production_toggled)
        statut_layout.addWidget(self.production_only_check)

        self.refresh_btn = QPushButton("Actualiser")
        self.refresh_btn.clicked.connect(self.load_data)
        statut_layout.addWidget(self.refresh_btn)

        filters_layout.addLayout(statut_layout)
        
        filters_group.setLayout(filters_layout)
        layout.addWidget(filters_group)
        
        # === Table principale ===
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Nom", "PrГ©nom", "Matricule", "Type Contrat", "Statut",
            "Nb Postes", "Niveau 1", "Niveau 2", "Niveau 3", "Niveau 4"
        ])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        # Masquer les colonnes polyvalence par dГ©faut (visibles seulement en mode production)
        for col in range(5, 10):
            self.table.setColumnHidden(col, True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.doubleClicked.connect(self.open_detail_dialog)
        layout.addWidget(self.table, 1)
        
        # === Stats globales ===
        stats_label = QLabel("Double-cliquez sur une ligne pour voir les dГ©tails complets")
        stats_label.setStyleSheet("color: #6b7280; font-style: italic; padding: 8px;")
        stats_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(stats_label)
        
        self.total_label = QLabel("Total : 0 personne(s)")
        self.total_label.setFont(QFont("Segoe UI", 11, QFont.Bold))
        self.total_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.total_label)

        # === Navigation de pagination ===
        pagination_bar = QHBoxLayout()
        self.prev_btn = QPushButton("в—Җ PrГ©cГ©dent")
        self.prev_btn.setEnabled(False)
        self.prev_btn.clicked.connect(self._prev_page)
        self.pagination_label = QLabel("Page 1/1")
        self.pagination_label.setAlignment(Qt.AlignCenter)
        self.next_btn = QPushButton("Suivant в–¶")
        self.next_btn.setEnabled(False)
        self.next_btn.clicked.connect(self._next_page)
        pagination_bar.addWidget(self.prev_btn)
        pagination_bar.addStretch()
        pagination_bar.addWidget(self.pagination_label)
        pagination_bar.addStretch()
        pagination_bar.addWidget(self.next_btn)
        layout.addLayout(pagination_bar)

        # === Actions ===
        actions = QHBoxLayout()
        actions.addStretch()

        self.dates_entree_btn = QPushButton("Affecter dates d'entrГ©e")
        self.dates_entree_btn.setStyleSheet("""
            QPushButton {
                background: #3b82f6;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2563eb;
            }
        """)
        self.dates_entree_btn.clicked.connect(self.open_dates_entree_dialog)
        actions.addWidget(self.dates_entree_btn)

        self.export_btn = QPushButton("Exporter la liste")
        self.export_btn.clicked.connect(self.export_list)
        actions.addWidget(self.export_btn)

        self.close_btn = QPushButton("Fermer")
        self.close_btn.clicked.connect(self.close)
        actions.addWidget(self.close_btn)

        layout.addLayout(actions)

        # Ajouter le widget de contenu au layout principal
        main_layout.addWidget(content_widget)

        # Appliquer l'Г©tat initial des colonnes polyvalence
        if is_gestion_production:
            for col in range(5, 10):
                self.table.setColumnHidden(col, False)

        # Charger les donnГ©es
        self.all_data = []
        self.load_data()
    
    def _get_current_filters(self) -> dict:
        """Retourne les filtres actifs sous forme de dict pour get_paginated()."""
        filters = {}
        if self.radio_actifs.isChecked():
            filters['statut'] = 'ACTIF'
        elif self.radio_inactifs.isChecked():
            filters['statut'] = 'INACTIF'
        search = self.search_input.text().strip()
        if search:
            filters['search'] = search
        if self.production_only_check.isChecked():
            filters['production_only'] = True
        return filters

    def _reload_page(self):
        """Remet l'offset Г  0 et recharge (aprГЁs changement de filtre)."""
        self._page_offset = 0
        self.load_data()

    def _prev_page(self):
        """Passe Г  la page prГ©cГ©dente."""
        if self._page_offset >= self._page_size:
            self._page_offset -= self._page_size
            self.load_data()

    def _next_page(self):
        """Passe Г  la page suivante."""
        if self._page_offset + self._page_size < self._total_count:
            self._page_offset += self._page_size
            self.load_data()

    def _update_pagination_controls(self):
        """Met Г  jour les boutons et le label de pagination."""
        current_page = self._page_offset // self._page_size + 1
        total_pages = max(1, (self._total_count + self._page_size - 1) // self._page_size)
        self.pagination_label.setText(f"Page {current_page}/{total_pages}")
        self.total_label.setText(f"Total : {self._total_count} personne(s)")
        self.prev_btn.setEnabled(self._page_offset > 0)
        self.next_btn.setEnabled(self._page_offset + self._page_size < self._total_count)

    def load_data(self):
        """Charge la page courante de personnels avec filtres (async, cГҙtГ© serveur)."""
        offset = self._page_offset
        limit = self._page_size
        filters = self._get_current_filters()

        def fetch_data(progress_callback=None):
            return PersonnelRepository.get_paginated(offset=offset, limit=limit, filters=filters)

        def on_success(result):
            rows, total = result
            self.all_data = rows
            self._total_count = total
            self._render_table()
            self._update_pagination_controls()

        def on_error(error_msg):
            logger.error(f"Erreur chargement donnees: {error_msg}")
            show_error_message(self, "Erreur", "Impossible de charger les donnГ©es", Exception(error_msg))

        worker = DbWorker(fetch_data)
        worker.signals.result.connect(on_success)
        worker.signals.error.connect(on_error)
        DbThreadPool.start(worker)

    def _render_table(self):
        """Affiche self.all_data (dГ©jГ  filtrГ©e cГҙtГ© serveur) dans la table."""
        self.table.setRowCount(0)

        for data in self.all_data:
            statut = data.get("statut", "").upper()
            row = self.table.rowCount()
            self.table.insertRow(row)

            # Col 0 : Nom (avec ID et statut en UserRole)
            nom_item = QTableWidgetItem(data.get("nom", ""))
            nom_item.setData(Qt.UserRole, data.get("id"))
            nom_item.setData(Qt.UserRole + 1, statut)
            self.table.setItem(row, 0, nom_item)

            # Col 1 : PrГ©nom
            self.table.setItem(row, 1, QTableWidgetItem(data.get("prenom", "")))

            # Col 2 : Matricule
            self.table.setItem(row, 2, QTableWidgetItem(data.get("matricule") or "-"))

            # Col 3 : Type Contrat
            self.table.setItem(row, 3, QTableWidgetItem(data.get("type_contrat") or "-"))

            # Col 4 : Statut avec couleur
            statut_item = QTableWidgetItem(statut)
            statut_item.setTextAlignment(Qt.AlignCenter)
            if statut == "ACTIF":
                statut_item.setBackground(QColor("#d1fae5"))
                statut_item.setForeground(QColor("#065f46"))
            else:
                statut_item.setBackground(QColor("#fee2e2"))
                statut_item.setForeground(QColor("#991b1b"))
            self.table.setItem(row, 4, statut_item)

            # Col 5 : Nb Postes
            self.table.setItem(row, 5, QTableWidgetItem(str(data.get("nb_postes", 0))))

            # Col 6-9 : Stats polyvalence avec couleurs
            n1_item = QTableWidgetItem(str(data.get("n1", 0)))
            n1_item.setBackground(QColor("#fef2f2"))
            n1_item.setForeground(QColor("#dc2626"))
            n1_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 6, n1_item)

            n2_item = QTableWidgetItem(str(data.get("n2", 0)))
            n2_item.setBackground(QColor("#fffbeb"))
            n2_item.setForeground(QColor("#d97706"))
            n2_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 7, n2_item)

            n3_item = QTableWidgetItem(str(data.get("n3", 0)))
            n3_item.setBackground(QColor("#f0fdf4"))
            n3_item.setForeground(QColor("#059669"))
            n3_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 8, n3_item)

            n4_item = QTableWidgetItem(str(data.get("n4", 0)))
            n4_item.setBackground(QColor("#eff6ff"))
            n4_item.setForeground(QColor("#2563eb"))
            n4_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 9, n4_item)

    def _on_production_toggled(self, checked):
        """Bascule les colonnes polyvalence et recharge quand la checkbox Production change."""
        for col in range(5, 10):
            self.table.setColumnHidden(col, not checked)
        self._reload_page()
    
    def open_detail_dialog(self):
        """Ouvre la fenГӘtre de dГ©tails pour l'opГ©rateur sГ©lectionnГ©."""
        selected = self.table.selectedItems()
        if not selected:
            return
        
        operateur_id = selected[0].data(Qt.UserRole)
        statut = selected[0].data(Qt.UserRole + 1)
        nom = selected[0].text()
        prenom = selected[1].text()

        # DГ©terminer si l'opГ©rateur est en production
        is_production = False
        for data in self.all_data:
            if data.get("id") == operateur_id:
                is_production = (data.get("numposte") or "") == "Production"
                break

        detail_dialog = DetailOperateurDialog(operateur_id, nom, prenom, statut, self, is_production=is_production)
        detail_dialog.operateur_status_changed.connect(self.on_operateur_status_changed)
        detail_dialog.exec_()
    
    def on_operateur_status_changed(self, operateur_id):
        """Callback quand le statut d'un opГ©rateur est modifiГ©."""
        self.load_data()
        self.data_changed.emit()

    def open_dates_entree_dialog(self):
        """Ouvre le dialogue d'affectation des dates d'entrГ©e"""
        dialog = AffecterDatesEntreeDialog(self)
        dialog.exec_()

    def export_list(self):
        """Exporte la liste en Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from PyQt5.QtWidgets import QFileDialog
            
            # DГ©terminer le nom du fichier selon les filtres actifs
            date_str = dt.date.today().strftime('%Y%m%d')
            parts = ["personnel"]

            if self.production_only_check.isChecked():
                parts.append("production")

            if self.radio_actifs.isChecked():
                parts.append("actifs")
            elif self.radio_inactifs.isChecked():
                parts.append("inactifs")

            parts.append(date_str)
            default_name = "_".join(parts) + ".xlsx"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Exporter la liste",
                default_name,
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Personnel"
            
            # En-tГӘtes selon le mode
            production_only = self.production_only_check.isChecked()
            headers = ["Nom", "PrГ©nom", "Matricule", "Type Contrat", "Statut"]
            if production_only:
                headers += ["Nb Postes", "Niveau 1", "Niveau 2", "Niveau 3", "Niveau 4"]
            ws.append(headers)

            # Style en-tГӘtes
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # DonnГ©es : rГ©cupГ©rer TOUS les rГ©sultats filtrГ©s (pas seulement la page courante)
            all_matching, _ = PersonnelRepository.get_paginated(
                offset=0, limit=10_000, filters=self._get_current_filters()
            )

            for data in all_matching:
                statut = data.get("statut", "").upper()
                row_data = [
                    data.get("nom", ""),
                    data.get("prenom", ""),
                    data.get("matricule") or "-",
                    data.get("type_contrat") or "-",
                    statut,
                ]
                if production_only:
                    row_data += [
                        data.get("nb_postes", 0),
                        data.get("n1", 0),
                        data.get("n2", 0),
                        data.get("n3", 0),
                        data.get("n4", 0),
                    ]
                ws.append(row_data)

                # Colorer la colonne statut (colonne 5)
                row_idx = ws.max_row
                statut_cell = ws.cell(row=row_idx, column=5)
                if statut == "ACTIF":
                    statut_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    statut_cell.font = Font(color="065F46", bold=True)
                else:
                    statut_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    statut_cell.font = Font(color="991B1B", bold=True)
                statut_cell.alignment = Alignment(horizontal="center")
            
            # Ajuster les largeurs
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
            QMessageBox.information(
                self, "Export rГ©ussi",
                f"Liste exportГ©e avec succГЁs !\n\n{file_path}"
            )
            
        except ImportError:
            QMessageBox.warning(
                self, "Module manquant",
                "Le module 'openpyxl' est requis pour l'export Excel.\n\n"
                "Installez-le avec : pip install openpyxl"
            )
        except Exception as e:
            logger.exception(f"Erreur export: {e}")
            show_error_message(self, "Erreur", "Impossible d'exporter", e)


class AffecterDatesEntreeDialog(QDialog):
    """Dialogue pour affecter les dates d'entrГ©e aux employГ©s sans date"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Affectation des dates d'entrГ©e")
        self.setMinimumWidth(1000)
        self.setMinimumHeight(600)

        self.date_widgets = {}  # Dictionnaire pour stocker les QDateEdit par operateur_id
        self.init_ui()
        self.load_personnel_sans_date()

    def init_ui(self):
        """Initialise l'interface"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Header
        header = QLabel("Affectation des dates d'entrГ©e")
        header.setFont(QFont("Segoe UI", 16, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        subtitle = QLabel("Saisissez la date d'entrГ©e pour chaque employГ© puis cliquez sur 'Enregistrer'")
        subtitle.setStyleSheet("color: #6b7280; font-style: italic;")
        subtitle.setAlignment(Qt.AlignCenter)
        layout.addWidget(subtitle)

        # Stats
        self.stats_label = QLabel()
        self.stats_label.setStyleSheet("color: #3b82f6; font-weight: bold; padding: 10px;")
        self.stats_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.stats_label)

        # Table des employГ©s
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Nom", "PrГ©nom", "Matricule", "Statut", "Date d'entrГ©e", "Action"
        ])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)  # Nom
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)  # PrГ©nom
        self.table.setColumnWidth(0, 150)  # Nom
        self.table.setColumnWidth(3, 100)  # Statut
        self.table.setColumnWidth(4, 150)  # Date
        self.table.setColumnWidth(5, 120)  # Action
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #f9fafb;
                gridline-color: #e5e7eb;
                border: 1px solid #d1d5db;
            }
            QHeaderView::section {
                background-color: #f3f4f6;
                color: #374151;
                font-weight: 600;
                padding: 8px;
                border: none;
            }
        """)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)

        # Boutons d'actions
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        self.close_btn = QPushButton("Fermer")
        self.close_btn.clicked.connect(self.close)
        self.close_btn.setStyleSheet("""
            QPushButton {
                background: #ef4444;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background: #dc2626;
            }
        """)
        btn_layout.addWidget(self.close_btn)

        layout.addLayout(btn_layout)

    def load_personnel_sans_date(self):
        """Charge les employГ©s sans date d'entrГ©e"""
        try:
            personnel = PersonnelRepository.get_personnel_sans_date_entree()

            # Mettre Г  jour les stats
            self.stats_label.setText(
                f"{len(personnel)} employГ©(s) sans date d'entrГ©e"
            )

            # Remplir la table
            self.table.setRowCount(len(personnel))
            self.date_widgets.clear()

            for row, emp in enumerate(personnel):
                operateur_id = emp['id']

                # Nom
                nom_item = QTableWidgetItem(emp['nom'])
                nom_item.setData(Qt.UserRole, operateur_id)  # Stocker l'ID
                self.table.setItem(row, 0, nom_item)

                # PrГ©nom
                prenom_item = QTableWidgetItem(emp['prenom'])
                self.table.setItem(row, 1, prenom_item)

                # Matricule
                matricule_item = QTableWidgetItem(emp['matricule'] or "вҖ”")
                self.table.setItem(row, 2, matricule_item)

                # Statut
                statut = emp['statut']
                statut_item = QTableWidgetItem(statut)
                if statut == "ACTIF":
                    statut_item.setForeground(QColor("#10b981"))
                else:
                    statut_item.setForeground(QColor("#6b7280"))
                self.table.setItem(row, 3, statut_item)

                # Date d'entrГ©e - QDateEdit
                date_edit = QDateEdit()
                date_edit.setCalendarPopup(True)
                date_edit.setDate(QDate.currentDate())
                date_edit.setDisplayFormat("dd/MM/yyyy")
                date_edit.setMinimumDate(QDate(1950, 1, 1))
                date_edit.setMaximumDate(QDate.currentDate())
                date_edit.setStyleSheet("""
                    QDateEdit {
                        padding: 6px;
                        border: 1px solid #d1d5db;
                        border-radius: 4px;
                    }
                    QDateEdit:focus {
                        border-color: #3b82f6;
                    }
                """)
                self.table.setCellWidget(row, 4, date_edit)
                self.date_widgets[operateur_id] = date_edit

                # Bouton Enregistrer
                btn_enregistrer = QPushButton("Enregistrer")
                btn_enregistrer.setStyleSheet("""
                    QPushButton {
                        background: #10b981;
                        color: white;
                        padding: 6px 12px;
                        border-radius: 4px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: #059669;
                    }
                """)
                btn_enregistrer.clicked.connect(
                    lambda checked, oid=operateur_id, r=row: self.enregistrer_date(oid, r)
                )
                self.table.setCellWidget(row, 5, btn_enregistrer)

        except Exception as e:
            logger.exception(f"Erreur chargement: {e}")
            show_error_message(self, "Erreur", "Erreur lors du chargement", e)

    def enregistrer_date(self, operateur_id, row):
        """Enregistre la date d'entrГ©e pour un employГ© spГ©cifique"""
        try:
            # RГ©cupГ©rer les informations
            nom = self.table.item(row, 0).text()
            prenom = self.table.item(row, 1).text()
            date_widget = self.date_widgets.get(operateur_id)

            if not date_widget:
                return

            date_entree = date_widget.date().toString("yyyy-MM-dd")
            date_display = date_widget.date().toString("dd/MM/yyyy")

            # Confirmer
            reply = QMessageBox.question(
                self,
                "Confirmation",
                f"Affecter la date {date_display} Г  {nom} {prenom} ?",
                QMessageBox.Yes | QMessageBox.No
            )

            if reply != QMessageBox.Yes:
                return

            # Enregistrer dans la base
            try:
                PersonnelRepository.save_date_entree(operateur_id, date_entree)

                # Logger
                log_hist(
                    "AFFECTATION_DATE_ENTREE",
                    f"Date d'entrГ©e affectГ©e: {date_display}",
                    operateur_id,
                    None
                )

                QMessageBox.information(
                    self,
                    "SuccГЁs",
                    f"Date d'entrГ©e enregistrГ©e pour {nom} {prenom}"
                )

                # Retirer la ligne du tableau
                self.table.removeRow(row)
                del self.date_widgets[operateur_id]

                # Mettre Г  jour les stats
                nb_restants = self.table.rowCount()
                self.stats_label.setText(f"{nb_restants} employГ©(s) sans date d'entrГ©e")

                if nb_restants == 0:
                    QMessageBox.information(
                        self,
                        "TerminГ©",
                        "Tous les employГ©s ont maintenant une date d'entrГ©e!"
                    )
                    self.close()

            except Exception as e:
                logger.exception(f"Erreur enregistrement: {e}")
                show_error_message(self, "Erreur", "Erreur lors de l'enregistrement", e)

        except Exception as e:
            logger.exception(f"Erreur: {e}")
            show_error_message(self, "Erreur", "Une erreur est survenue", e)


if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication
    
    app = QApplication(sys.argv)
    dialog = GestionPersonnelDialog()
    dialog.show()
    sys.exit(app.exec_())