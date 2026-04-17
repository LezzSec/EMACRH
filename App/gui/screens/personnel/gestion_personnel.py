# -*- coding: utf-8 -*-
# gestion_personnel.py – Fenêtre principale de gestion du personnel.

from infrastructure.logging.logging_config import get_logger
logger = get_logger(__name__)

from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QLineEdit, QGroupBox,
    QMessageBox, QAbstractItemView, QWidget, QComboBox,
)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal
from PyQt5.QtGui import QFont, QColor

from domain.repositories.personnel_repo import PersonnelRepository
from gui.workers.db_worker import DbWorker, DbThreadPool
from gui.components.emac_ui_kit import add_custom_title_bar, show_error_message
from gui.view_models.personnel_list_view_model import PersonnelListViewModel
from domain.services.admin.auth_service import get_current_user

from gui.screens.personnel.detail_operateur_dialog import DetailOperateurDialog

import datetime as dt


class GestionPersonnelDialog(QDialog):
    """
    Fenêtre principale de gestion du personnel.
    Affiche tous les opérateurs avec filtrage par statut.
    """

    data_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Gestion du Personnel")
        self.setGeometry(100, 100, 1000, 600)

        self.vm = PersonnelListViewModel(page_size=50)

        # Debounce recherche : attend 350 ms après la dernière frappe avant de requêter
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(350)
        self._search_timer.timeout.connect(self._reload_page)

        # Layout principal avec marges nulles
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Barre de titre personnalisée
        title_bar = add_custom_title_bar(self, "Gestion du Personnel")
        main_layout.addWidget(title_bar)

        # Widget de contenu
        content_widget = QWidget()
        layout = QVBoxLayout(content_widget)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # === Header ===
        header = QLabel("Gestion du Personnel")
        header.setFont(QFont("Segoe UI", 18, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        subtitle = QLabel("Vue complète du personnel")
        subtitle.setStyleSheet("color: #6b7280; font-style: italic;")
        subtitle.setAlignment(Qt.AlignCenter)
        layout.addWidget(subtitle)

        # === Filtres ===
        filters_group = QGroupBox("Filtres")
        filters_layout = QVBoxLayout()

        # Ligne 1 : Recherche
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Recherche :"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Nom ou prénom...")
        self.search_input.textChanged.connect(self._search_timer.start)
        search_layout.addWidget(self.search_input, 1)
        filters_layout.addLayout(search_layout)

        # Ligne 2 : Vue (combo preset) + bouton actualiser discret
        vue_layout = QHBoxLayout()
        vue_layout.addWidget(QLabel("Vue :"))

        current_user = get_current_user()
        is_gestion_production = (
            current_user and
            current_user.get('role_nom') == 'gestion_production'
        )

        self.vue_combo = QComboBox()
        self.vue_combo.addItem("Actifs en production", ("ACTIF", True))
        self.vue_combo.addItem("Tous les actifs", ("ACTIF", False))
        self.vue_combo.addItem("Inactifs", ("INACTIF", False))
        self.vue_combo.addItem("Personnel administratif", ("ACTIF", False))
        self.vue_combo.addItem("Tout le personnel", (None, False))
        self.vue_combo.setCurrentIndex(0 if is_gestion_production else 1)
        self.vue_combo.setMinimumWidth(220)
        self.vue_combo.currentIndexChanged.connect(self._on_vue_changed)
        vue_layout.addWidget(self.vue_combo)

        vue_layout.addStretch()

        self.refresh_btn = QPushButton("\u27f3")
        self.refresh_btn.setToolTip("Actualiser les données depuis le serveur")
        self.refresh_btn.setFixedWidth(36)
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background: transparent;
                color: #6b7280;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                font-size: 16px;
            }
            QPushButton:hover { background: #f3f4f6; color: #1e40af; }
        """)
        self.refresh_btn.clicked.connect(self.load_data)
        vue_layout.addWidget(self.refresh_btn)

        filters_layout.addLayout(vue_layout)

        filters_group.setLayout(filters_layout)
        layout.addWidget(filters_group)

        # === Table principale ===
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Nom", "Prénom", "Matricule", "Type Contrat", "Statut",
            "Nb Postes", "Niveau 1", "Niveau 2", "Niveau 3", "Niveau 4"
        ])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        # Masquer les colonnes polyvalence par défaut (visibles seulement en mode production)
        for col in range(5, 10):
            self.table.setColumnHidden(col, True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.doubleClicked.connect(self.open_detail_dialog)
        self.table.itemSelectionChanged.connect(self._on_selection_changed)
        layout.addWidget(self.table, 1)

        self.total_label = QLabel("Total : 0 personne(s)")
        self.total_label.setFont(QFont("Segoe UI", 11, QFont.Bold))
        self.total_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.total_label)

        # === Navigation de pagination ===
        pagination_bar = QHBoxLayout()
        self.prev_btn = QPushButton("◀ Précédent")
        self.prev_btn.setEnabled(False)
        self.prev_btn.clicked.connect(self._prev_page)
        self.pagination_label = QLabel("Page 1/1")
        self.pagination_label.setAlignment(Qt.AlignCenter)
        self.next_btn = QPushButton("Suivant ▶")
        self.next_btn.setEnabled(False)
        self.next_btn.clicked.connect(self._next_page)
        pagination_bar.addWidget(self.prev_btn)
        pagination_bar.addStretch()
        pagination_bar.addWidget(self.pagination_label)
        pagination_bar.addStretch()
        pagination_bar.addWidget(self.next_btn)
        layout.addLayout(pagination_bar)

        # === Actions ===
        actions = QHBoxLayout()

        self.detail_btn = QPushButton("Voir le détail")
        self.detail_btn.setEnabled(False)
        self.detail_btn.setToolTip("Sélectionnez une ligne pour afficher le détail")
        self.detail_btn.clicked.connect(self.open_detail_dialog)
        actions.addWidget(self.detail_btn)

        actions.addStretch()

        self.export_btn = QPushButton("Exporter la liste")
        self.export_btn.clicked.connect(self.export_list)
        actions.addWidget(self.export_btn)

        self.close_btn = QPushButton("Fermer")
        self.close_btn.clicked.connect(self.close)
        actions.addWidget(self.close_btn)

        layout.addLayout(actions)
        main_layout.addWidget(content_widget)

        # Appliquer l'état initial des colonnes polyvalence selon la vue sélectionnée
        _, production_only_init = self.vue_combo.currentData()
        for col in range(5, 10):
            self.table.setColumnHidden(col, not production_only_init)

        self.all_data = []
        self.load_data()

    # ------------------------------------------------------------------
    # Synchronisation des filtres UI → ViewModel
    # ------------------------------------------------------------------

    def _sync_filters(self) -> None:
        """Lit les filtres depuis l'UI et les pousse dans le ViewModel."""
        statut, production_only = self.vue_combo.currentData()
        self.vm.set_filters(
            statut=statut,
            search=self.search_input.text(),
            production_only=production_only,
        )

    def _on_vue_changed(self) -> None:
        """Bascule les colonnes polyvalence et recharge quand la vue change."""
        _, production_only = self.vue_combo.currentData()
        for col in range(5, 10):
            self.table.setColumnHidden(col, not production_only)
        self._reload_page()

    def _on_selection_changed(self) -> None:
        """Active/désactive le bouton Voir le détail selon la sélection."""
        self.detail_btn.setEnabled(bool(self.table.selectedItems()))

    # ------------------------------------------------------------------
    # Navigation de page
    # ------------------------------------------------------------------

    def _reload_page(self) -> None:
        """Remet l'offset à 0 et recharge (après changement de filtre)."""
        self._sync_filters()
        self.vm.reset_page()
        self.load_data()

    def _prev_page(self) -> None:
        if self.vm.go_to_prev_page():
            self.load_data()

    def _next_page(self) -> None:
        if self.vm.go_to_next_page():
            self.load_data()

    def _update_pagination_controls(self) -> None:
        """Met à jour les boutons et le label de pagination."""
        self.pagination_label.setText(
            f"Page {self.vm.current_page}/{self.vm.total_pages}"
        )
        self.total_label.setText(f"Total : {self.vm.total_count} personne(s)")
        self.prev_btn.setEnabled(self.vm.has_prev)
        self.next_btn.setEnabled(self.vm.has_next)

    # ------------------------------------------------------------------
    # Chargement des données
    # ------------------------------------------------------------------

    def load_data(self) -> None:
        """Charge la page courante de personnels avec filtres (async, côté serveur)."""
        self._sync_filters()

        def on_success(result):
            rows, _ = result
            self.all_data = rows
            self._render_table()
            self._update_pagination_controls()

        def on_error(error_msg):
            logger.error(f"Erreur chargement donnees: {error_msg}")
            show_error_message(self, "Erreur", "Impossible de charger les données", Exception(error_msg))

        worker = DbWorker(self.vm.fetch_page)
        worker.signals.result.connect(on_success)
        worker.signals.error.connect(on_error)
        DbThreadPool.start(worker)

    def _render_table(self) -> None:
        """Affiche self.all_data (déjà filtrée côté serveur) dans la table."""
        self.table.setRowCount(0)

        for data in self.all_data:
            statut = data.get("statut", "").upper()
            row = self.table.rowCount()
            self.table.insertRow(row)

            # Col 0 : Nom (avec ID et statut en UserRole)
            nom_item = QTableWidgetItem(data.get("nom", ""))
            nom_item.setData(Qt.UserRole, data.get("id"))
            nom_item.setData(Qt.UserRole + 1, statut)
            self.table.setItem(row, 0, nom_item)

            # Col 1 : Prénom
            self.table.setItem(row, 1, QTableWidgetItem(data.get("prenom", "")))

            # Col 2 : Matricule
            self.table.setItem(row, 2, QTableWidgetItem(data.get("matricule") or "-"))

            # Col 3 : Type Contrat
            self.table.setItem(row, 3, QTableWidgetItem(data.get("type_contrat") or "-"))

            # Col 4 : Statut avec couleur
            statut_item = QTableWidgetItem(statut)
            statut_item.setTextAlignment(Qt.AlignCenter)
            if statut == "ACTIF":
                statut_item.setBackground(QColor("#d1fae5"))
                statut_item.setForeground(QColor("#065f46"))
            else:
                statut_item.setBackground(QColor("#fee2e2"))
                statut_item.setForeground(QColor("#991b1b"))
            self.table.setItem(row, 4, statut_item)

            # Col 5 : Nb Postes
            self.table.setItem(row, 5, QTableWidgetItem(str(data.get("nb_postes", 0))))

            # Col 6-9 : Stats polyvalence avec couleurs
            n1_item = QTableWidgetItem(str(data.get("n1", 0)))
            n1_item.setBackground(QColor("#fef2f2"))
            n1_item.setForeground(QColor("#dc2626"))
            n1_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 6, n1_item)

            n2_item = QTableWidgetItem(str(data.get("n2", 0)))
            n2_item.setBackground(QColor("#fffbeb"))
            n2_item.setForeground(QColor("#d97706"))
            n2_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 7, n2_item)

            n3_item = QTableWidgetItem(str(data.get("n3", 0)))
            n3_item.setBackground(QColor("#f0fdf4"))
            n3_item.setForeground(QColor("#059669"))
            n3_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 8, n3_item)

            n4_item = QTableWidgetItem(str(data.get("n4", 0)))
            n4_item.setBackground(QColor("#eff6ff"))
            n4_item.setForeground(QColor("#2563eb"))
            n4_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 9, n4_item)

    # ------------------------------------------------------------------
    # Actions utilisateur
    # ------------------------------------------------------------------

    def open_detail_dialog(self) -> None:
        """Ouvre la fenêtre de détails pour l'opérateur sélectionné."""
        selected = self.table.selectedItems()
        if not selected:
            return

        operateur_id = selected[0].data(Qt.UserRole)
        statut = selected[0].data(Qt.UserRole + 1)
        nom = selected[0].text()
        prenom = selected[1].text()

        is_production = False
        for data in self.all_data:
            if data.get("id") == operateur_id:
                is_production = (data.get("numposte") or "") == "Production"
                break

        detail_dialog = DetailOperateurDialog(operateur_id, nom, prenom, statut, self, is_production=is_production)
        detail_dialog.operateur_status_changed.connect(self.on_operateur_status_changed)
        detail_dialog.exec_()

    def on_operateur_status_changed(self, operateur_id: int) -> None:
        """Callback quand le statut d'un opérateur est modifié."""
        self.load_data()
        self.data_changed.emit()


    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def export_list(self) -> None:
        """Exporte la liste en Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from PyQt5.QtWidgets import QFileDialog

            date_str = dt.date.today().strftime('%Y%m%d')
            parts = ["personnel"]

            statut_key, production_only = self.vue_combo.currentData()
            if production_only:
                parts.append("production")
            if statut_key == "ACTIF":
                parts.append("actifs")
            elif statut_key == "INACTIF":
                parts.append("inactifs")

            parts.append(date_str)
            default_name = "_".join(parts) + ".xlsx"

            file_path, _ = QFileDialog.getSaveFileName(
                self, "Exporter la liste",
                default_name,
                "Excel Files (*.xlsx)"
            )
            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Personnel"

            _, production_only = self.vue_combo.currentData()
            headers = ["Nom", "Prénom", "Matricule", "Type Contrat", "Statut"]
            if production_only:
                headers += ["Nb Postes", "Niveau 1", "Niveau 2", "Niveau 3", "Niveau 4"]
            ws.append(headers)

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Récupérer TOUS les résultats filtrés (pas seulement la page courante)
            all_matching, _ = PersonnelRepository.get_paginated(
                offset=0, limit=10_000, filters=self.vm.get_current_filters()
            )

            for data in all_matching:
                statut = data.get("statut", "").upper()
                row_data = [
                    data.get("nom", ""),
                    data.get("prenom", ""),
                    data.get("matricule") or "-",
                    data.get("type_contrat") or "-",
                    statut,
                ]
                if production_only:
                    row_data += [
                        data.get("nb_postes", 0),
                        data.get("n1", 0),
                        data.get("n2", 0),
                        data.get("n3", 0),
                        data.get("n4", 0),
                    ]
                ws.append(row_data)

                row_idx = ws.max_row
                statut_cell = ws.cell(row=row_idx, column=5)
                if statut == "ACTIF":
                    statut_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    statut_cell.font = Font(color="065F46", bold=True)
                else:
                    statut_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    statut_cell.font = Font(color="991B1B", bold=True)
                statut_cell.alignment = Alignment(horizontal="center")

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except Exception:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

            wb.save(file_path)

            QMessageBox.information(
                self, "Export réussi",
                f"Liste exportée avec succès !\n\n{file_path}"
            )

        except Exception as e:
            logger.exception(f"Erreur export: {e}")
            show_error_message(self, "Erreur", "Impossible d'exporter", e)


if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication

    app = QApplication(sys.argv)
    dialog = GestionPersonnelDialog()
    dialog.show()
    sys.exit(app.exec_())
