
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QLineEdit, QComboBox, QGroupBox,
    QMessageBox, QAbstractItemView, QWidget, QTextEdit, QTabWidget, QCheckBox
)
from PyQt5.QtCore import Qt, QDate, pyqtSignal
from PyQt5.QtGui import QFont, QColor

from core.db.configbd import get_connection as get_db_connection
from core.services.logger import log_hist

import datetime as dt


# -------- Helpers DB --------
def _cursor(conn):
    """Retourne (cursor, dict_mode)."""
    try:
        cur = conn.cursor(dictionary=True)
        return cur, True
    except TypeError:
        cur = conn.cursor()
        return cur, False


def _rows(cur, dict_mode):
    """Retourne une liste de dicts quelle que soit la lib DB."""
    if dict_mode:
        return cur.fetchall()
    names = [d[0] for d in cur.description]
    return [dict(zip(names, r)) for r in cur.fetchall()]


class DetailOperateurDialog(QDialog):
    """Fenêtre modale affichant tous les détails d'un opérateur inactif."""
    
    operateur_reactivated = pyqtSignal(int)  # Signal émis si réactivation
    
    def __init__(self, operateur_id, nom, prenom, parent=None):
        super().__init__(parent)
        self.operateur_id = operateur_id
        self.setWindowTitle(f"Détails - {nom} {prenom}")
        self.setGeometry(200, 150, 900, 600)
        
        layout = QVBoxLayout(self)
        
        # === Header avec infos opérateur ===
        header = QLabel(f" {nom} {prenom}")
        header.setFont(QFont("Arial", 18, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        status_label = QLabel("Statut : INACTIF")
        status_label.setStyleSheet("color: #dc2626; font-weight: bold; font-size: 14px;")
        status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(status_label)
        
        # === Onglets ===
        tabs = QTabWidget()
        
        # Onglet 1 : Polyvalences
        poly_tab = QWidget()
        poly_layout = QVBoxLayout(poly_tab)
        
        poly_label = QLabel(" Polyvalences et Compétences")
        poly_label.setFont(QFont("Arial", 12, QFont.Bold))
        poly_layout.addWidget(poly_label)
        
        self.poly_table = QTableWidget()
        self.poly_table.setColumnCount(6)
        self.poly_table.setHorizontalHeaderLabels([
            "Poste", "Niveau", "Date Évaluation", "Prochaine Éval.", "Ancienneté", "Statut"
        ])
        self.poly_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.poly_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        poly_layout.addWidget(self.poly_table)
        
        # Stats des niveaux
        stats_box = QHBoxLayout()
        self.stat_n1 = self._create_mini_stat("N1", "0", "#ef4444")
        self.stat_n2 = self._create_mini_stat("N2", "0", "#f59e0b")
        self.stat_n3 = self._create_mini_stat("N3", "0", "#10b981")
        self.stat_n4 = self._create_mini_stat("N4", "0", "#3b82f6")
        self.stat_total = self._create_mini_stat("Total", "0", "#6b7280")
        
        stats_box.addWidget(self.stat_n1)
        stats_box.addWidget(self.stat_n2)
        stats_box.addWidget(self.stat_n3)
        stats_box.addWidget(self.stat_n4)
        stats_box.addWidget(self.stat_total)
        poly_layout.addLayout(stats_box)
        
        tabs.addTab(poly_tab, "Polyvalences")
        
        # Onglet 2 : Résumé / Notes
        summary_tab = QWidget()
        summary_layout = QVBoxLayout(summary_tab)
        
        summary_label = QLabel(" Résumé du Parcours")
        summary_label.setFont(QFont("Arial", 12, QFont.Bold))
        summary_layout.addWidget(summary_label)
        
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        summary_layout.addWidget(self.summary_text)
        
        tabs.addTab(summary_tab, "Résumé")
        
        # Onglet 3 : Historique des modifications
        history_tab = QWidget()
        history_layout = QVBoxLayout(history_tab)
        
        history_label = QLabel(" Historique")
        history_label.setFont(QFont("Arial", 12, QFont.Bold))
        history_layout.addWidget(history_label)
        
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(3)
        self.history_table.setHorizontalHeaderLabels(["Date", "Action", "Description"])
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.history_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        history_layout.addWidget(self.history_table)
        
        tabs.addTab(history_tab, "Historique")
        
        layout.addWidget(tabs, 1)
        
        # === Boutons d'action ===
        actions = QHBoxLayout()
        
        self.reactivate_btn = QPushButton(" Réactiver l'opérateur")
        self.reactivate_btn.setStyleSheet("""
            QPushButton {
                background: #10b981;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background: #059669;
            }
        """)
        self.reactivate_btn.clicked.connect(self.reactivate_operateur)
        actions.addWidget(self.reactivate_btn)
        
        actions.addStretch()
        
        self.export_btn = QPushButton(" Exporter le profil")
        self.export_btn.clicked.connect(self.export_profile)
        actions.addWidget(self.export_btn)
        
        self.close_btn = QPushButton("Fermer")
        self.close_btn.clicked.connect(self.close)
        actions.addWidget(self.close_btn)
        
        layout.addLayout(actions)
        
        # Charger les données
        self.load_data()
    
    def _create_mini_stat(self, label, value, color):
        """Crée une mini-carte de statistique."""
        widget = QWidget()
        widget.setStyleSheet(f"""
            QWidget {{
                background: {color};
                border-radius: 6px;
                padding: 8px;
            }}
        """)
        
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(2)
        
        val_label = QLabel(value)
        val_label.setFont(QFont("Arial", 16, QFont.Bold))
        val_label.setStyleSheet("color: white;")
        val_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(val_label)
        
        text_label = QLabel(label)
        text_label.setStyleSheet("color: white; font-size: 10px;")
        text_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(text_label)
        
        widget.value_label = val_label
        return widget
    
    def load_data(self):
        """Charge toutes les données de l'opérateur."""
        self.load_polyvalences()
        self.load_summary()
        self.load_history()
    
    def load_polyvalences(self):
        """Charge les polyvalences."""
        try:
            connection = get_db_connection()
            cursor, dict_mode = _cursor(connection)
            
            query = """
                SELECT 
                    ps.poste_code,
                    p.niveau,
                    p.date_evaluation,
                    p.prochaine_evaluation
                FROM polyvalence p
                JOIN postes ps ON p.poste_id = ps.id
                WHERE p.operateur_id = %s
                ORDER BY ps.poste_code
            """
            
            cursor.execute(query, (self.operateur_id,))
            rows = _rows(cursor, dict_mode)
            
            cursor.close()
            connection.close()
            
            self.poly_table.setRowCount(0)
            
            niveaux = {1: 0, 2: 0, 3: 0, 4: 0}
            
            for r in rows:
                row = self.poly_table.rowCount()
                self.poly_table.insertRow(row)
                
                poste = r.get("poste_code", "")
                niveau = r.get("niveau")
                date_eval = r.get("date_evaluation")
                date_next = r.get("prochaine_evaluation")
                
                # Poste
                self.poly_table.setItem(row, 0, QTableWidgetItem(str(poste)))
                
                # Niveau
                niveau_item = QTableWidgetItem(str(niveau) if niveau else "N/A")
                if niveau:
                    niveaux[int(niveau)] += 1
                    if niveau == 1:
                        niveau_item.setBackground(QColor("#fef2f2"))
                        niveau_item.setForeground(QColor("#dc2626"))
                    elif niveau == 2:
                        niveau_item.setBackground(QColor("#fffbeb"))
                        niveau_item.setForeground(QColor("#d97706"))
                    elif niveau == 3:
                        niveau_item.setBackground(QColor("#f0fdf4"))
                        niveau_item.setForeground(QColor("#059669"))
                    elif niveau == 4:
                        niveau_item.setBackground(QColor("#eff6ff"))
                        niveau_item.setForeground(QColor("#2563eb"))
                niveau_item.setTextAlignment(Qt.AlignCenter)
                self.poly_table.setItem(row, 1, niveau_item)
                
                # Dates
                self.poly_table.setItem(row, 2, QTableWidgetItem(self._format_date(date_eval)))
                self.poly_table.setItem(row, 3, QTableWidgetItem(self._format_date(date_next)))
                
                # Ancienneté
                anciennete = self._calculate_anciennete(date_eval)
                self.poly_table.setItem(row, 4, QTableWidgetItem(anciennete))
                
                # Statut (À jour / En retard)
                statut_item = QTableWidgetItem()
                if date_next:
                    today = dt.date.today()
                    if isinstance(date_next, str):
                        next_date = dt.datetime.strptime(date_next, "%Y-%m-%d").date()
                    elif hasattr(date_next, "date"):
                        next_date = date_next.date()
                    else:
                        next_date = date_next
                    
                    if next_date < today:
                        statut_item.setText("️ En retard")
                        statut_item.setForeground(QColor("#dc2626"))
                    else:
                        statut_item.setText(" À jour")
                        statut_item.setForeground(QColor("#059669"))
                else:
                    statut_item.setText("N/A")
                
                statut_item.setTextAlignment(Qt.AlignCenter)
                self.poly_table.setItem(row, 5, statut_item)
            
            # Mise à jour des stats
            self.stat_n1.value_label.setText(str(niveaux[1]))
            self.stat_n2.value_label.setText(str(niveaux[2]))
            self.stat_n3.value_label.setText(str(niveaux[3]))
            self.stat_n4.value_label.setText(str(niveaux[4]))
            self.stat_total.value_label.setText(str(sum(niveaux.values())))
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de charger les polyvalences :\n{e}")
    
    def load_summary(self):
        """Génère un résumé textuel du parcours."""
        try:
            connection = get_db_connection()
            cursor, dict_mode = _cursor(connection)
            
            # Récupérer les infos
            cursor.execute(
                "SELECT nom, prenom, statut FROM operateurs WHERE id = %s",
                (self.operateur_id,)
            )
            op = cursor.fetchone()
            
            if not op:
                return
            
            nom = op["nom"] if dict_mode else op[0]
            prenom = op["prenom"] if dict_mode else op[1]
            
            # Statistiques
            cursor.execute(
                """SELECT COUNT(*) as total, 
                   SUM(CASE WHEN niveau = 1 THEN 1 ELSE 0 END) as n1,
                   SUM(CASE WHEN niveau = 2 THEN 1 ELSE 0 END) as n2,
                   SUM(CASE WHEN niveau = 3 THEN 1 ELSE 0 END) as n3,
                   SUM(CASE WHEN niveau = 4 THEN 1 ELSE 0 END) as n4,
                   MIN(date_evaluation) as premiere_eval,
                   MAX(date_evaluation) as derniere_eval
                   FROM polyvalence WHERE operateur_id = %s""",
                (self.operateur_id,)
            )
            stats = cursor.fetchone()
            
            if dict_mode:
                total = stats["total"]
                n1, n2, n3, n4 = stats["n1"], stats["n2"], stats["n3"], stats["n4"]
                premiere = stats["premiere_eval"]
                derniere = stats["derniere_eval"]
            else:
                total = stats[0]
                n1, n2, n3, n4 = stats[1], stats[2], stats[3], stats[4]
                premiere = stats[5]
                derniere = stats[6]
            
            cursor.close()
            connection.close()
            
            # Générer le résumé
            summary = f"""
═══════════════════════════════════════════════
  PROFIL : {nom} {prenom}
═══════════════════════════════════════════════

 STATISTIQUES GÉNÉRALES
─────────────────────────────────────────────
• Nombre total de postes occupés : {total}
• Répartition des niveaux :
  - Niveau 1 : Opérateur nouveau à encadrer par titulaire N3/4 ou absent du poste depuis plus de 12 mois. (< 80%)     : {n1} poste(s)
  - Niveau 2 : Opérateur formé et apte à conduire le poste seul. Certaines notions sont en cours d'acquisition. (> 80%)     : {n2} poste(s)
  - Niveau 3 : Opérateur titulaire, formé, apte à conduire le poste et apte à former. (> 90%)       : {n3} poste(s)
  - Niveau 4 : N3 + Leader ou Polyvalent (maîtrise 3 postes d'une ligne : mél. interne, cylindres, conditionnement) (> 90%)     : {n4} poste(s)

 PÉRIODE D'ACTIVITÉ
─────────────────────────────────────────────
• Première évaluation : {self._format_date(premiere)}
• Dernière évaluation : {self._format_date(derniere)}

 POINTS FORTS
─────────────────────────────────────────────
"""
            
            if n4 > 0:
                summary += f"• {n4} poste(s) maîtrisé(s) au niveau Référent (N4)\n"
            if n3 > 0:
                summary += f"• {n3} poste(s) maîtrisé(s) au niveau Expert (N3)\n"
            
            if n3 + n4 >= 3:
                summary += "• Opérateur polyvalent (3+ postes de niveau 3/4)\n"
            
            summary += """
 REMARQUES
─────────────────────────────────────────────
L'opérateur est actuellement INACTIF.
Toutes ses compétences et évaluations sont conservées
dans le système pour référence future.

Vous pouvez réactiver cet opérateur à tout moment
en cliquant sur le bouton "Réactiver l'opérateur".
"""
            
            self.summary_text.setPlainText(summary)
            
        except Exception as e:
            self.summary_text.setPlainText(f"Erreur lors du chargement du résumé :\n{e}")
    
    def load_history(self):
        """Charge l'historique des modifications."""
        try:
            connection = get_db_connection()
            cursor, dict_mode = _cursor(connection)
            
            cursor.execute(
                """SELECT date_time, action, description 
                   FROM historique 
                   WHERE operateur_id = %s 
                   ORDER BY date_time DESC 
                   LIMIT 50""",
                (self.operateur_id,)
            )
            rows = _rows(cursor, dict_mode)
            
            cursor.close()
            connection.close()
            
            self.history_table.setRowCount(0)
            
            for r in rows:
                row = self.history_table.rowCount()
                self.history_table.insertRow(row)
                
                date_str = self._format_datetime(r.get("date_time"))
                action = r.get("action", "")
                desc = r.get("description", "")
                
                self.history_table.setItem(row, 0, QTableWidgetItem(date_str))
                self.history_table.setItem(row, 1, QTableWidgetItem(action))
                self.history_table.setItem(row, 2, QTableWidgetItem(desc))
            
        except Exception as e:
            QMessageBox.warning(self, "Historique", f"Impossible de charger l'historique :\n{e}")
    
    def reactivate_operateur(self):
        """Réactive l'opérateur (passe le statut à ACTIF)."""
        reply = QMessageBox.question(
            self, "Confirmer la réactivation",
            "Êtes-vous sûr de vouloir réactiver cet opérateur ?\n\n"
            "Son statut passera à ACTIF et il réapparaîtra dans les listes principales.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        try:
            connection = get_db_connection()
            cursor, _ = _cursor(connection)
            
            cursor.execute(
                "UPDATE operateurs SET statut = 'ACTIF' WHERE id = %s",
                (self.operateur_id,)
            )
            
            connection.commit()
            cursor.close()
            connection.close()
            
            QMessageBox.information(
                self, "Réactivation réussie",
                " L'opérateur a été réactivé avec succès !\n\n"
                "Il apparaîtra de nouveau dans les listes d'opérateurs actifs."
            )
            
            self.operateur_reactivated.emit(self.operateur_id)
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de réactiver l'opérateur :\n{e}")
    
    def export_profile(self):
        """Exporte le profil complet en texte."""
        from PyQt5.QtWidgets import QFileDialog
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Exporter le profil",
            f"profil_operateur_{self.operateur_id}.txt",
            "Text Files (*.txt)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(self.summary_text.toPlainText())
                f.write("\n\n")
                f.write("═" * 60 + "\n")
                f.write("DÉTAILS DES POLYVALENCES\n")
                f.write("═" * 60 + "\n\n")
                
                for row in range(self.poly_table.rowCount()):
                    poste = self.poly_table.item(row, 0).text()
                    niveau = self.poly_table.item(row, 1).text()
                    date_eval = self.poly_table.item(row, 2).text()
                    date_next = self.poly_table.item(row, 3).text()
                    anciennete = self.poly_table.item(row, 4).text()
                    
                    f.write(f"• {poste} - Niveau {niveau}\n")
                    f.write(f"  Évaluation : {date_eval}\n")
                    f.write(f"  Prochaine  : {date_next}\n")
                    f.write(f"  Ancienneté : {anciennete}\n\n")
            
            QMessageBox.information(self, "Export réussi", f" Profil exporté :\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible d'exporter le profil :\n{e}")
    
    def _format_date(self, date_val):
        """Formate une date."""
        if date_val is None:
            return "N/A"
        if isinstance(date_val, str):
            try:
                return dt.datetime.strptime(date_val, "%Y-%m-%d").strftime("%d/%m/%Y")
            except:
                return date_val
        if hasattr(date_val, "strftime"):
            return date_val.strftime("%d/%m/%Y")
        return str(date_val)
    
    def _format_datetime(self, datetime_val):
        """Formate un datetime."""
        if datetime_val is None:
            return "N/A"
        if isinstance(datetime_val, str):
            try:
                return dt.datetime.fromisoformat(datetime_val).strftime("%d/%m/%Y %H:%M")
            except:
                return datetime_val
        if hasattr(datetime_val, "strftime"):
            return datetime_val.strftime("%d/%m/%Y %H:%M")
        return str(datetime_val)
    
    def _calculate_anciennete(self, date_eval):
        """Calcule l'ancienneté."""
        if date_eval is None:
            return "N/A"
        try:
            if isinstance(date_eval, str):
                date_obj = dt.datetime.strptime(date_eval, "%Y-%m-%d").date()
            elif hasattr(date_eval, "date"):
                date_obj = date_eval.date()
            else:
                date_obj = date_eval
            
            delta = dt.date.today() - date_obj
            years = delta.days // 365
            months = (delta.days % 365) // 30
            
            if years > 0:
                return f"{years} an(s) {months} mois"
            elif months > 0:
                return f"{months} mois"
            else:
                return f"{delta.days} jour(s)"
        except:
            return "N/A"


class OperateursInactifsDialog(QDialog):
    """
    Fenêtre principale de gestion des opérateurs inactifs.
    Vue synthétique avec possibilité d'ouvrir les détails complets.
    """
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Gestion des Opérateurs Inactifs")
        self.setGeometry(100, 100, 1000, 600)
        
        layout = QVBoxLayout(self)
        
        # === Header ===
        header = QLabel(" Opérateurs Inactifs")
        header.setFont(QFont("Arial", 18, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        subtitle = QLabel("Liste complète des opérateurs avec statut INACTIF")
        subtitle.setStyleSheet("color: #6b7280; font-style: italic;")
        subtitle.setAlignment(Qt.AlignCenter)
        layout.addWidget(subtitle)
        
        # === Filtres ===
        filters = QHBoxLayout()
        
        filters.addWidget(QLabel(" Recherche :"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Nom ou prénom...")
        self.search_input.textChanged.connect(self.filter_table)
        filters.addWidget(self.search_input, 1)
        
        self.show_stats_check = QCheckBox("Afficher les statistiques")
        self.show_stats_check.setChecked(True)
        self.show_stats_check.toggled.connect(self.toggle_stats_columns)
        filters.addWidget(self.show_stats_check)
        
        self.refresh_btn = QPushButton(" Actualiser")
        self.refresh_btn.clicked.connect(self.load_data)
        filters.addWidget(self.refresh_btn)
        
        layout.addLayout(filters)
        
        # === Table principale ===
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "Nom", "Prénom", "Nb Postes", "Niveau 1", "Niveau 2", "Niveau 3", "Niveau 4"
        ])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.doubleClicked.connect(self.open_detail_dialog)
        layout.addWidget(self.table, 1)
        
        # === Stats globales ===
        stats_label = QLabel("Double-cliquez sur une ligne pour voir les détails complets")
        stats_label.setStyleSheet("color: #6b7280; font-style: italic; padding: 8px;")
        stats_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(stats_label)
        
        self.total_label = QLabel("Total : 0 opérateur(s)")
        self.total_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.total_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.total_label)
        
        # === Actions ===
        actions = QHBoxLayout()
        actions.addStretch()
        
        self.export_btn = QPushButton(" Exporter la liste")
        self.export_btn.clicked.connect(self.export_list)
        actions.addWidget(self.export_btn)
        
        self.close_btn = QPushButton("Fermer")
        self.close_btn.clicked.connect(self.close)
        actions.addWidget(self.close_btn)
        
        layout.addLayout(actions)
        
        # Charger les données
        self.all_data = []
        self.load_data()
    
    def load_data(self):
        """Charge tous les opérateurs inactifs avec leurs stats."""
        try:
            connection = get_db_connection()
            cursor, dict_mode = _cursor(connection)
            
            query = """
                SELECT 
                    o.id,
                    o.nom,
                    o.prenom,
                    COUNT(p.id) as nb_postes,
                    SUM(CASE WHEN p.niveau = 1 THEN 1 ELSE 0 END) as n1,
                    SUM(CASE WHEN p.niveau = 2 THEN 1 ELSE 0 END) as n2,
                    SUM(CASE WHEN p.niveau = 3 THEN 1 ELSE 0 END) as n3,
                    SUM(CASE WHEN p.niveau = 4 THEN 1 ELSE 0 END) as n4
                FROM operateurs o
                LEFT JOIN polyvalence p ON o.id = p.operateur_id
                WHERE UPPER(o.statut) = 'INACTIF'
                GROUP BY o.id, o.nom, o.prenom
                ORDER BY o.nom, o.prenom
            """
            
            cursor.execute(query)
            self.all_data = _rows(cursor, dict_mode)
            
            cursor.close()
            connection.close()
            
            self.filter_table()
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible de charger les données :\n{e}")
    
    def filter_table(self):
        """Filtre et affiche les données dans la table."""
        search_text = self.search_input.text().lower()
        
        self.table.setRowCount(0)
        
        count = 0
        for data in self.all_data:
            nom = data.get("nom", "").lower()
            prenom = data.get("prenom", "").lower()
            
            if search_text and search_text not in nom and search_text not in prenom:
                continue
            
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            # Stocker l'ID en UserRole
            nom_item = QTableWidgetItem(data.get("nom", ""))
            nom_item.setData(Qt.UserRole, data.get("id"))
            self.table.setItem(row, 0, nom_item)
            
            self.table.setItem(row, 1, QTableWidgetItem(data.get("prenom", "")))
            self.table.setItem(row, 2, QTableWidgetItem(str(data.get("nb_postes", 0))))
            
            # Colonnes stats avec couleurs
            n1_item = QTableWidgetItem(str(data.get("n1", 0)))
            n1_item.setBackground(QColor("#fef2f2"))
            n1_item.setForeground(QColor("#dc2626"))
            n1_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 3, n1_item)
            
            n2_item = QTableWidgetItem(str(data.get("n2", 0)))
            n2_item.setBackground(QColor("#fffbeb"))
            n2_item.setForeground(QColor("#d97706"))
            n2_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 4, n2_item)
            
            n3_item = QTableWidgetItem(str(data.get("n3", 0)))
            n3_item.setBackground(QColor("#f0fdf4"))
            n3_item.setForeground(QColor("#059669"))
            n3_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 5, n3_item)
            
            n4_item = QTableWidgetItem(str(data.get("n4", 0)))
            n4_item.setBackground(QColor("#eff6ff"))
            n4_item.setForeground(QColor("#2563eb"))
            n4_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 6, n4_item)
            
            count += 1
        
        self.total_label.setText(f"Total : {count} opérateur(s) inactif(s)")
    
    def toggle_stats_columns(self, checked):
        """Affiche/masque les colonnes de statistiques."""
        for col in range(3, 7):  # Colonnes N1, N2, N3, N4
            self.table.setColumnHidden(col, not checked)
    
    def open_detail_dialog(self):
        """Ouvre la fenêtre de détails pour l'opérateur sélectionné."""
        selected = self.table.selectedItems()
        if not selected:
            return
        
        operateur_id = selected[0].data(Qt.UserRole)
        nom = selected[0].text()
        prenom = selected[1].text()
        
        detail_dialog = DetailOperateurDialog(operateur_id, nom, prenom, self)
        detail_dialog.operateur_reactivated.connect(self.on_operateur_reactivated)
        detail_dialog.exec_()
    
    def on_operateur_reactivated(self, operateur_id):
        """Callback quand un opérateur est réactivé."""
        QMessageBox.information(
            self, "Opérateur réactivé",
            "L'opérateur a été réactivé avec succès.\n"
            "La liste va être actualisée."
        )
        self.load_data()
    
    def export_list(self):
        """Exporte la liste en Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from PyQt5.QtWidgets import QFileDialog
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Exporter la liste",
                f"operateurs_inactifs_{dt.date.today().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Opérateurs Inactifs"
            
            # En-têtes
            headers = ["Nom", "Prénom", "Nb Postes", "Niveau 1", "Niveau 2", "Niveau 3", "Niveau 4"]
            ws.append(headers)
            
            # Style en-têtes
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Données
            for data in self.all_data:
                ws.append([
                    data.get("nom", ""),
                    data.get("prenom", ""),
                    data.get("nb_postes", 0),
                    data.get("n1", 0),
                    data.get("n2", 0),
                    data.get("n3", 0),
                    data.get("n4", 0)
                ])
            
            # Ajuster les largeurs
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
            QMessageBox.information(
                self, "Export réussi",
                f" Liste exportée avec succès !\n\n{file_path}"
            )
            
        except ImportError:
            QMessageBox.warning(
                self, "Module manquant",
                "Le module 'openpyxl' est requis pour l'export Excel.\n\n"
                "Installez-le avec : pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Impossible d'exporter :\n{e}")


# Test autonome
if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication
    
    app = QApplication(sys.argv)
    dialog = OperateursInactifsDialog()
    dialog.show()
    sys.exit(app.exec_())